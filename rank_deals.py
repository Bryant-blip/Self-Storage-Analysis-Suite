"""
rank_deals.py

Scores and ranks every deal in the reports/ folder.

Deal Score is the canonical score from db_utils.recalculate_scores
(YoC 50% | Population 35% | Land Cost Efficiency 15%, min-max normalized
across processed deals.db rows, hard-gated on YoC >= 10% and Pop >= 30k).
This script looks each report up in deals.db by report_path and uses the
stored deal_score/population_3mi so results match the dashboard and
rank_reports.py. Reports with no matching deals.db row (or a gated-out
score of NULL) are still listed, sorted after scored deals by YoC, with
a blank Deal Score.

The YoC Score / Pop Score / LCE Score columns are this script's own
informational per-metric breakdown (0-100, absolute thresholds — see
score_yoc/score_population/score_lce below); they do not feed into the
Deal Score and may not match db_utils' relative normalization.

Output: reports/deal_rankings.xlsx
Usage:  python rank_deals.py
"""

import glob
import math
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing openpyxl — run: pip install openpyxl")
    sys.exit(1)

from crexi import census_pop as census_module
from db_utils import get_db, recalculate_scores

REPORTS      = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
OUTPUT_PATH  = os.path.join(REPORTS, "deal_rankings.xlsx")
CENSUS_KEY   = os.environ.get("CENSUS_API_KEY", "")
MIN_POP      = 30_000
MAX_POP      = 200_000
YOC_MAX      = 0.12    # 12% YoC → score 100
LCE_BEST     = 0.10    # land ≤10% of total cost → score 100
LCE_WORST    = 0.50    # land ≥50% of total cost → score 0


# ── Proforma calculations ─────────────────────────────────────────────────────

def calc_proforma(ws) -> dict | None:
    """
    Read raw assumption cells and derive all key metrics.
    Returns None if required cells are missing.
    """
    try:
        acres         = float(ws["C5"].value)
        yield_pct     = float(ws["E5"].value)
        rent_psf      = float(ws["E6"].value)
        occupancy     = float(ws["E7"].value)
        expense_ratio = float(ws["E8"].value)
        cap_rate      = float(ws["E9"].value)
        cost_per_sqft = float(ws["E10"].value)
    except (TypeError, ValueError):
        return None

    # C6 (land cost) may be None for older reports that missed the price backfill
    land_cost_raw = ws["C6"].value
    land_cost     = float(land_cost_raw) if land_cost_raw is not None else None

    net_rentable    = acres * 43_560 * yield_pct
    monthly_gross   = net_rentable * rent_psf * occupancy
    monthly_noi     = monthly_gross * (1 - expense_ratio)
    annual_noi      = monthly_noi * 12
    construction    = cost_per_sqft * net_rentable
    if land_cost is not None:
        total_cost  = construction + land_cost
        land_pct    = land_cost / total_cost if total_cost else 0
    else:
        total_cost  = construction
        land_pct    = None
    valuation       = annual_noi / cap_rate if cap_rate else 0
    equity_value    = valuation - total_cost
    yoc             = annual_noi / total_cost if total_cost else 0

    return {
        "acres":         acres,
        "land_cost":     land_cost,
        "rent_psf":      rent_psf,
        "net_rentable":  net_rentable,
        "annual_noi":    annual_noi,
        "total_cost":    total_cost,
        "valuation":     valuation,
        "equity_value":  equity_value,
        "yoc":           yoc,
        "land_pct":      land_pct,
    }


# ── Scoring functions (each returns 0–100) ────────────────────────────────────

def score_yoc(yoc: float) -> float:
    return min(100.0, max(0.0, (yoc / YOC_MAX) * 100))


def score_population(pop: int) -> float:
    if pop <= 0:
        return 0.0
    if pop <= MIN_POP:
        return 0.0
    if pop >= MAX_POP:
        return 100.0
    # Log scale: feels more intuitive for population spread
    log_min = math.log(MIN_POP)
    log_max = math.log(MAX_POP)
    return min(100.0, max(0.0, (math.log(pop) - log_min) / (log_max - log_min) * 100))


def score_lce(land_pct) -> float:
    """Land cost efficiency — lower land % = better score. None → neutral 50."""
    if land_pct is None:
        return 50.0
    if land_pct >= LCE_WORST:
        return 0.0
    if land_pct <= LCE_BEST:
        return 100.0
    return max(0.0, (LCE_WORST - land_pct) / (LCE_WORST - LCE_BEST) * 100)


def get_population_3mi(address: str) -> int:
    """
    Fallback population lookup for reports with no matching deals.db row
    (see collect_deals — the DB's stored population_3mi is preferred).

    Resolves the subject city/place from the address (state name lookup,
    same as census_pop's own address parsing), then delegates to
    check_population_gate for the subject-city-centroid-based 3-mile pool.
    Returns 0 if the city can't be resolved.
    """
    city, state = census_module.parse_city_state_from_address(address)
    if not city or not state:
        return 0

    place_key = census_module.get_place_name_lookup().get((city.lower(), state))
    if not place_key:
        return 0

    info = census_module.load_place_centroids().get(place_key)
    if not info:
        return 0

    result = census_module.check_population_gate(
        info["lat"], info["lng"], address, CENSUS_KEY,
    )
    return result["population_3mi"]


# ── Market name from folder path ──────────────────────────────────────────────

_STATE_ABBREV = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
    "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
    "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
    "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
    "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
    "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
    "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
    "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
    "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
    "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
    "WI": "Wisconsin", "WY": "Wyoming",
}

import re as _re

def market_from_address(address: str) -> str:
    """Extract state name from address string, e.g. 'Durham, NC 27703' -> 'North Carolina'."""
    m = _re.search(r",\s*([A-Z]{2})\s+\d{5}", address)
    if m:
        return _STATE_ABBREV.get(m.group(1), m.group(1))
    return "Unknown"


# ── Main ──────────────────────────────────────────────────────────────────────

def collect_deals() -> list[dict]:
    paths = sorted(
        p for p in glob.glob(os.path.join(REPORTS, "**", "*.xlsx"), recursive=True)
        if os.path.basename(p) != "deal_rankings.xlsx"
    )

    # Pull canonical population_3mi / deal_score from deals.db, keyed by
    # report_path, so this spreadsheet matches the dashboard and
    # rank_reports.py. Recalculate first in case scores are stale/NULL.
    # If deals.db is missing or unreadable, fall back to an empty lookup —
    # every report is then treated as "not in DB" (see per-report handling
    # below), so the script still runs.
    db_rows = {}
    try:
        conn = get_db()
        try:
            recalculate_scores(conn)
            for r in conn.execute(
                "SELECT report_path, population_3mi, deal_score FROM deals "
                "WHERE report_path IS NOT NULL"
            ).fetchall():
                key = os.path.normcase(os.path.abspath(r["report_path"]))
                db_rows[key] = r
        finally:
            conn.close()
    except Exception as exc:
        print(f"  WARNING: could not read deals.db ({exc}) — "
              f"population/deal_score will use fallback lookups\n")

    deals = []
    for path in paths:
        rel = os.path.relpath(path, os.path.dirname(os.path.abspath(__file__)))

        try:
            wb = openpyxl.load_workbook(path)
        except Exception as exc:
            print(f"  SKIP (can't open): {rel} — {exc}")
            continue

        proforma = next(
            (wb[n] for n in wb.sheetnames
             if n.strip().lower() in {"proforma", "initial look proforma", "initial proforma"}),
            None,
        )
        if proforma is None:
            print(f"  SKIP (no proforma tab): {rel}")
            continue

        address = str(proforma["B3"].value or "").strip()
        metrics = calc_proforma(proforma)
        if metrics is None:
            print(f"  SKIP (missing assumption cells): {rel}")
            continue

        print(f"  Scoring: {address[:55]}")

        db_row = db_rows.get(os.path.normcase(os.path.abspath(path)))
        if db_row is not None and db_row["population_3mi"] is not None:
            population = db_row["population_3mi"]
        else:
            population = get_population_3mi(address)
        total = db_row["deal_score"] if db_row is not None else None

        yoc_s = score_yoc(metrics["yoc"])
        pop_s = score_population(population)
        lce_s = score_lce(metrics["land_pct"])

        deals.append({
            "path":         path,
            "rel_path":     rel,
            "address":      address,
            "market":       market_from_address(address),
            "acres":        metrics["acres"],
            "land_cost":    metrics["land_cost"],
            "rent_psf":     metrics["rent_psf"],
            "annual_noi":   metrics["annual_noi"],
            "total_cost":   metrics["total_cost"],
            "equity_value": metrics["equity_value"],
            "yoc":          metrics["yoc"],
            "land_pct":     metrics["land_pct"],
            "population":   population,
            "yoc_score":    yoc_s,
            "pop_score":    pop_s,
            "lce_score":    lce_s,
            "deal_score":   total,
        })

    # Scored deals first (by deal_score desc); deals with no DB score
    # (no matching deals.db row, or gated out — deal_score NULL) sort
    # after, by YoC desc.
    deals.sort(key=lambda d: (
        d["deal_score"] is not None,
        d["deal_score"] if d["deal_score"] is not None else d["yoc"],
    ), reverse=True)
    for i, d in enumerate(deals, 1):
        d["rank"] = i

    return deals


def write_rankings(deals: list[dict]) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deal Rankings"

    # ── Styles ───────────────────────────────────────────────────────────────
    HDR_FILL  = PatternFill("solid", fgColor="1F3864")   # dark navy
    HDR_FONT  = Font(bold=True, color="FFFFFF", size=10)
    GOLD_FILL = PatternFill("solid", fgColor="FFD700")
    SILV_FILL = PatternFill("solid", fgColor="C0C0C0")
    BRON_FILL = PatternFill("solid", fgColor="CD7F32")
    ALT_FILL  = PatternFill("solid", fgColor="F2F2F2")
    BORDER    = Border(bottom=Side(style="thin", color="D0D0D0"))

    def fmt_pct(v):    return f"{v:.1%}"
    def fmt_dol(v):    return f"${v:,.0f}"
    def fmt_score(v):  return f"{v:.1f}"

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:Q1")
    title = ws["A1"]
    title.value = "Deal Rankings — Deal Score is the canonical dashboard score (YoC 50% | Population 35% | Land Cost Efficiency 15%)"
    title.font  = Font(bold=True, size=13, color="1F3864")
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:Q2")
    sub = ws["A2"]
    sub.value = (
        f"Deal Score sourced from deals.db (blank = no matching report_path row, or hard-gated out: "
        f"YoC < 10% or Pop < 30k)  |  YoC/Pop/LCE Score columns are informational only "
        f"(YoC 0–12%+, Pop 30k–200k+ log scale, Land Cost % 50%→0/10%→100) and do not feed Deal Score"
    )
    sub.font      = Font(italic=True, size=9, color="666666")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    # ── Headers ───────────────────────────────────────────────────────────────
    headers = [
        ("Rank",           5),
        ("Address",        38),
        ("Market",         14),
        ("Deal Score",      9),
        ("YoC",             7),
        ("YoC Score",       8),
        ("Pop (3 mi)",      11),
        ("Pop Score",       8),
        ("Land Cost %",    10),
        ("LCE Score",       8),
        ("Acres",           6),
        ("Land Cost",      11),
        ("Rent/sqft",       9),
        ("Annual NOI",     11),
        ("Total Cost",     11),
        ("Equity Value",   11),
        ("Report",          9),
    ]

    HDR_ROW = 3
    for col, (label, width) in enumerate(headers, 1):
        cell = ws.cell(row=HDR_ROW, column=col, value=label)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[HDR_ROW].height = 28

    # ── Data rows ─────────────────────────────────────────────────────────────
    for i, d in enumerate(deals):
        row = HDR_ROW + 1 + i
        fill = (GOLD_FILL if d["rank"] == 1
                else SILV_FILL if d["rank"] == 2
                else BRON_FILL if d["rank"] == 3
                else ALT_FILL if i % 2 == 1 else None)

        def cell(col, value, num_fmt=None, hyperlink=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = Font(bold=(d["rank"] <= 3), size=10)
            c.alignment = Alignment(vertical="center",
                                    horizontal="center" if col != 2 else "left")
            if fill:
                c.fill = fill
            if num_fmt:
                c.number_format = num_fmt
            if hyperlink:
                c.hyperlink = hyperlink
                c.style     = "Hyperlink"
            c.border = BORDER
            return c

        cell(1,  d["rank"])
        cell(2,  d["address"])
        cell(3,  d["market"])
        cell(4,  round(d["deal_score"], 1) if d["deal_score"] is not None else None)
        cell(5,  d["yoc"]).number_format = '0.0%'
        cell(6,  round(d["yoc_score"],   1))
        cell(7,  d["population"],  '#,##0')
        cell(8,  round(d["pop_score"],   1))
        cell(9,  d["land_pct"]).number_format = '0.0%'
        cell(10, round(d["lce_score"],   1))
        cell(11, round(d["acres"], 2),   '0.00')
        cell(12, d["land_cost"],         '"$"#,##0')
        cell(13, d["rent_psf"],          '"$"0.00')
        cell(14, d["annual_noi"],        '"$"#,##0')
        cell(15, d["total_cost"],        '"$"#,##0')
        cell(16, d["equity_value"],      '"$"#,##0')
        # Report hyperlink — relative path as file:// link
        abs_path = os.path.abspath(d["path"])
        file_url = "file:///" + abs_path.replace("\\", "/")
        cell(18, "Open", hyperlink=file_url)

        ws.row_dimensions[row].height = 18

    # ── Freeze panes, auto-filter ──────────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=HDR_ROW + 1, column=1)
    ws.auto_filter.ref = f"A{HDR_ROW}:{get_column_letter(len(headers))}{HDR_ROW}"

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    try:
        wb.save(OUTPUT_PATH)
        print(f"\nSaved: {os.path.relpath(OUTPUT_PATH, os.path.dirname(os.path.abspath(__file__)))}")
    except PermissionError:
        print(f"\nERROR: Could not save deal_rankings.xlsx — close it in Excel first, then re-run.")


def main():
    from dotenv import load_dotenv
    load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

    print(f"Scoring {len(glob.glob(os.path.join(REPORTS, '**/*.xlsx'), recursive=True))-1} deals...\n")
    deals = collect_deals()
    print(f"\n{'-'*60}")
    print(f"{'Rank':<5} {'Score':>6}  {'YoC':>6}  {'Pop (3mi)':>11}  Address")
    print(f"{'-'*60}")
    for d in deals[:10]:
        score_str = f"{d['deal_score']:>5.1f}" if d["deal_score"] is not None else "  N/A"
        print(f"  #{d['rank']:<3} {score_str}  {d['yoc']:>5.1%}  "
              f"{d['population']:>11,}  {d['address'][:45]}")
    if len(deals) > 10:
        print(f"  ... and {len(deals)-10} more")

    write_rankings(deals)
    print(f"\nTop 3 deals:")
    for d in deals[:3]:
        print(f"  #{d['rank']} {d['address']}")
        land_str = f"{d['land_pct']:.0%} of cost" if d["land_pct"] is not None else "land cost unknown"
        score_str = f"{d['deal_score']:.1f}" if d["deal_score"] is not None else "N/A"
        print(f"     Score {score_str}  |  YoC {d['yoc']:.1%}  |  "
              f"Pop {d['population']:,}  |  Land {land_str}")


if __name__ == "__main__":
    main()
