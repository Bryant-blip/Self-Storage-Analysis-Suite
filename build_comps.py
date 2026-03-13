import math
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
SUBJECT_LAT   = 40.397
SUBJECT_LON   = -111.830
SUBJECT_ADDR  = "740 E Utah Highland Dr, Lehi, UT 84043"
DATE_PULLED   = "03/10/2026"
SEARCH_RADIUS = 4.0

# Output path (Windows format via os.path.join relative to this file)
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR  = os.path.join(BASE_DIR, "output")
OUTPUT_PATH = os.path.join(OUTPUT_DIR,
    "storage_comps_740_E_Utah_Highland_Dr_Lehi_UT_84043_20260310.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def haversine(lat1, lon1, lat2, lon2):
    R = 3958.8
    phi1, phi2 = math.radians(lat1), math.radians(lat2)
    dphi = math.radians(lat2 - lat1)
    dlam = math.radians(lon2 - lon1)
    a = math.sin(dphi/2)**2 + math.cos(phi1)*math.cos(phi2)*math.sin(dlam/2)**2
    return round(2 * R * math.asin(math.sqrt(a)), 2)

def drive_min(miles):
    return round(miles / 25 * 60)

# ─────────────────────────────────────────────────────────────────────────────
# FACILITY MASTER LIST
# Coordinates estimated from street address geocoding (straight-line distances)
# ─────────────────────────────────────────────────────────────────────────────
FACILITIES = [
    {"id":1, "name":"Stock-N-Lock Self Storage",
     "address":"580 S 850 E, Lehi, UT 84043",
     "lat":40.383, "lon":-111.840,
     "phone":"(801) 766-1142",
     "website":"https://www.lehiutahselfstorage.com"},
    {"id":2, "name":"Public Storage (708 W Main St)",
     "address":"708 W Main St, Lehi, UT 84043",
     "lat":40.389, "lon":-111.862,
     "phone":"(844) 726-4531",
     "website":"https://www.publicstorage.com/self-storage-ut-lehi/5417.html"},
    {"id":3, "name":"AAA Self Storage",
     "address":"874 W 1500 N, Lehi, UT 84043",
     "lat":40.408, "lon":-111.863,
     "phone":"(801) 272-8192",
     "website":"https://aaalehi.com"},
    {"id":4, "name":"Prime Storage - Lehi (Pointe Meadow)",
     "address":"1985 N Pointe Meadow Dr, Lehi, UT 84043",
     "lat":40.420, "lon":-111.852,
     "phone":"(801) 609-3040",
     "website":"https://primestorage.com/locations/ut/lehi/UT09/"},
    {"id":5, "name":"AF Storage",
     "address":"320 S 100 W, American Fork, UT 84003",
     "lat":40.371, "lon":-111.799,
     "phone":"(801) 756-7766",
     "website":"https://www.afstorage.com"},
    {"id":6, "name":"Harbor Storage",
     "address":"72 E Frontage Rd, American Fork, UT 84003",
     "lat":40.372, "lon":-111.793,
     "phone":"(801) 756-2928",
     "website":"https://www.afharborstorage.com"},
    {"id":7, "name":"1st Lehi Storage",
     "address":"2100 N 960 W, Lehi, UT 84043",
     "lat":40.421, "lon":-111.869,
     "phone":"(801) 768-3595",
     "website":"https://www.1st-lehi-storage.com"},
    {"id":8, "name":"Lehi Indoor Self Storage",
     "address":"1099 S 1100 W, Lehi, UT 84043",
     "lat":40.375, "lon":-111.872,
     "phone":"(801) 901-2298",
     "website":"https://www.lehiindoorstorage.com"},
    {"id":9, "name":"Prime Storage - American Fork",
     "address":"420 E 620 S, American Fork, UT 84003",
     "lat":40.367, "lon":-111.788,
     "phone":"(888) 846-6503",
     "website":"https://primestorage.com/locations/ut/american-fork/UT10/"},
    {"id":10, "name":"Fort Knox Storage of Lehi",
     "address":"2400 N 1200 W, Lehi, UT 84043",
     "lat":40.426, "lon":-111.874,
     "phone":"(801) 766-4777",
     "website":"https://www.lehiselfstorage.com"},
    {"id":11, "name":"Extra Space Storage - Cedar Hills",
     "address":"9978 N 4700 W, Cedar Hills, UT 84003",
     "lat":40.420, "lon":-111.765,
     "phone":"(801) 999-1515",
     "website":"https://www.extraspace.com/storage/facilities/us/utah/cedar_hills/7534/"},
    {"id":12, "name":"Extra Space Storage - Thanksgiving Way",
     "address":"4285 N Thanksgiving Way, Lehi, UT 84043",
     "lat":40.432, "lon":-111.891,
     "phone":"(801) 981-4615",
     "website":"https://www.extraspace.com/storage/facilities/us/utah/lehi/1000000376/"},
]

# ─────────────────────────────────────────────────────────────────────────────
# UNIT PRICING DATA
# All rates sourced directly from facility websites or verified aggregators.
# N/A = rate not published / not available online.
# ─────────────────────────────────────────────────────────────────────────────
PRICING = [
    # ── 1. Stock-N-Lock Self Storage ─────────────────────────────────────────
    # Source: lehiutahselfstorage.com (facility website, standard listed prices)
    # No separate online vs. walk-in rate published; 30% off promo via Storage.com
    {"fac_id":1,"unit_type":"5x5",    "sqft":25,    "cc":"No",  "online":50,  "instore":50,   "notes":"Drive-up. 30% off promo Jan-Mar 2026 (Source: Storage.com).","url":"https://www.lehiutahselfstorage.com/pages/rent"},
    {"fac_id":1,"unit_type":"5x10",   "sqft":50,    "cc":"No",  "online":75,  "instore":75,   "notes":"Drive-up or interior. 30% off promo Jan-Mar 2026.","url":"https://www.lehiutahselfstorage.com/pages/rent"},
    {"fac_id":1,"unit_type":"10x10",  "sqft":100,   "cc":"No",  "online":110, "instore":110,  "notes":"Interior unit ($105-$115 range; mid used). 30% off promo Jan-Mar 2026.","url":"https://www.lehiutahselfstorage.com/pages/rent"},
    {"fac_id":1,"unit_type":"10x15",  "sqft":150,   "cc":"No",  "online":155, "instore":155,  "notes":"Drive-up. 30% off promo Jan-Mar 2026.","url":"https://www.lehiutahselfstorage.com/pages/rent"},
    {"fac_id":1,"unit_type":"10x20",  "sqft":200,   "cc":"No",  "online":185, "instore":185,  "notes":"Drive-up. 30% off promo Jan-Mar 2026.","url":"https://www.lehiutahselfstorage.com/pages/rent"},
    {"fac_id":1,"unit_type":"10x30",  "sqft":300,   "cc":"No",  "online":240, "instore":240,  "notes":"Drive-up ($235-$250 range; mid used). 30% off promo Jan-Mar 2026.","url":"https://www.lehiutahselfstorage.com/pages/rent"},
    # ── 2. Public Storage 708 W Main ─────────────────────────────────────────
    # Source: storage.com verified listing — confirmed web vs. standard rates
    {"fac_id":2,"unit_type":"5x5",    "sqft":25,    "cc":"Yes", "online":29,  "instore":39,   "notes":"50% off 1st month; $29 admin fee at move-in.","url":"https://www.storage.com/self-storage/utah/lehi/public-storage-lehi-708-w-main-st-218015/"},
    {"fac_id":2,"unit_type":"5x10",   "sqft":50,    "cc":"Yes", "online":68,  "instore":90,   "notes":"50% off 1st month; $29 admin fee.","url":"https://www.storage.com/self-storage/utah/lehi/public-storage-lehi-708-w-main-st-218015/"},
    {"fac_id":2,"unit_type":"10x10",  "sqft":100,   "cc":"Yes", "online":85,  "instore":113,  "notes":"Web-only special; $29 admin fee.","url":"https://www.storage.com/self-storage/utah/lehi/public-storage-lehi-708-w-main-st-218015/"},
    {"fac_id":2,"unit_type":"10x15",  "sqft":150,   "cc":"Yes", "online":104, "instore":138,  "notes":"$1 first month rent; $29 admin fee.","url":"https://www.storage.com/self-storage/utah/lehi/public-storage-lehi-708-w-main-st-218015/"},
    {"fac_id":2,"unit_type":"10x20",  "sqft":200,   "cc":"Yes", "online":128, "instore":170,  "notes":"2nd month free; $29 admin fee.","url":"https://www.storage.com/self-storage/utah/lehi/public-storage-lehi-708-w-main-st-218015/"},
    # ── 3. AAA Self Storage ───────────────────────────────────────────────────
    # Source: aaalehi.com — no online vs. walk-in distinction published
    {"fac_id":3,"unit_type":"5x10",   "sqft":50,    "cc":"No",  "online":75,  "instore":"N/A","notes":"Ground-level drive-up. No separate online rate published.","url":"https://aaalehi.com"},
    {"fac_id":3,"unit_type":"10x10",  "sqft":100,   "cc":"No",  "online":100, "instore":"N/A","notes":"Ground-level drive-up. No separate online rate published.","url":"https://aaalehi.com"},
    {"fac_id":3,"unit_type":"10x20",  "sqft":200,   "cc":"No",  "online":150, "instore":"N/A","notes":"Ground-level drive-up. No separate online rate published.","url":"https://aaalehi.com"},
    {"fac_id":3,"unit_type":"10x25",  "sqft":250,   "cc":"No",  "online":175, "instore":"N/A","notes":"Ground-level drive-up. No separate online rate published.","url":"https://aaalehi.com"},
    {"fac_id":3,"unit_type":"10x30",  "sqft":300,   "cc":"No",  "online":200, "instore":"N/A","notes":"Ground-level drive-up. No separate online rate published.","url":"https://aaalehi.com"},
    # ── 4. Prime Storage - Lehi (Pointe Meadow) ──────────────────────────────
    # Source: primestorage.com — online vs. standard confirmed; no CC at location
    {"fac_id":4,"unit_type":"5x5",    "sqft":25,    "cc":"No",  "online":31,  "instore":55,   "notes":"Inside/ground floor. First Month FREE on select sizes.","url":"https://primestorage.com/locations/ut/lehi/UT09/"},
    {"fac_id":4,"unit_type":"10x10",  "sqft":100,   "cc":"No",  "online":93,  "instore":139,  "notes":"Drive-up. First Month FREE promo.","url":"https://primestorage.com/locations/ut/lehi/UT09/"},
    {"fac_id":4,"unit_type":"10x15",  "sqft":150,   "cc":"No",  "online":104, "instore":188,  "notes":"Drive-up. First Month FREE promo.","url":"https://primestorage.com/locations/ut/lehi/UT09/"},
    {"fac_id":4,"unit_type":"10x20",  "sqft":200,   "cc":"No",  "online":164, "instore":224,  "notes":"Drive-up. First Month FREE promo.","url":"https://primestorage.com/locations/ut/lehi/UT09/"},
    {"fac_id":4,"unit_type":"10x25",  "sqft":250,   "cc":"No",  "online":252, "instore":269,  "notes":"Drive-up.","url":"https://primestorage.com/locations/ut/lehi/UT09/"},
    {"fac_id":4,"unit_type":"10x30",  "sqft":300,   "cc":"No",  "online":293, "instore":319,  "notes":"Drive-up.","url":"https://primestorage.com/locations/ut/lehi/UT09/"},
    {"fac_id":4,"unit_type":"Parking","sqft":"N/A", "cc":"No",  "online":412, "instore":420,  "notes":"30-ft outdoor parking space.","url":"https://primestorage.com/locations/ut/lehi/UT09/"},
    # ── 5. AF Storage ─────────────────────────────────────────────────────────
    # Source: afstorage.com — no separate online rate; both CC and drive-up available
    {"fac_id":5,"unit_type":"5x5",    "sqft":25,    "cc":"Yes", "online":48,  "instore":"N/A","notes":"Climate-controlled indoor. No online discount noted.","url":"https://www.afstorage.com/pages/storage-unit-prices"},
    {"fac_id":5,"unit_type":"5x10",   "sqft":50,    "cc":"Yes", "online":78,  "instore":"N/A","notes":"CC indoor $78; non-CC drive-up also avail at $68/mo.","url":"https://www.afstorage.com/pages/storage-unit-prices"},
    {"fac_id":5,"unit_type":"10x10",  "sqft":100,   "cc":"Yes", "online":118, "instore":"N/A","notes":"CC $118; non-CC drive-up $108/mo.","url":"https://www.afstorage.com/pages/storage-unit-prices"},
    {"fac_id":5,"unit_type":"10x15",  "sqft":150,   "cc":"Yes", "online":148, "instore":"N/A","notes":"CC $148; non-CC drive-up $138/mo.","url":"https://www.afstorage.com/pages/storage-unit-prices"},
    {"fac_id":5,"unit_type":"10x20",  "sqft":200,   "cc":"Yes", "online":178, "instore":"N/A","notes":"CC $178; non-CC drive-up $168/mo.","url":"https://www.afstorage.com/pages/storage-unit-prices"},
    {"fac_id":5,"unit_type":"10x25",  "sqft":250,   "cc":"No",  "online":188, "instore":"N/A","notes":"Drive-up only (non-CC).","url":"https://www.afstorage.com/pages/storage-unit-prices"},
    {"fac_id":5,"unit_type":"10x30",  "sqft":300,   "cc":"No",  "online":218, "instore":"N/A","notes":"Drive-up only (non-CC).","url":"https://www.afstorage.com/pages/storage-unit-prices"},
    # ── 6. Harbor Storage ─────────────────────────────────────────────────────
    # Source: afharborstorage.com — category pricing only (no specific unit sizes)
    {"fac_id":6,"unit_type":"5x10",   "sqft":50,    "cc":"N/A", "online":70,  "instore":"N/A","notes":"Listed as 'Small Unit' $70/mo. Buy 1 month get 1 month FREE promo.","url":"https://www.afharborstorage.com"},
    {"fac_id":6,"unit_type":"10x10",  "sqft":100,   "cc":"N/A", "online":117, "instore":"N/A","notes":"Listed as 'Medium Unit' $117/mo. Buy 1 month get 1 month FREE promo.","url":"https://www.afharborstorage.com"},
    {"fac_id":6,"unit_type":"10x20",  "sqft":200,   "cc":"N/A", "online":160, "instore":"N/A","notes":"Listed as 'Large Unit' $160/mo. Buy 1 month get 1 month FREE promo.","url":"https://www.afharborstorage.com"},
    # ── 7. 1st Lehi Storage ───────────────────────────────────────────────────
    # Source: 1st-lehi-storage.com — limited pricing (only 3 non-standard sizes)
    {"fac_id":7,"unit_type":"10x20",  "sqft":200,   "cc":"N/A", "online":300, "instore":"N/A","notes":"Drive-up. Limited pricing data on website.","url":"https://www.1st-lehi-storage.com"},
    {"fac_id":7,"unit_type":"10x30",  "sqft":300,   "cc":"N/A", "online":275, "instore":"N/A","notes":"Drive-up. Facility lists as 10x32 at $275. Limited data.","url":"https://www.1st-lehi-storage.com"},
    # ── 8. Lehi Indoor Self Storage ───────────────────────────────────────────
    # Source: lehiindoorstorage.com — all CC; prices vary by floor (shown = lowest)
    {"fac_id":8,"unit_type":"5x5",    "sqft":25,    "cc":"Yes", "online":67,  "instore":"N/A","notes":"All CC (3-story indoor facility). Lowest rate shown; range $67-$79. $25 admin fee.","url":"https://www.lehiindoorstorage.com/1099-s-1100-w-lehi-ut-84043"},
    {"fac_id":8,"unit_type":"5x10",   "sqft":50,    "cc":"Yes", "online":71,  "instore":"N/A","notes":"All CC. Lowest rate shown; range $71-$89. $25 admin fee.","url":"https://www.lehiindoorstorage.com/1099-s-1100-w-lehi-ut-84043"},
    {"fac_id":8,"unit_type":"10x10",  "sqft":100,   "cc":"Yes", "online":100, "instore":"N/A","notes":"All CC. Lowest rate shown; range $100-$139. $25 admin fee.","url":"https://www.lehiindoorstorage.com/1099-s-1100-w-lehi-ut-84043"},
    {"fac_id":8,"unit_type":"10x15",  "sqft":150,   "cc":"Yes", "online":160, "instore":"N/A","notes":"All CC. Lowest rate shown; range $160-$192. $25 admin fee.","url":"https://www.lehiindoorstorage.com/1099-s-1100-w-lehi-ut-84043"},
    {"fac_id":8,"unit_type":"10x20",  "sqft":200,   "cc":"Yes", "online":181, "instore":"N/A","notes":"All CC. Range $181-$189. $25 admin fee.","url":"https://www.lehiindoorstorage.com/1099-s-1100-w-lehi-ut-84043"},
    {"fac_id":8,"unit_type":"10x25",  "sqft":250,   "cc":"Yes", "online":229, "instore":"N/A","notes":"All CC. $25 admin fee.","url":"https://www.lehiindoorstorage.com/1099-s-1100-w-lehi-ut-84043"},
    {"fac_id":8,"unit_type":"10x30",  "sqft":300,   "cc":"Yes", "online":305, "instore":"N/A","notes":"All CC. Range $305-$318. $25 admin fee.","url":"https://www.lehiindoorstorage.com/1099-s-1100-w-lehi-ut-84043"},
    # ── 9. Prime Storage - American Fork ──────────────────────────────────────
    # Source: primestorage.com — online vs. standard confirmed; no CC at location
    {"fac_id":9,"unit_type":"5x5",    "sqft":25,    "cc":"No",  "online":37,  "instore":64,   "notes":"1st Month FREE promo.","url":"https://primestorage.com/locations/ut/american-fork/UT10/"},
    {"fac_id":9,"unit_type":"5x10",   "sqft":50,    "cc":"No",  "online":49,  "instore":84,   "notes":"1st Month FREE promo.","url":"https://primestorage.com/locations/ut/american-fork/UT10/"},
    {"fac_id":9,"unit_type":"10x10",  "sqft":100,   "cc":"No",  "online":96,  "instore":124,  "notes":"Standard drive-up.","url":"https://primestorage.com/locations/ut/american-fork/UT10/"},
    {"fac_id":9,"unit_type":"10x15",  "sqft":150,   "cc":"No",  "online":108, "instore":159,  "notes":"Standard drive-up.","url":"https://primestorage.com/locations/ut/american-fork/UT10/"},
    {"fac_id":9,"unit_type":"10x20",  "sqft":200,   "cc":"No",  "online":155, "instore":209,  "notes":"Limited inventory noted.","url":"https://primestorage.com/locations/ut/american-fork/UT10/"},
    {"fac_id":9,"unit_type":"10x30",  "sqft":300,   "cc":"No",  "online":237, "instore":325,  "notes":"Limited inventory noted.","url":"https://primestorage.com/locations/ut/american-fork/UT10/"},
    {"fac_id":9,"unit_type":"Parking","sqft":"N/A", "cc":"No",  "online":102, "instore":119,  "notes":"35-ft outdoor parking. 1st Month FREE promo.","url":"https://primestorage.com/locations/ut/american-fork/UT10/"},
    # ── 10. Fort Knox Storage of Lehi ─────────────────────────────────────────
    # Source: RentCafe / StorageCafe aggregated listings (official site 404)
    {"fac_id":10,"unit_type":"5x5",   "sqft":25,    "cc":"Yes", "online":57,  "instore":"N/A","notes":"CC unit. Rate from RentCafe/StorageCafe aggregated listing.","url":"https://www.lehiselfstorage.com"},
    {"fac_id":10,"unit_type":"5x10",  "sqft":50,    "cc":"Yes", "online":78,  "instore":"N/A","notes":"CC unit. Aggregated estimate ($75-$80 range).","url":"https://www.lehiselfstorage.com"},
    {"fac_id":10,"unit_type":"10x10", "sqft":100,   "cc":"Yes", "online":125, "instore":"N/A","notes":"CC unit. Rate from RentCafe aggregated listing.","url":"https://www.lehiselfstorage.com"},
    {"fac_id":10,"unit_type":"10x15", "sqft":150,   "cc":"Yes", "online":155, "instore":"N/A","notes":"CC unit. Aggregated estimate ($150-$170 range).","url":"https://www.lehiselfstorage.com"},
    {"fac_id":10,"unit_type":"10x20", "sqft":200,   "cc":"Yes", "online":170, "instore":"N/A","notes":"CC unit. Aggregated estimate; verify with facility directly.","url":"https://www.lehiselfstorage.com"},
    # ── 11. Extra Space Storage - Cedar Hills ─────────────────────────────────
    # Source: storage.com verified listing — web vs. standard rates confirmed
    {"fac_id":11,"unit_type":"5x5",   "sqft":25,    "cc":"Yes", "online":26,  "instore":28,   "notes":"CC, upstairs via elevator. Flash sale. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/highland/84003/9978-north-4700-west/"},
    {"fac_id":11,"unit_type":"5x10",  "sqft":50,    "cc":"Yes", "online":71,  "instore":75,   "notes":"CC, 1st floor. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/highland/84003/9978-north-4700-west/"},
    {"fac_id":11,"unit_type":"10x10", "sqft":100,   "cc":"Yes", "online":100, "instore":105,  "notes":"CC, upstairs via elevator. Non-CC also avail $92/$98. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/highland/84003/9978-north-4700-west/"},
    {"fac_id":11,"unit_type":"10x15", "sqft":150,   "cc":"Yes", "online":133, "instore":141,  "notes":"CC, upstairs via elevator. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/highland/84003/9978-north-4700-west/"},
    {"fac_id":11,"unit_type":"10x20", "sqft":200,   "cc":"Yes", "online":135, "instore":143,  "notes":"CC flash sale (upstairs). 1st-floor CC $214/$226. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/highland/84003/9978-north-4700-west/"},
    {"fac_id":11,"unit_type":"10x25", "sqft":250,   "cc":"Yes", "online":254, "instore":269,  "notes":"CC, 1st floor. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/highland/84003/9978-north-4700-west/"},
    {"fac_id":11,"unit_type":"10x30", "sqft":300,   "cc":"No",  "online":395, "instore":405,  "notes":"Drive-up, non-CC. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/highland/84003/9978-north-4700-west/"},
    # ── 12. Extra Space Storage - Thanksgiving Way ────────────────────────────
    # Source: storage.com verified listing — web vs. standard confirmed
    # NOTE: ~4.0 mi straight-line (borderline radius)
    {"fac_id":12,"unit_type":"5x5",   "sqft":25,    "cc":"Yes", "online":28,  "instore":30,   "notes":"CC, upstairs via lift. $29 admin fee. ~4.0 mi from subject (borderline radius).","url":"https://www.storage.com/self-storage/utah/lehi/extra-space-storage-8488-lehi-thanksgiving-way-232460/"},
    {"fac_id":12,"unit_type":"5x10",  "sqft":50,    "cc":"No",  "online":38,  "instore":40,   "notes":"Interior, 1st floor. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/lehi/extra-space-storage-8488-lehi-thanksgiving-way-232460/"},
    {"fac_id":12,"unit_type":"10x10", "sqft":100,   "cc":"Yes", "online":79,  "instore":83,   "notes":"CC, 1st floor interior. Non-CC also avail $66/$69. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/lehi/extra-space-storage-8488-lehi-thanksgiving-way-232460/"},
    {"fac_id":12,"unit_type":"10x15", "sqft":150,   "cc":"Yes", "online":107, "instore":114,  "notes":"CC, 1st floor interior. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/lehi/extra-space-storage-8488-lehi-thanksgiving-way-232460/"},
    {"fac_id":12,"unit_type":"10x20", "sqft":200,   "cc":"Yes", "online":156, "instore":166,  "notes":"CC, 1st floor interior. $29 admin fee.","url":"https://www.storage.com/self-storage/utah/lehi/extra-space-storage-8488-lehi-thanksgiving-way-232460/"},
]

# ─────────────────────────────────────────────────────────────────────────────
# COMPUTE DISTANCES & SORT
# ─────────────────────────────────────────────────────────────────────────────
fac_lookup = {f["id"]: f for f in FACILITIES}

for f in FACILITIES:
    d = haversine(SUBJECT_LAT, SUBJECT_LON, f["lat"], f["lon"])
    f["distance"]   = d
    f["drive_time"] = drive_min(d)

FACILITIES.sort(key=lambda x: x["distance"])

for p in PRICING:
    fac = fac_lookup[p["fac_id"]]
    p["distance"]   = fac["distance"]
    p["drive_time"] = fac["drive_time"]
    p["fac_name"]   = fac["name"]
    p["fac_addr"]   = fac["address"]

UNIT_ORDER = {"5x5":0,"5x10":1,"10x10":2,"10x15":3,"10x20":4,"10x25":5,"10x30":6,"Parking":7}

PRICING.sort(key=lambda p: (p["distance"], p["fac_name"], UNIT_ORDER.get(p["unit_type"], 99)))

# ─────────────────────────────────────────────────────────────────────────────
# STYLES
# ─────────────────────────────────────────────────────────────────────────────
FILL_BLUE   = PatternFill("solid", fgColor="BDD7EE")
FILL_GREEN  = PatternFill("solid", fgColor="E2EFDA")
FILL_ORANGE = PatternFill("solid", fgColor="FCE4D6")
FILL_ALTROW = PatternFill("solid", fgColor="F2F2F2")
BOLD        = Font(bold=True)
thin = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def style_header(cell, fill):
    cell.font      = BOLD
    cell.fill      = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = BORDER

def style_body(cell, ri):
    cell.border    = BORDER
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    if ri % 2 == 0:
        cell.fill = FILL_ALTROW

def auto_width(ws, min_w=8, max_w=55):
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        mx = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col_letter].width = max(min_w, min(mx + 2, max_w))

# ─────────────────────────────────────────────────────────────────────────────
# BUILD WORKBOOK
# ─────────────────────────────────────────────────────────────────────────────
wb = Workbook()

# ── TAB 1: Comps Detail ───────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Comps Detail"

CD_HDR = [
    "Facility Name", "Address", "Distance (mi)", "Drive Time (min)",
    "Unit Type", "Sq Ft", "Climate Controlled",
    "Online Rate ($/mo)", "In-Store Rate ($/mo)",
    "Notes / Promotions", "Date Pulled", "Source URL"
]

for ci, h in enumerate(CD_HDR, 1):
    style_header(ws1.cell(row=1, column=ci, value=h), FILL_BLUE)
ws1.row_dimensions[1].height = 32
ws1.freeze_panes = "A2"

for ri, p in enumerate(PRICING, 2):
    row = [
        p["fac_name"], p["fac_addr"], p["distance"], p["drive_time"],
        p["unit_type"], p["sqft"], p["cc"],
        p["online"], p["instore"],
        p["notes"], DATE_PULLED, p["url"]
    ]
    for ci, val in enumerate(row, 1):
        c = ws1.cell(row=ri, column=ci, value=val)
        style_body(c, ri)
        if ci == 3:
            c.number_format = "0.00"
        elif ci in (8, 9) and isinstance(val, (int, float)):
            c.number_format = "$#,##0"

auto_width(ws1)
ws1.column_dimensions["A"].width = 38
ws1.column_dimensions["B"].width = 36
ws1.column_dimensions["J"].width = 50
ws1.column_dimensions["L"].width = 62

# ── TAB 2: Market Summary ─────────────────────────────────────────────────────
ws2 = wb.create_sheet("Market Summary")

MS_HDR = [
    "Unit Type", "Sq Ft",
    "Avg Online Rate", "Min Online Rate", "Max Online Rate",
    "Avg In-Store Rate", "Min In-Store Rate", "Max In-Store Rate",
    "# of Comps"
]

for ci, h in enumerate(MS_HDR, 1):
    style_header(ws2.cell(row=1, column=ci, value=h), FILL_GREEN)
ws2.row_dimensions[1].height = 32
ws2.freeze_panes = "A2"

online_map  = defaultdict(list)
instore_map = defaultdict(list)
sqft_map    = {}

for p in PRICING:
    ut = p["unit_type"]
    sqft_map[ut] = p["sqft"]
    if isinstance(p["online"],  (int, float)): online_map[ut].append(p["online"])
    if isinstance(p["instore"], (int, float)): instore_map[ut].append(p["instore"])

UT_ORDER_LIST = ["5x5","5x10","10x10","10x15","10x20","10x25","10x30","Parking"]

def sagg(lst):
    return round(sum(lst)/len(lst), 2) if lst else "N/A"

for ri, ut in enumerate(UT_ORDER_LIST, 2):
    ol  = online_map.get(ut, [])
    ins = instore_map.get(ut, [])
    nc  = max(len(ol), len(ins))
    row = [
        ut, sqft_map.get(ut, "N/A"),
        sagg(ol),  (min(ol)  if ol  else "N/A"), (max(ol)  if ol  else "N/A"),
        sagg(ins), (min(ins) if ins else "N/A"), (max(ins) if ins else "N/A"),
        nc if nc > 0 else "N/A"
    ]
    for ci, val in enumerate(row, 1):
        c = ws2.cell(row=ri, column=ci, value=val)
        style_body(c, ri)
        if ci in (3, 4, 5, 6, 7, 8) and isinstance(val, (int, float)):
            c.number_format = "$#,##0.00"

auto_width(ws2)
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 10

# ── TAB 3: Facility List ──────────────────────────────────────────────────────
ws3 = wb.create_sheet("Facility List")

FL_HDR = ["Facility Name", "Address", "Distance (mi)", "Drive Time (min)", "Phone", "Website"]

for ci, h in enumerate(FL_HDR, 1):
    style_header(ws3.cell(row=1, column=ci, value=h), FILL_ORANGE)
ws3.row_dimensions[1].height = 32
ws3.freeze_panes = "A2"

for ri, f in enumerate(FACILITIES, 2):
    row = [f["name"], f["address"], f["distance"], f["drive_time"], f["phone"], f["website"]]
    for ci, val in enumerate(row, 1):
        c = ws3.cell(row=ri, column=ci, value=val)
        style_body(c, ri)
        if ci == 3:
            c.number_format = "0.00"

auto_width(ws3)
ws3.column_dimensions["A"].width = 40
ws3.column_dimensions["B"].width = 36
ws3.column_dimensions["F"].width = 64

# ─────────────────────────────────────────────────────────────────────────────
# SAVE
# ─────────────────────────────────────────────────────────────────────────────
os.makedirs(OUTPUT_DIR, exist_ok=True)
wb.save(OUTPUT_PATH)
print(f"\n  Saved: {OUTPUT_PATH}")

# ─────────────────────────────────────────────────────────────────────────────
# SUMMARY REPORT
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "="*68)
print("  STORAGE COMPS SUMMARY  --  740 E Utah Highland Dr, Lehi UT 84043")
print(f"  Search Radius: {SEARCH_RADIUS} mi (straight-line)  |  Date: {DATE_PULLED}")
print("="*68)

print(f"\n  Total Facilities Found Within ~{SEARCH_RADIUS} mi: {len(FACILITIES)}")
print()
print(f"  {'DIST':>5}  {'DRIVE':>5}  FACILITY")
print(f"  {'(mi)':>5}  {'(min)':>5}")
print("  " + "-"*62)
for f in FACILITIES:
    print(f"  {f['distance']:5.2f}  {f['drive_time']:5d}  {f['name']}")
    print(f"         {'':<7}       {f['address']}")

v10x10_on  = [p["online"]  for p in PRICING if p["unit_type"]=="10x10" and isinstance(p["online"],  (int,float))]
v10x10_in  = [p["instore"] for p in PRICING if p["unit_type"]=="10x10" and isinstance(p["instore"], (int,float))]
v10x20_on  = [p["online"]  for p in PRICING if p["unit_type"]=="10x20" and isinstance(p["online"],  (int,float))]
v10x20_in  = [p["instore"] for p in PRICING if p["unit_type"]=="10x20" and isinstance(p["instore"], (int,float))]

print()
print(f"  10x10  Online Rate  :  ${min(v10x10_on):.0f} - ${max(v10x10_on):.0f}/mo  "
      f"(avg ${sum(v10x10_on)/len(v10x10_on):.0f})   [{len(v10x10_on)} comps]")
if v10x10_in:
    print(f"  10x10  In-Store Rate:  ${min(v10x10_in):.0f} - ${max(v10x10_in):.0f}/mo  "
          f"(avg ${sum(v10x10_in)/len(v10x10_in):.0f})   [{len(v10x10_in)} comps]")
print()
print(f"  10x20  Online Rate  :  ${min(v10x20_on):.0f} - ${max(v10x20_on):.0f}/mo  "
      f"(avg ${sum(v10x20_on)/len(v10x20_on):.0f})  [{len(v10x20_on)} comps]")
if v10x20_in:
    print(f"  10x20  In-Store Rate:  ${min(v10x20_in):.0f} - ${max(v10x20_in):.0f}/mo  "
          f"(avg ${sum(v10x20_in)/len(v10x20_in):.0f})  [{len(v10x20_in)} comps]")

all_on = [(p["fac_name"], p["unit_type"], p["online"])
          for p in PRICING if isinstance(p["online"], (int,float))]
cheapest = min(all_on, key=lambda x: x[2])
priciest = max(all_on, key=lambda x: x[2])

print()
print(f"  Cheapest unit : {cheapest[0]}")
print(f"                  {cheapest[1]}  @  ${cheapest[2]:.0f}/mo (online rate)")
print()
print(f"  Priciest unit : {priciest[0]}")
print(f"                  {priciest[1]}  @  ${priciest[2]:.0f}/mo (online rate)")
print()
print(f"  Excel file    : {OUTPUT_PATH}")
print("="*68)
