"""
Storage Tools — Web Application (FastAPI)

Endpoints:
  POST /api/quick-estimate    — instant cost estimate (no API call)
  POST /api/comps             — start market comps agent (SSE stream)
  POST /api/accurate-estimate — start accurate cost agent (SSE stream)
  GET  /api/download/{file}   — download generated Excel file
  POST /api/register          — create account
  POST /api/login             — get JWT token
  GET  /api/usage             — get user's usage stats
  GET  /api/admin/users       — admin user dashboard
  GET  /admin                 — admin dashboard page
"""

import os
from dotenv import load_dotenv
load_dotenv()

import json
import asyncio
from datetime import date, datetime, timedelta
from typing import Optional
from collections import Counter

from fastapi import FastAPI, Request, Depends, HTTPException, status
from fastapi.responses import (
    HTMLResponse, FileResponse, StreamingResponse, JSONResponse,
)
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from sqlalchemy.orm import Session
from sqlalchemy import func

from passlib.context import CryptContext
from jose import jwt, JWTError

from claude_agent_sdk import (
    query, ClaudeAgentOptions, ResultMessage, AssistantMessage, TextBlock,
)

from database import init_db, get_db, User, UsageLog

# ── Config ────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# JWT settings
SECRET_KEY = os.environ.get("JWT_SECRET", "")
if not SECRET_KEY or SECRET_KEY == "change-me-in-production-use-a-real-secret":
    raise RuntimeError("JWT_SECRET must be set in .env — generate a strong random string (32+ chars)")
ALGORITHM = "HS256"
TOKEN_EXPIRE_HOURS = 24

# Server-side API key — users never see or touch this
SERVER_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")

# Admin email — only this user can access /api/admin endpoints
ADMIN_EMAIL = os.environ.get("ADMIN_EMAIL", "").lower()

# Password hashing
pwd_ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")


# ── Cost Data (same as desktop app) ──────────────────────────────────────────
# ── Property Type Cost Data ──────────────────────────────────────────────────
# Per-SF costs and lump-sum items for each property type.
# Sources: RSMeans 2024/2025 national averages, adjusted by location factor.

PROPERTY_TYPES = {
    "storage_driveup": {
        "label": "Self-Storage: Drive-Up",
        "has_stories": False,
        "per_sf": [
            ("Site Work & Grading",        5.50),
            ("Concrete Slab / Foundation", 8.00),
            ("Steel Structure",           18.00),
            ("Metal Roofing",              5.50),
            ("Electrical & Lighting",      4.00),
            ("Paving & Parking",           4.50),
        ],
        "lump": [
            ("Roll-Up Doors",    "~1 per 125 SF", lambda sf: int(sf / 125) * 1100),
            ("Security System",  "Lump sum",      lambda sf: 30000),
            ("Office Buildout",  "~400 SF",       lambda sf: 400 * 130),
        ],
    },
    "storage_cc": {
        "label": "Self-Storage: Climate Controlled",
        "has_stories": True,
        "per_sf": [
            ("Site Work & Grading",        5.50),
            ("Concrete Slab / Foundation", 9.50),
            ("Steel Structure",           22.00),
            ("Metal Roofing",              5.50),
            ("HVAC System",               12.00),
            ("Insulation",                 4.00),
            ("Interior Corridors",         6.00),
            ("Fire Suppression",           3.50),
            ("Electrical & Lighting",      5.00),
            ("Paving & Parking",           4.50),
        ],
        "lump": [
            ("Roll-Up / Entry Doors", "~1 per 100 SF", lambda sf: int(sf / 100) * 1200),
            ("Elevator",              "If > 20k SF",    lambda sf: 120000 if sf > 20000 else 0),
            ("Security System",       "Lump sum",       lambda sf: 40000),
            ("Office Buildout",       "~400 SF",        lambda sf: 400 * 140),
        ],
    },
    "retail_qsr": {
        "label": "Retail / QSR (Shell)",
        "has_stories": False,
        "per_sf": [
            ("Site Work & Grading",          8.00),
            ("Foundation",                  12.00),
            ("Structural Framing",          18.00),
            ("Exterior Walls & Façade",     22.00),
            ("Roofing",                      8.00),
            ("Storefront / Windows",        10.00),
            ("Electrical & Lighting",        8.00),
            ("Plumbing (Rough)",             6.00),
            ("HVAC (Shell Prep)",            5.00),
            ("Paving & Parking",             6.00),
        ],
        "lump": [
            ("Drive-Thru Infrastructure", "If applicable", lambda sf: 45000 if sf < 4000 else 0),
            ("Grease Trap / Hood Prep",   "QSR standard",  lambda sf: 25000),
            ("Signage Allowance",         "Lump sum",      lambda sf: 15000),
            ("ADA / Restrooms",           "Lump sum",      lambda sf: 20000),
        ],
    },
    "warehouse": {
        "label": "Warehouse / Distribution",
        "has_stories": False,
        "per_sf": [
            ("Site Work & Grading",          4.00),
            ("Concrete Slab / Foundation",  7.50),
            ("Steel Structure (Clear Span)", 16.00),
            ("Metal Roofing & Walls",        6.50),
            ("Dock Doors & Levelers",        3.00),
            ("Electrical & Lighting",        4.50),
            ("Fire Suppression (ESFR)",      3.50),
            ("Paving & Truck Courts",        5.00),
        ],
        "lump": [
            ("Office Buildout",   "~1,000 SF",   lambda sf: 1000 * 130),
            ("Security / Fencing", "Lump sum",   lambda sf: 35000),
        ],
    },
    "medical_office": {
        "label": "Medical Office",
        "has_stories": True,
        "per_sf": [
            ("Site Work & Grading",          6.00),
            ("Foundation",                  12.00),
            ("Structural Framing",          24.00),
            ("Exterior Walls & Façade",     18.00),
            ("Roofing",                      7.00),
            ("Interior Build-Out",          35.00),
            ("HVAC (Medical Grade)",        18.00),
            ("Plumbing (Medical)",          12.00),
            ("Electrical & Lighting",       12.00),
            ("Fire Suppression",             4.00),
            ("Paving & Parking",             5.00),
        ],
        "lump": [
            ("Elevator",           "If multi-story", lambda sf: 150000 if sf > 10000 else 0),
            ("Medical Gas Systems", "Lump sum",      lambda sf: 40000),
            ("ADA Compliance",     "Lump sum",       lambda sf: 25000),
        ],
    },
    "multifamily": {
        "label": "Multifamily — Garden Style",
        "has_stories": True,
        "per_sf": [
            ("Site Work & Grading",          5.50),
            ("Foundation",                  10.00),
            ("Wood / Steel Framing",        22.00),
            ("Exterior Walls & Siding",     14.00),
            ("Roofing",                      6.00),
            ("Windows & Doors",              8.00),
            ("Interior Finishes",           28.00),
            ("HVAC (Per Unit)",             10.00),
            ("Plumbing (Per Unit)",         12.00),
            ("Electrical & Lighting",        8.00),
            ("Fire Suppression",             3.50),
            ("Paving & Parking",             4.00),
        ],
        "lump": [
            ("Elevator",           "If 3+ stories", lambda sf: 160000 if sf > 30000 else 0),
            ("Clubhouse / Amenity", "Lump sum",     lambda sf: 80000),
            ("Landscaping",        "Lump sum",       lambda sf: 50000),
        ],
    },
}

QUALITY_MULT = {"Economy": 0.85, "Average": 1.00, "Premium": 1.15}

SOFT_COSTS = [
    ("Architectural & Engineering",   0.050),
    ("Permits & Impact Fees",         0.025),
    ("Geotechnical / Environmental",  0.008),
    ("Survey & Land Planning",        0.004),
    ("Legal & Closing",               0.008),
    ("Builder's Risk Insurance",      0.007),
    ("Construction Loan Interest",    0.040),
    ("Property Taxes During Const.",  0.008),
    ("Contingency",                   0.075),
]

LOCATION_FACTORS = {
    "new york": 1.42, "manhattan": 1.48, "brooklyn": 1.42, "bronx": 1.42,
    "los angeles": 1.18, "chicago": 1.12, "houston": 0.88, "phoenix": 0.92,
    "philadelphia": 1.15, "san antonio": 0.85, "san diego": 1.15,
    "dallas": 0.90, "austin": 0.92, "fort worth": 0.89,
    "jacksonville": 0.87, "san francisco": 1.38, "san jose": 1.30,
    "columbus": 0.93, "charlotte": 0.88, "indianapolis": 0.92,
    "seattle": 1.15, "denver": 0.96, "nashville": 0.90,
    "atlanta": 0.90, "portland": 1.08, "las vegas": 0.98,
    "memphis": 0.85, "louisville": 0.90, "baltimore": 0.98,
    "milwaukee": 1.02, "albuquerque": 0.90, "tucson": 0.90,
    "fresno": 1.05, "sacramento": 1.12, "miami": 0.95,
    "tampa": 0.90, "orlando": 0.90, "st louis": 0.98,
    "pittsburgh": 1.00, "raleigh": 0.88, "minneapolis": 1.05,
    "cleveland": 0.98, "detroit": 1.00, "boston": 1.25,
    "honolulu": 1.35, "anchorage": 1.28, "kansas city": 0.95,
    "oklahoma city": 0.85, "omaha": 0.90, "virginia beach": 0.90,
    "colorado springs": 0.93, "tulsa": 0.84, "arlington": 0.89,
    "new orleans": 0.88, "bakersfield": 1.05, "boise": 0.92,
    "richmond": 0.90, "des moines": 0.92, "salt lake city": 0.93,
    "birmingham": 0.85, "spokane": 1.00, "rochester": 1.02,
}


def _lookup_location_factor(city_text: str) -> tuple[float, str]:
    lower = city_text.lower().strip().rstrip(",").strip()
    for city, factor in LOCATION_FACTORS.items():
        if city in lower:
            return factor, city.title()
    return 1.00, ""


# ── System Prompts ───────────────────────────────────────────────────────────
COMPS_SYSTEM_PROMPT = """
You are a self-storage market research analyst. Output a formatted Excel file.
MINIMIZE web fetches — every extra fetch costs money.

## Research Strategy (strictly in order)

1. ONE WebSearch: "self storage [location] prices site:sparefoot.com"
2. Fetch the top SpareFoot result — extract EVERY facility listed on that page.
3. If SpareFoot has fewer than 5 facilities, do ONE more WebSearch:
   "self storage near [location] pricing" and fetch the top 2 results.
4. Only visit individual facility websites if a facility has no pricing data
   from the above sources. Limit to 3 individual sites maximum.

Find every facility within the search radius — do not skip any.
Total fetches allowed: 6 maximum.

## Collect per facility:
Name | Address | Phone | Website | Online rates by unit size | In-store rates |
Climate controlled (Yes/No) | Promotions | Distance (mi) & drive time (min)
from subject using Haversine formula, 25 mph estimate.

Include only facilities within the search radius. Sort closest first.

## Excel — 3 Tabs

Tab 1 "Comps Detail" — one row per facility x unit type
Columns: Facility Name | Address | Distance (mi) | Unit Type | Sq Ft |
  Climate Controlled | Online Rate ($/mo) | In-Store Rate ($/mo) |
  Notes/Promotions | Date Pulled | Source URL | Drive Time (min)
Format: bold header (#BDD7EE), currency, auto-width columns, freeze row 1

Tab 2 "Market Summary" — two sections separated by a blank row
  Section 1 header: "CLIMATE CONTROLLED UNITS" (bold, green #E2EFDA)
  Section 2 header: "DRIVE UP / STANDARD UNITS" (bold, orange #FCE4D6)
  Each section columns: Unit Type | Sq Ft | Avg Online | Min Online | Max Online |
    Avg In-Store | Min In-Store | Max In-Store | # Comps
  Currency format on all rate columns.

Tab 3 "Facility List" — one row per facility
Columns: Facility Name | Address | Distance (mi) | Drive Time (min) | Phone | Website
Format: bold header (#FCE4D6), sorted by distance

## Rules
- Never fabricate pricing. Mark missing data as N/A.
- Distance required on every row — exclude facility if unknown.
- Write Excel via openpyxl using the Bash tool.
"""

COST_AGENT_PROMPT = """
You are a self-storage construction cost analyst. Your job is to research CURRENT,
ACCURATE construction costs for a specific city and building type, then write an
Excel cost estimate.

MINIMIZE web fetches — every extra fetch costs money. Max 3 fetches total.

## Research Strategy
1. ONE WebSearch: "self storage construction cost per square foot [city] [year]"
2. ONE WebSearch: "USACE area cost factor [city] [year]" OR
   "RSMeans construction cost [city] [year]"
3. Fetch the 1-2 most relevant results to get real numbers.

## What to find:
- Current $/SF for self-storage construction in or near that city
- Location cost factor vs national average
- Breakdown by component if available (site work, steel, concrete, HVAC, etc.)
- Any recent project cost data for self-storage in the area

## Output
Write an Excel file using openpyxl with these details:

Tab 1 "Cost Estimate":
Row 1: "Self Storage Construction Cost Estimate" (bold, size 14)
Row 2: Building Type
Row 3: Total SF — put the numeric SF value in cell B3
Row 4: City + location factor found
Row 5: Quality level
Row 6: Date + sources used
Row 7: blank
Row 8: Header row (bold, #BDD7EE fill): Component | $/SF | Total Cost
Rows 9+: One row per cost component with real researched $/SF values

CRITICAL MATH RULES — follow these exactly:
- Column B = $/SF (numeric value, formatted as currency)
- Column C = Total Cost — MUST use an Excel formula: =B{row}*$B$3
  Do NOT type a hardcoded number in column C. Every Total Cost cell MUST be a formula.
- Hard Cost Subtotal row: Column C = SUM formula of all hard cost rows above
- Each Soft Cost row: Column C = formula referencing Hard Cost Subtotal * percentage
- TOTAL ESTIMATED COST row: = Hard Cost Subtotal + Soft Cost Subtotal (formula)
- TOTAL $/SF row: = TOTAL / $B$3 (formula)
- Double-check: $/SF * Total SF MUST equal Total Cost for every row. Using formulas guarantees this.

Then: Hard Cost Subtotal (bold)
Then: Itemized Soft Costs (~22.5% total: A&E 5%, Permits 2.5%, Geotech 0.8%, Survey 0.4%, Legal 0.8%, Insurance 0.7%, Loan Interest 4%, Property Tax 0.8%, Contingency 7.5%)
Then: TOTAL ESTIMATED COST (bold, #E2EFDA fill) with $/SF and total

Tab 2 "Sources":
List every URL and source used with what data came from each.

## Rules
- Use REAL data from your research, not generic estimates.
- If you cannot find city-specific data, use the closest metro area and note that.
- Always include the source of each number.
- Currency format on all dollar columns.
- Auto-width all columns.
- NEVER hardcode Total Cost values — always use formulas so the math is guaranteed correct.
"""


# ── FastAPI App ──────────────────────────────────────────────────────────────
app = FastAPI(title="Storage Tools")

app.mount("/static", StaticFiles(directory=os.path.join(BASE_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(BASE_DIR, "templates"))


@app.on_event("startup")
def startup():
    init_db()


# ── Auth helpers ─────────────────────────────────────────────────────────────
def create_token(email: str) -> str:
    expire = datetime.utcnow() + timedelta(hours=TOKEN_EXPIRE_HOURS)
    return jwt.encode({"sub": email, "exp": expire}, SECRET_KEY, algorithm=ALGORITHM)


def get_current_user(request: Request) -> str:
    """Extract user email from Authorization header or cookie."""
    token = None
    auth = request.headers.get("Authorization", "")
    if auth.startswith("Bearer "):
        token = auth[7:]
    else:
        token = request.cookies.get("token")

    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        email = payload.get("sub")
        if not email:
            raise HTTPException(status_code=401, detail="Invalid token")
        return email
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid or expired token")


def _get_server_api_key() -> str:
    """Get the server-side API key from environment."""
    if not SERVER_API_KEY:
        raise HTTPException(500, "Server API key not configured. Contact administrator.")
    return SERVER_API_KEY


def _check_rate_limit(user: User, db: Session):
    """Check if user has exceeded their per-user daily search limit."""
    limit = user.daily_limit if user.daily_limit is not None else 0
    if limit <= 0:
        raise HTTPException(429, "No searches available. Contact administrator for access.")
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    count = db.query(func.count(UsageLog.id)).filter(
        UsageLog.user_id == user.id,
        UsageLog.created_at >= today_start,
        UsageLog.action.in_(["comps", "accurate_estimate"]),
    ).scalar()
    if count >= limit:
        raise HTTPException(429, f"Daily limit reached ({limit} searches/day). Try again tomorrow.")


def _log_usage(user_id: int, action: str, location: str, db: Session):
    """Record a usage event."""
    db.add(UsageLog(user_id=user_id, action=action, location=location))
    db.commit()


# ── Request/Response Models ──────────────────────────────────────────────────
class RegisterRequest(BaseModel):
    email: str
    password: str

class LoginRequest(BaseModel):
    email: str
    password: str

class QuickEstimateRequest(BaseModel):
    building_type: str
    sf: float
    city: str = ""
    quality: str = "Average"
    loan_rate: float = 8.5      # annual interest rate %
    const_months: int = 12      # construction period in months

class CompsRequest(BaseModel):
    location: str
    radius: float = 5.0

class AccurateEstimateRequest(BaseModel):
    building_type: str
    sf: float
    city: str
    quality: str = "Average"
    stories: int = 1
    loan_rate: float = 8.5
    const_months: int = 12


# ── Pages ────────────────────────────────────────────────────────────────────
@app.get("/", response_class=HTMLResponse)
async def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.get("/admin", response_class=HTMLResponse)
async def admin_page(request: Request):
    return templates.TemplateResponse("admin.html", {"request": request})


# ── Auth API ─────────────────────────────────────────────────────────────────
@app.post("/api/register")
async def register(req: RegisterRequest, db: Session = Depends(get_db)):
    existing = db.query(User).filter(User.email == req.email).first()
    if existing:
        raise HTTPException(400, "Email already registered")

    user = User(
        email=req.email,
        password_hash=pwd_ctx.hash(req.password),
    )
    db.add(user)
    db.commit()
    db.refresh(user)

    token = create_token(req.email)
    is_admin = (req.email.lower() == ADMIN_EMAIL)
    return {"token": token, "email": req.email, "is_admin": is_admin}


@app.post("/api/login")
async def login(req: LoginRequest, db: Session = Depends(get_db)):
    user = db.query(User).filter(User.email == req.email).first()
    if not user or not pwd_ctx.verify(req.password, user.password_hash):
        raise HTTPException(401, "Invalid email or password")
    if not user.is_active:
        raise HTTPException(403, "Account is disabled")

    token = create_token(req.email)
    is_admin = (req.email.lower() == ADMIN_EMAIL)

    # Auto-grant admin unlimited searches
    if is_admin and (user.daily_limit is None or user.daily_limit == 0):
        user.daily_limit = 99999
        db.commit()

    return {"token": token, "email": req.email, "is_admin": is_admin}


# ── Usage stats ──────────────────────────────────────────────────────────────
@app.get("/api/usage")
async def get_usage(request: Request, db: Session = Depends(get_db)):
    email = get_current_user(request)
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(404, "User not found")

    # Today's usage
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    today_count = db.query(func.count(UsageLog.id)).filter(
        UsageLog.user_id == user.id,
        UsageLog.created_at >= today_start,
        UsageLog.action.in_(["comps", "accurate_estimate"]),
    ).scalar()

    # All-time usage
    total_count = db.query(func.count(UsageLog.id)).filter(
        UsageLog.user_id == user.id,
    ).scalar()

    # Last 10 searches
    recent = db.query(UsageLog).filter(
        UsageLog.user_id == user.id,
    ).order_by(UsageLog.created_at.desc()).limit(10).all()

    return {
        "today": today_count,
        "daily_limit": user.daily_limit if user.daily_limit is not None else 0,
        "total": total_count,
        "recent": [
            {"action": r.action, "location": r.location or "", "date": r.created_at.isoformat()}
            for r in recent
        ],
    }


# ── Admin dashboard ──────────────────────────────────────────────────────────
@app.get("/api/admin/users")
async def admin_users(request: Request, db: Session = Depends(get_db)):
    email = get_current_user(request)
    if ADMIN_EMAIL and email.lower() != ADMIN_EMAIL:
        raise HTTPException(403, "Admin access only")

    users = db.query(User).order_by(User.created_at.desc()).all()
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)

    result = []
    for u in users:
        total_searches = db.query(func.count(UsageLog.id)).filter(
            UsageLog.user_id == u.id,
            UsageLog.action.in_(["comps", "accurate_estimate"]),
        ).scalar()
        today_searches = db.query(func.count(UsageLog.id)).filter(
            UsageLog.user_id == u.id,
            UsageLog.action.in_(["comps", "accurate_estimate"]),
            UsageLog.created_at >= today_start,
        ).scalar()
        last_use = db.query(UsageLog.created_at).filter(
            UsageLog.user_id == u.id,
        ).order_by(UsageLog.created_at.desc()).first()

        result.append({
            "id": u.id,
            "email": u.email,
            "created": u.created_at.isoformat() if u.created_at else "",
            "active": u.is_active,
            "subscription": u.subscription_tier or "free",
            "daily_limit": u.daily_limit if u.daily_limit is not None else 0,
            "total_searches": total_searches,
            "today_searches": today_searches,
            "last_active": last_use[0].isoformat() if last_use else "never",
        })

    return {"users": result, "total_users": len(result)}


@app.post("/api/admin/toggle-user/{user_id}")
async def admin_toggle_user(user_id: int, request: Request, db: Session = Depends(get_db)):
    email = get_current_user(request)
    if ADMIN_EMAIL and email.lower() != ADMIN_EMAIL:
        raise HTTPException(403, "Admin access only")

    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(404, "User not found")
    user.is_active = not user.is_active
    db.commit()
    return {"ok": True, "email": user.email, "active": user.is_active}


@app.post("/api/admin/set-tier/{user_id}")
async def admin_set_tier(user_id: int, request: Request, db: Session = Depends(get_db)):
    email = get_current_user(request)
    if ADMIN_EMAIL and email.lower() != ADMIN_EMAIL:
        raise HTTPException(403, "Admin access only")
    body = await request.json()
    tier = body.get("tier", "free")
    if tier not in ("free", "pro", "enterprise"):
        raise HTTPException(400, "Invalid tier")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(404, "User not found")
    user.subscription_tier = tier
    db.commit()
    return {"ok": True, "email": user.email, "tier": tier}


@app.post("/api/admin/set-limit/{user_id}")
async def admin_set_limit(user_id: int, request: Request, db: Session = Depends(get_db)):
    email = get_current_user(request)
    if ADMIN_EMAIL and email.lower() != ADMIN_EMAIL:
        raise HTTPException(403, "Admin access only")
    body = await request.json()
    limit = int(body.get("limit", 0))
    if limit < 0:
        raise HTTPException(400, "Limit must be >= 0")
    user = db.query(User).filter(User.id == user_id).first()
    if not user:
        raise HTTPException(404, "User not found")
    user.daily_limit = limit
    db.commit()
    return {"ok": True, "email": user.email, "daily_limit": limit}


# ── Quick Estimate (no API call) ─────────────────────────────────────────────
@app.post("/api/quick-estimate")
async def quick_estimate(req: QuickEstimateRequest, request: Request, db: Session = Depends(get_db)):
    email = get_current_user(request)
    user = db.query(User).filter(User.email == email).first()

    sf = req.sf
    if sf <= 0:
        raise HTTPException(400, "SF must be positive")

    btype = req.building_type
    quality = req.quality
    q_mult = QUALITY_MULT.get(quality, 1.0)

    loc_factor, matched_city = _lookup_location_factor(req.city)

    ptype = PROPERTY_TYPES.get(btype)
    if not ptype:
        raise HTTPException(400, f"Unknown property type: {btype}")

    per_sf_items = ptype["per_sf"]
    lump_items = ptype["lump"]

    rows = []
    total_hard = 0.0

    for name, base_psf in per_sf_items:
        adj_psf = base_psf * q_mult * loc_factor
        cost = adj_psf * sf
        rows.append({"name": name, "psf": f"${adj_psf:,.2f}", "cost": f"${cost:,.0f}", "cost_raw": cost})
        total_hard += cost

    for name, note, calc_fn in lump_items:
        cost = calc_fn(sf) * q_mult * loc_factor
        if cost > 0:
            rows.append({"name": name, "psf": note, "cost": f"${cost:,.0f}", "cost_raw": cost})
            total_hard += cost

    # Calculate construction loan interest from user inputs
    loan_rate = max(0, min(req.loan_rate, 25))  # cap at 25%
    const_months = max(1, min(req.const_months, 48))  # cap at 48 months
    # Average outstanding balance is ~50% of total (draws happen over time)
    loan_interest = total_hard * (loan_rate / 100) * (const_months / 12) * 0.5

    soft_items = []
    total_soft = 0.0
    for name, pct in SOFT_COSTS:
        if name == "Construction Loan Interest":
            # Use calculated value instead of flat percentage
            soft_items.append({
                "name": f"Construction Loan Interest ({loan_rate:.1f}% × {const_months}mo)",
                "pct": f"{loan_interest / total_hard:.1%}" if total_hard > 0 else "0.0%",
                "cost": f"${loan_interest:,.0f}",
                "cost_raw": loan_interest,
            })
            total_soft += loan_interest
        else:
            amt = total_hard * pct
            soft_items.append({"name": name, "pct": f"{pct:.1%}", "cost": f"${amt:,.0f}", "cost_raw": amt})
            total_soft += amt

    grand_total = total_hard + total_soft
    total_psf = grand_total / sf if sf > 0 else 0

    # Log usage (free action, but still track it)
    if user:
        _log_usage(user.id, "quick_estimate", req.city or "N/A", db)

    return {
        "building_type": ptype["label"],
        "sf": sf,
        "city": req.city or "National Avg",
        "quality": quality,
        "location_factor": loc_factor,
        "matched_city": matched_city,
        "hard_cost_rows": rows,
        "total_hard": f"${total_hard:,.0f}",
        "total_hard_raw": total_hard,
        "soft_cost_rows": soft_items,
        "total_soft": f"${total_soft:,.0f}",
        "grand_total": f"${grand_total:,.0f}",
        "grand_total_raw": grand_total,
        "total_psf": f"${total_psf:,.2f}",
    }


# ── Quick Estimate Excel Export ───────────────────────────────────────────────
@app.post("/api/quick-estimate-export")
async def quick_estimate_export(req: QuickEstimateRequest, request: Request, db: Session = Depends(get_db)):
    """Generate an Excel file from a quick estimate with a Sources tab."""
    email = get_current_user(request)

    sf = req.sf
    if sf <= 0:
        raise HTTPException(400, "SF must be positive")

    btype = req.building_type
    quality = req.quality
    q_mult = QUALITY_MULT.get(quality, 1.0)
    loc_factor, matched_city = _lookup_location_factor(req.city)

    ptype = PROPERTY_TYPES.get(btype)
    if not ptype:
        raise HTTPException(400, f"Unknown property type: {btype}")

    loan_rate = max(0, min(req.loan_rate, 25))
    const_months = max(1, min(req.const_months, 48))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, numbers, Alignment

    wb = Workbook()

    # ── Tab 1: Cost Estimate ──
    ws = wb.active
    ws.title = "Cost Estimate"

    header_font = Font(bold=True, size=14)
    bold_font = Font(bold=True)
    currency_fmt = '#,##0'
    currency_psf_fmt = '$#,##0.00'
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    total_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.merge_cells("A1:C1")
    ws["A1"] = "Construction Cost Estimate — Quick Estimate"
    ws["A1"].font = header_font

    ws["A2"] = "Property Type:"
    ws["B2"] = ptype["label"]
    ws["A3"] = "Total SF:"
    ws["B3"] = sf
    ws["B3"].number_format = '#,##0'
    ws["A4"] = "City:"
    ws["B4"] = f"{req.city or 'National Avg'} (Location Factor: {loc_factor:.2f}x{' — ' + matched_city if matched_city else ''})"
    ws["A5"] = "Quality:"
    ws["B5"] = quality
    ws["A6"] = "Date:"
    ws["B6"] = date.today().strftime("%B %d, %Y")
    ws["A7"] = "Loan Rate:"
    ws["B7"] = f"{loan_rate:.1f}% annual × {const_months} months"

    row = 9
    for col, label in [(1, "Component"), (2, "$/SF"), (3, "Total Cost")]:
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = bold_font
        cell.fill = header_fill
    row += 1

    hard_start = row
    total_hard = 0.0
    for name, base_psf in ptype["per_sf"]:
        adj_psf = base_psf * q_mult * loc_factor
        cost = adj_psf * sf
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=adj_psf).number_format = currency_psf_fmt
        ws.cell(row=row, column=3, value=cost).number_format = currency_fmt
        total_hard += cost
        row += 1

    for name, note, calc_fn in ptype["lump"]:
        cost = calc_fn(sf) * q_mult * loc_factor
        if cost > 0:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=note)
            ws.cell(row=row, column=3, value=cost).number_format = currency_fmt
            total_hard += cost
            row += 1

    # Hard cost subtotal
    ws.cell(row=row, column=1, value="HARD COST SUBTOTAL").font = bold_font
    ws.cell(row=row, column=2, value=total_hard / sf if sf > 0 else 0).number_format = currency_psf_fmt
    ws.cell(row=row, column=2).font = bold_font
    ws.cell(row=row, column=3, value=total_hard).number_format = currency_fmt
    ws.cell(row=row, column=3).font = bold_font
    row += 2

    # Soft costs header
    ws.cell(row=row, column=1, value="Soft Costs").font = bold_font
    ws.cell(row=row, column=2, value="% of Hard").font = bold_font
    row += 1

    total_soft = 0.0
    loan_interest = total_hard * (loan_rate / 100) * (const_months / 12) * 0.5

    for name, pct in SOFT_COSTS:
        if name == "Construction Loan Interest":
            ws.cell(row=row, column=1, value=f"Construction Loan Interest ({loan_rate:.1f}% × {const_months}mo)")
            effective_pct = loan_interest / total_hard if total_hard > 0 else 0
            ws.cell(row=row, column=2, value=f"{effective_pct:.1%}")
            ws.cell(row=row, column=3, value=loan_interest).number_format = currency_fmt
            total_soft += loan_interest
        else:
            amt = total_hard * pct
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=f"{pct:.1%}")
            ws.cell(row=row, column=3, value=amt).number_format = currency_fmt
            total_soft += amt
        row += 1

    # Soft cost subtotal
    ws.cell(row=row, column=1, value="SOFT COST SUBTOTAL").font = bold_font
    ws.cell(row=row, column=3, value=total_soft).number_format = currency_fmt
    ws.cell(row=row, column=3).font = bold_font
    row += 2

    # Grand total
    grand_total = total_hard + total_soft
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = total_fill
        ws.cell(row=row, column=col).font = bold_font
    ws.cell(row=row, column=1, value="TOTAL ESTIMATED COST")
    ws.cell(row=row, column=2, value=grand_total / sf if sf > 0 else 0).number_format = currency_psf_fmt
    ws.cell(row=row, column=3, value=grand_total).number_format = currency_fmt

    # Auto-width columns
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 18

    # ── Tab 2: Sources & Assumptions ──
    ws2 = wb.create_sheet("Sources & Assumptions")
    ws2["A1"] = "Sources & Assumptions"
    ws2["A1"].font = header_font

    ws2["A3"] = "Category"
    ws2["B3"] = "Assumption"
    ws2["C3"] = "Source"
    for col in range(1, 4):
        ws2.cell(row=3, column=col).font = bold_font
        ws2.cell(row=3, column=col).fill = header_fill

    sources = [
        ("Base Cost Data", "Per-SF costs are national averages for this property type",
         "RSMeans Building Construction Cost Data 2024/2025 — https://www.rsmeans.com"),
        ("Location Factor", f"{loc_factor:.2f}x applied to {matched_city or 'national average'}",
         "USACE Area Cost Factors (PAX) — https://www.usace.army.mil/Cost-Engineering/"),
        ("Quality Multiplier", f"{quality} = {q_mult:.2f}x adjustment",
         "RSMeans quality adjustment methodology — Economy (0.85x), Average (1.00x), Premium (1.15x)"),
        ("A&E (5%)", "Architectural & Engineering fees as % of hard costs",
         "AIA Compensation Report — https://www.aia.org/resources/compensation-report"),
        ("Permits & Impact Fees (2.5%)", "Building permits and municipal impact fees",
         "National Association of Home Builders — https://www.nahb.org/advocacy/impact-fees"),
        ("Geotech / Environmental (0.8%)", "Geotechnical study, Phase I/II ESA",
         "ASTM E1527-21 standard practice — typical Phase I ESA cost ranges"),
        ("Survey & Land Planning (0.4%)", "ALTA survey, site planning",
         "NSPS/ALTA Standards — https://www.nsps.us.com"),
        ("Legal & Closing (0.8%)", "Legal fees, title, closing costs",
         "Industry standard — varies by deal complexity and jurisdiction"),
        ("Builder's Risk Insurance (0.7%)", "Construction-period insurance coverage",
         "IRMI (International Risk Management Institute) — https://www.irmi.com"),
        ("Construction Loan Interest", f"{loan_rate:.1f}% annual rate × {const_months} months × 50% avg draw",
         "User-entered rate. Formula: hard costs × rate × (months/12) × 0.5 average outstanding balance"),
        ("Property Tax During Const. (0.8%)", "Ad valorem tax on land during construction",
         "County assessor rates — varies by jurisdiction"),
        ("Contingency (7.5%)", "Unforeseen conditions, change orders",
         "AACE International Recommended Practice 18R-97 — Class 3 estimate contingency range"),
    ]

    for i, (cat, assumption, source) in enumerate(sources, start=4):
        ws2.cell(row=i, column=1, value=cat)
        ws2.cell(row=i, column=2, value=assumption)
        ws2.cell(row=i, column=3, value=source)

    disclaimer_row = len(sources) + 6
    ws2.cell(row=disclaimer_row, column=1, value="DISCLAIMER").font = bold_font
    ws2.cell(row=disclaimer_row + 1, column=1,
             value="This is a preliminary estimate based on published cost data and standard industry assumptions. "
                   "Actual costs will vary based on site conditions, local labor markets, material pricing, "
                   "and project-specific requirements. Verify all figures with local contractors and consultants.")

    ws2.column_dimensions["A"].width = 35
    ws2.column_dimensions["B"].width = 55
    ws2.column_dimensions["C"].width = 75

    # Save and return
    safe_city = (req.city or "national").replace(" ", "_").replace(",", "")
    filename = f"quick_estimate_{safe_city}_{date.today().strftime('%b-%d-%y')}.xlsx"
    filepath = os.path.join(OUTPUT_DIR, filename)
    wb.save(filepath)

    return FileResponse(filepath, filename=filename,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Market Comps (SSE stream) ────────────────────────────────────────────────
@app.post("/api/comps")
async def run_comps(req: CompsRequest, request: Request, db: Session = Depends(get_db)):
    email = get_current_user(request)
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(404, "User not found")

    _check_rate_limit(user, db)
    api_key = _get_server_api_key()

    if req.radius > 7:
        raise HTTPException(400, "Maximum radius is 7 miles")

    location = req.location.strip()
    if not location:
        raise HTTPException(400, "Location is required")

    _log_usage(user.id, "comps", location, db)

    today = date.today().strftime("%b-%d-%y")
    safe_loc = location.replace(" ", "_").replace(",", "").replace("/", "-")
    output_file = os.path.join(OUTPUT_DIR, f"storage_comps_{safe_loc}_{today}.xlsx")

    async def event_stream():
        # Set server API key for this request
        os.environ["ANTHROPIC_API_KEY"] = api_key

        prompt = f"""
Find self-storage market rent comps for:
  Location : {location}
  Radius   : {req.radius} miles
  Date     : {date.today().strftime("%B %d, %Y")}
  Save to  : {output_file}

Instructions:
1. Search SpareFoot for ALL self-storage facilities within {req.radius} miles of {location}.
2. Find every facility in the radius — do not stop early.
3. Calculate distance/drive time from "{location}" for each facility.
4. Write the Excel file using openpyxl (3-tab format per system prompt).
5. Print a brief summary: facilities found, 10x10 price range, 10x20 price range.

No fabricated data — mark missing as N/A.
"""
        try:
            async for message in query(
                prompt=prompt,
                options=ClaudeAgentOptions(
                    system_prompt=COMPS_SYSTEM_PROMPT,
                    allowed_tools=["WebSearch", "WebFetch", "Bash", "Write"],
                    permission_mode="acceptEdits",
                    cwd=BASE_DIR,
                    max_turns=25,
                    model="claude-haiku-4-5",
                ),
            ):
                if isinstance(message, AssistantMessage):
                    for block in message.content:
                        if isinstance(block, TextBlock) and block.text.strip():
                            yield f"data: {json.dumps({'type': 'log', 'text': block.text})}\n\n"
                elif isinstance(message, ResultMessage):
                    fname = os.path.basename(output_file) if os.path.exists(output_file) else None
                    yield f"data: {json.dumps({'type': 'done', 'file': fname, 'result': message.result or ''})}\n\n"
        except Exception as exc:
            yield f"data: {json.dumps({'type': 'error', 'text': str(exc)})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


# ── Accurate Estimate (SSE stream) ───────────────────────────────────────────
@app.post("/api/accurate-estimate")
async def run_accurate_estimate(req: AccurateEstimateRequest, request: Request, db: Session = Depends(get_db)):
    email = get_current_user(request)
    user = db.query(User).filter(User.email == email).first()
    if not user:
        raise HTTPException(404, "User not found")

    _check_rate_limit(user, db)
    api_key = _get_server_api_key()

    if not req.city.strip():
        raise HTTPException(400, "City is required for accurate estimate")

    _log_usage(user.id, "accurate_estimate", req.city, db)

    today = date.today().strftime("%b-%d-%y")
    safe_city = req.city.replace(" ", "_").replace(",", "")
    ptype = PROPERTY_TYPES.get(req.building_type)
    btype_label = ptype["label"] if ptype else req.building_type
    output_file = os.path.join(OUTPUT_DIR, f"cost_estimate_{safe_city}_{today}.xlsx")

    async def event_stream():
        # Set server API key for this request
        os.environ["ANTHROPIC_API_KEY"] = api_key

        has_stories = ptype["has_stories"] if ptype else False
        stories = max(1, min(req.stories, 5)) if has_stories else 1
        stories_line = f"\n  Stories       : {stories}" if stories > 1 else ""
        stories_instruction = ""
        if stories > 1:
            stories_instruction = f"""
This is a {stories}-story building. Research actual multi-story construction cost
impacts for this property type and city. Account for structural, mechanical,
vertical transportation, and code compliance differences vs single-story.
"""

        prompt = f"""
Research and create a construction cost estimate for:
  Property Type : {btype_label}
  Total SF      : {req.sf:,.0f}{stories_line}
  City          : {req.city}
  Quality       : {req.quality}
  Date          : {date.today().strftime("%B %d, %Y")}
  Save to       : {output_file}
{stories_instruction}
Search for CURRENT construction costs specific to {btype_label} in {req.city} or the nearest major metro.
Find real $/SF data for this property type, not generic national averages.
Include itemized soft costs: A&E 5%, Permits & Impact Fees 2.5%, Geotech/Environmental 0.8%, Survey & Land Planning 0.4%, Legal & Closing 0.8%, Builder's Risk Insurance 0.7%, Property Taxes During Construction 0.8%, Contingency 7.5%.
For Construction Loan Interest, use: hard costs × {req.loan_rate}% annual rate × {req.const_months} months / 12 × 0.5 average draw factor.
Write the Excel file per the system prompt format.
"""
        try:
            async for message in query(
                prompt=prompt,
                options=ClaudeAgentOptions(
                    system_prompt=COST_AGENT_PROMPT,
                    allowed_tools=["WebSearch", "WebFetch", "Bash", "Write"],
                    permission_mode="acceptEdits",
                    cwd=BASE_DIR,
                    max_turns=10,
                    model="claude-haiku-4-5",
                ),
            ):
                if isinstance(message, AssistantMessage):
                    for block in message.content:
                        if isinstance(block, TextBlock) and block.text.strip():
                            yield f"data: {json.dumps({'type': 'log', 'text': block.text})}\n\n"
                elif isinstance(message, ResultMessage):
                    fname = os.path.basename(output_file) if os.path.exists(output_file) else None
                    yield f"data: {json.dumps({'type': 'done', 'file': fname, 'result': message.result or ''})}\n\n"
        except Exception as exc:
            yield f"data: {json.dumps({'type': 'error', 'text': str(exc)})}\n\n"

    return StreamingResponse(event_stream(), media_type="text/event-stream")


# ── File download ────────────────────────────────────────────────────────────
@app.get("/api/download/{filename}")
async def download_file(filename: str, request: Request, token: Optional[str] = None):
    # Accept token via query param for direct downloads
    if token:
        try:
            payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
            if not payload.get("sub"):
                raise HTTPException(401, "Invalid token")
        except JWTError:
            raise HTTPException(401, "Invalid or expired token")
    else:
        get_current_user(request)
    # Sanitize — prevent path traversal
    safe = os.path.basename(filename)
    path = os.path.join(OUTPUT_DIR, safe)
    if not os.path.exists(path):
        raise HTTPException(404, "File not found")
    return FileResponse(path, filename=safe,
                        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# ── Run ──────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
