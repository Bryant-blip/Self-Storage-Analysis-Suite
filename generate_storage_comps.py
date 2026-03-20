"""
Generate an Excel file with self-storage market analysis for a specific property.
Uses openpyxl to create three tabs: Comps Detail, Market Summary, and Facility List.
"""

import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_storage_comps_excel():
    """Create and populate the storage comps Excel file."""

    # Define output directory and filename
    output_dir = r"C:\Users\jones\Claude code Projects\Real Estate Project\output"
    os.makedirs(output_dir, exist_ok=True)

    filename = "storage_comps_424_N_Beeline_Highway_Payson_AZ_85441_Mar-19-26.xlsx"
    filepath = os.path.join(output_dir, filename)

    # Create workbook
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    header_font = Font(bold=True)

    market_summary_cc_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    market_summary_du_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    facility_list_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    # Define facilities data
    facilities = [
        {
            "name": "Swiss Village Self Storage",
            "address": "635 N Beeline Hwy, Payson, AZ 85541",
            "distance": 0.70,
            "drive_time": 2,
            "phone": "(928) 474-6339",
            "website": "sparefoot.com"
        },
        {
            "name": "Rim Country Self-Storage",
            "address": "208 N Tonto St, Payson, AZ 85541",
            "distance": 0.82,
            "drive_time": 2,
            "phone": "(928) 468-7867",
            "website": "storagepro.com"
        },
        {
            "name": "Star Valley Storage",
            "address": "83 N Orion Dr, Payson, AZ 85541",
            "distance": 1.04,
            "drive_time": 2,
            "phone": "(928) 474-5593",
            "website": "sparefoot.com"
        },
        {
            "name": "Mclane Self Storage",
            "address": "100 S Mc Lane Rd, Payson, AZ 85541",
            "distance": 1.37,
            "drive_time": 3,
            "phone": "(928) 474-1566",
            "website": "uhaul.com"
        },
        {
            "name": "Alpine Mini Storage",
            "address": "203 E Sherwood Dr, Payson, AZ 85541",
            "distance": 1.42,
            "drive_time": 3,
            "phone": "(928) 474-6656",
            "website": "moverscorp.com"
        },
        {
            "name": "Dudleys Mini Storage",
            "address": "206 S Dudley St, Payson, AZ 85541",
            "distance": 1.87,
            "drive_time": 4,
            "phone": "(928) 474-3987",
            "website": "sparefoot.com"
        },
        {
            "name": "A Storage Place",
            "address": "1500 W Red Baron Rd, Payson, AZ 85541",
            "distance": 2.47,
            "drive_time": 6,
            "phone": "(928) 474-9383",
            "website": "storagepro.com"
        },
        {
            "name": "Jot Mini Storage",
            "address": "900 S Mclane Rd, Payson, AZ 85541",
            "distance": 2.53,
            "drive_time": 6,
            "phone": "(928) 474-0334",
            "website": "uhaul.com"
        },
        {
            "name": "Country Meadows Storage",
            "address": "4230 N Az Highway 87, Payson, AZ 85541",
            "distance": 3.64,
            "drive_time": 9,
            "phone": "(928) 476-4165",
            "website": "moverscorp.com"
        }
    ]

    # Climate controlled status for each facility
    climate_controlled = {
        "Swiss Village Self Storage": "Unknown",
        "Rim Country Self-Storage": "Unknown",
        "Star Valley Storage": "Unknown",
        "Mclane Self Storage": "Unknown",
        "Alpine Mini Storage": "Unknown",
        "Dudleys Mini Storage": "Unknown",
        "A Storage Place": "Unknown",
        "Jot Mini Storage": "Unknown",
        "Country Meadows Storage": "Unknown"
    }

    unit_types = [
        {"type": "10x10", "sq_ft": 100},
        {"type": "10x20", "sq_ft": 200}
    ]

    date_pulled = "03/19/2026"

    # ===== TAB 1: COMPS DETAIL =====
    ws_detail = wb.create_sheet("Comps Detail", 0)

    # Headers
    headers_detail = [
        "Facility Name", "Address", "Distance (mi)", "Unit Type", "Sq Ft",
        "Climate Controlled", "Online Rate ($/mo)", "In-Store Rate ($/mo)",
        "Notes/Promotions", "Date Pulled", "Source URL", "Drive Time (min)"
    ]

    ws_detail.append(headers_detail)

    # Format header
    for col_num, header in enumerate(headers_detail, 1):
        cell = ws_detail.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Add data rows
    for facility in facilities:
        for unit in unit_types:
            ws_detail.append([
                facility["name"],
                facility["address"],
                facility["distance"],
                unit["type"],
                unit["sq_ft"],
                climate_controlled[facility["name"]],
                "N/A",
                "N/A",
                "",
                date_pulled,
                facility["website"],
                facility["drive_time"]
            ])

    # Auto-size columns
    column_widths_detail = [25, 35, 12, 10, 8, 18, 16, 16, 20, 12, 20, 14]
    for col_num, width in enumerate(column_widths_detail, 1):
        ws_detail.column_dimensions[get_column_letter(col_num)].width = width

    # Freeze header
    ws_detail.freeze_panes = "A2"

    # ===== TAB 2: MARKET SUMMARY =====
    ws_summary = wb.create_sheet("Market Summary", 1)

    row_num = 1

    # Section 1: Climate Controlled Units
    ws_summary.append(["CLIMATE CONTROLLED UNITS"])
    cell = ws_summary.cell(row=row_num, column=1)
    cell.font = Font(bold=True)
    cell.fill = market_summary_cc_fill
    row_num += 1

    # Headers for climate controlled
    headers_summary = [
        "Unit Type", "Sq Ft", "Avg Online", "Min Online", "Max Online",
        "Avg In-Store", "Min In-Store", "Max In-Store", "# Comps"
    ]
    ws_summary.append(headers_summary)

    # Format header
    for col_num, header in enumerate(headers_summary, 1):
        cell = ws_summary.cell(row=row_num, column=col_num)
        cell.font = header_font
        cell.fill = market_summary_cc_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 1

    # Add climate controlled data rows (all N/A)
    for unit in unit_types:
        ws_summary.append([
            unit["type"],
            unit["sq_ft"],
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            0
        ])
        row_num += 1

    # Blank row
    row_num += 1

    # Section 2: Drive Up / Standard Units
    ws_summary.append(["DRIVE UP / STANDARD UNITS"])
    cell = ws_summary.cell(row=row_num, column=1)
    cell.font = Font(bold=True)
    cell.fill = market_summary_du_fill
    row_num += 1

    # Headers for drive up
    ws_summary.append(headers_summary)

    # Format header
    for col_num, header in enumerate(headers_summary, 1):
        cell = ws_summary.cell(row=row_num, column=col_num)
        cell.font = header_font
        cell.fill = market_summary_du_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    row_num += 1

    # Add drive up data rows (all N/A)
    for unit in unit_types:
        ws_summary.append([
            unit["type"],
            unit["sq_ft"],
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            "N/A",
            0
        ])
        row_num += 1

    # Auto-size columns
    column_widths_summary = [15, 10, 12, 12, 12, 14, 14, 14, 10]
    for col_num, width in enumerate(column_widths_summary, 1):
        ws_summary.column_dimensions[get_column_letter(col_num)].width = width

    # Freeze header (freeze at row 2 for first section and row after blank)
    ws_summary.freeze_panes = "A2"

    # ===== TAB 3: FACILITY LIST =====
    ws_facilities = wb.create_sheet("Facility List", 2)

    # Headers
    headers_facilities = [
        "Facility Name", "Address", "Distance (mi)", "Drive Time (min)", "Phone", "Website"
    ]

    ws_facilities.append(headers_facilities)

    # Format header
    for col_num, header in enumerate(headers_facilities, 1):
        cell = ws_facilities.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = facility_list_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sort facilities by distance and add data rows
    sorted_facilities = sorted(facilities, key=lambda x: x["distance"])
    for facility in sorted_facilities:
        ws_facilities.append([
            facility["name"],
            facility["address"],
            facility["distance"],
            facility["drive_time"],
            facility["phone"],
            facility["website"]
        ])

    # Auto-size columns
    column_widths_facilities = [30, 40, 14, 16, 15, 20]
    for col_num, width in enumerate(column_widths_facilities, 1):
        ws_facilities.column_dimensions[get_column_letter(col_num)].width = width

    # Freeze header
    ws_facilities.freeze_panes = "A2"

    # Save workbook
    wb.save(filepath)

    return filepath, len(facilities)


if __name__ == "__main__":
    filepath, num_facilities = create_storage_comps_excel()
    print(f"Success! Excel file created at:")
    print(f"{filepath}")
    print(f"\nSummary:")
    print(f"- Total facilities found: {num_facilities}")
    print(f"- Market radius: 5 miles from 424 N Beeline Hwy, Payson, AZ 85441")
    print(f"- Unit types analyzed: 10x10 (100 sq ft), 10x20 (200 sq ft)")
    print(f"- Tabs created: Comps Detail, Market Summary, Facility List")
