"""
db_utils.py

Shared SQLite database utilities for the Storage Intel dashboard.
Imported by both app.py (Flask) and crexi_watcher.py (pipeline) to avoid
circular dependencies.

Database: data/deals.db
"""

import os
import sqlite3
from datetime import datetime, timezone

DB = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "deals.db")


def get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA journal_mode=WAL")
    return conn


def init_db():
    """Create all tables if they don't exist. Safe to call multiple times."""
    os.makedirs(os.path.dirname(DB), exist_ok=True)
    conn = get_db()
    try:
        conn.executescript("""
            CREATE TABLE IF NOT EXISTS deals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                listing_id TEXT UNIQUE,
                market TEXT,
                address TEXT,
                zip_code TEXT,
                lat REAL,
                lng REAL,
                asking_price REAL,
                acres REAL,
                price_per_acre REAL,
                avg_psf REAL,
                avg_psf_drive_up REAL,
                avg_psf_climate REAL,
                population_3mi INTEGER,
                population_density_per_sqmi REAL,
                construction_cost_per_sqft REAL,
                yield_on_cost REAL,
                deal_score REAL,
                nearby_facility_count INTEGER,
                report_path TEXT,
                crexi_url TEXT,
                scraped_at TEXT,
                processed_at TEXT,
                skip_reason TEXT,
                zip_pool_count INTEGER,
                pop_gate_passed TEXT,
                city_name TEXT
            );

            CREATE TABLE IF NOT EXISTS comps (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                listing_id TEXT,
                facility_name TEXT,
                facility_address TEXT,
                distance_miles REAL,
                unit_size TEXT,
                unit_type TEXT,
                web_rate REAL,
                in_store_rate REAL,
                rate_per_sqft REAL,
                scraped_at TEXT,
                FOREIGN KEY (listing_id) REFERENCES deals(listing_id)
            );

            CREATE TABLE IF NOT EXISTS census_cache (
                zip_code TEXT PRIMARY KEY,
                population INTEGER,
                population_density_per_sqmi REAL,
                queried_at TEXT
            );

            CREATE TABLE IF NOT EXISTS watcher_runs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                job_id TEXT UNIQUE,
                market TEXT,
                max_deals INTEGER,
                dry_run INTEGER,
                status TEXT,            -- running | finished | stopped | error
                deals_found INTEGER DEFAULT 0,
                started_at TEXT,
                finished_at TEXT,
                exit_code INTEGER
            );

            CREATE INDEX IF NOT EXISTS idx_deals_market ON deals(market);
            CREATE INDEX IF NOT EXISTS idx_deals_zip ON deals(zip_code);
            CREATE INDEX IF NOT EXISTS idx_deals_processed ON deals(processed_at);
            CREATE INDEX IF NOT EXISTS idx_comps_listing ON comps(listing_id);
            CREATE INDEX IF NOT EXISTS idx_watcher_runs_started ON watcher_runs(started_at DESC);
        """)
        conn.commit()

        # Add population_density_per_sqmi to census_cache if it was created
        # by the older census_pop.py (which had a 2-column schema)
        try:
            conn.execute("ALTER TABLE census_cache ADD COLUMN population_density_per_sqmi REAL")
            conn.commit()
        except sqlite3.OperationalError:
            pass  # Column already exists

        # Add pop_gate_passed for existing DBs created before the triple-gate upgrade
        try:
            conn.execute("ALTER TABLE deals ADD COLUMN pop_gate_passed TEXT")
            conn.commit()
        except sqlite3.OperationalError:
            pass  # Column already exists

        # Add city_name for existing DBs created before city_name was added
        try:
            conn.execute("ALTER TABLE deals ADD COLUMN city_name TEXT")
            conn.commit()
        except sqlite3.OperationalError:
            pass  # Column already exists

    finally:
        conn.close()


def recalculate_scores(conn: sqlite3.Connection):
    """
    Normalize YoC, population_3mi, and land_cost_efficiency across all
    processed, non-skipped deals, then write:
        deal_score = yoc_norm*50 + pop_norm*35 + lce_norm*15

    Hard gates (deals failing either get deal_score = NULL):
        population_3mi  >= 30,000
        yield_on_cost   >= 0.10   (10.0%)

    Land cost efficiency (LCE) = avg_psf / (price_per_acre / 43560)
        Higher value = more rental income per dollar of land cost = better.

    Normalization: (value - min) / (max - min) * 100, clamped 0–100.
    Deals with NULL for any metric receive a neutral 50 for that component.
    """
    rows = conn.execute("""
        SELECT listing_id, yield_on_cost, population_3mi,
               avg_psf, price_per_acre
        FROM deals
        WHERE processed_at IS NOT NULL AND skip_reason IS NULL
    """).fetchall()

    if not rows:
        return

    # Compute LCE for each row
    records = []
    for r in rows:
        yoc  = r["yield_on_cost"]
        pop  = r["population_3mi"]
        psf  = r["avg_psf"]
        ppa  = r["price_per_acre"]
        lce  = (psf / (ppa / 43560)) if (psf and ppa and ppa > 0) else None
        records.append({
            "listing_id": r["listing_id"],
            "yoc":  yoc,
            "pop":  pop,
            "lce":  lce,
            "passes_gates": (
                yoc is not None and yoc >= 0.10 and
                (pop is None or pop >= 30000)   # null pop = no data, not a disqualifier
            ),
        })

    def _norm(values):
        """Normalize a list of floats (None → neutral 50). Returns same-length list."""
        valid = [v for v in values if v is not None]
        if not valid or len(valid) < 2:
            return [50.0] * len(values)
        mn, mx = min(valid), max(valid)
        if mx == mn:
            return [50.0] * len(values)
        return [
            max(0.0, min(100.0, (v - mn) / (mx - mn) * 100)) if v is not None else 50.0
            for v in values
        ]

    yoc_norms = _norm([r["yoc"]  for r in records])
    pop_norms = _norm([r["pop"]  for r in records])
    lce_norms = _norm([r["lce"]  for r in records])

    for rec, yn, pn, ln in zip(records, yoc_norms, pop_norms, lce_norms):
        if rec["passes_gates"]:
            score = round(yn * 0.50 + pn * 0.35 + ln * 0.15, 1)
        else:
            score = None
        conn.execute(
            "UPDATE deals SET deal_score = ? WHERE listing_id = ?",
            (score, rec["listing_id"]),
        )
    conn.commit()


def calc_proforma_cells(ws) -> dict:
    """
    Read raw assumption cells from an openpyxl proforma worksheet.
    Returns dict with keys: acres, asking_price, avg_psf, cost_per_sqft,
    yield_pct, occupancy, expense_ratio, cap_rate, facility_type.
    Any cell that is missing or non-numeric → None.

    Mixed-facility reports (mixed_proforma_template.xlsx) don't have the
    single E5-E10 assumptions block — instead they carry a CC mini-proforma
    (rows 13-20) and a DU mini-proforma (rows 22-29) that roll up into a
    main summary. When E5 is empty but B15 (CC rentable sqft) has a value,
    this is detected as the mixed layout and the return dict is built from
    a sqft-weighted blend of the two mini-proformas instead. In that case
    an extra "du_psf" key (drive-up rent $/sqft from D24) is also returned.
    """
    def _float(cell_ref):
        try:
            v = ws[cell_ref].value
            return float(v) if v is not None else None
        except (TypeError, ValueError):
            return None

    def _str(cell_ref):
        v = ws[cell_ref].value
        return str(v).strip() if v is not None else None

    acres = _float("C5")
    e5 = _float("E5")
    cc_sqft = _float("B15")

    if e5 is None and cc_sqft is not None:
        # Mixed layout: CC mini-proforma rows 13-20, DU mini-proforma rows 22-29
        du_sqft = _float("B24")
        cc_rent = _float("D15")
        du_rent = _float("D24")
        cc_cost = _float("D18")
        du_cost = _float("D27")
        cc_occ  = _float("D16")
        du_occ  = _float("D25")
        cc_exp  = _float("D17")
        du_exp  = _float("D26")

        def _blend(cc_val, du_val):
            if cc_val is None or du_val is None or du_sqft is None:
                return None
            total = cc_sqft + du_sqft
            return (cc_val * cc_sqft + du_val * du_sqft) / total if total else None

        def _blend_if_diff(cc_val, du_val):
            if cc_val is None or du_val is None:
                return None
            if cc_val == du_val:
                return cc_val
            return _blend(cc_val, du_val)

        yield_pct = None
        if acres and du_sqft is not None:
            land_sqft = acres * 43560
            yield_pct = (cc_sqft + du_sqft) / land_sqft if land_sqft else None

        return {
            "acres":          acres,
            "asking_price":   _float("C6"),
            "avg_psf":        _blend(cc_rent, du_rent),
            "yield_pct":      yield_pct,
            "occupancy":      _blend_if_diff(cc_occ, du_occ),
            "expense_ratio":  _blend_if_diff(cc_exp, du_exp),
            "cap_rate":       _float("C8"),
            "cost_per_sqft":  _blend(cc_cost, du_cost),
            "facility_type":  "mixed",
            "du_psf":         du_rent,
        }

    return {
        "acres":          acres,
        "asking_price":   _float("C6"),
        "avg_psf":        _float("E6"),
        "yield_pct":      e5,
        "occupancy":      _float("E7"),
        "expense_ratio":  _float("E8"),
        "cap_rate":       _float("E9"),
        "cost_per_sqft":  _float("E10"),
        "facility_type":  _str("E3"),
    }


def _calc_yoc(cells: dict):
    """Derive yield_on_cost from raw proforma cells. Returns None if inputs missing."""
    try:
        acres         = cells["acres"]
        rent_psf      = cells["avg_psf"]
        yield_pct     = cells["yield_pct"]
        occupancy     = cells["occupancy"]
        expense_ratio = cells["expense_ratio"]
        cap_rate      = cells["cap_rate"]  # noqa: F841 (unused in YoC but kept for completeness)
        cost_per_sqft = cells["cost_per_sqft"]
        asking_price  = cells["asking_price"]

        if None in (acres, rent_psf, yield_pct, occupancy, expense_ratio, cost_per_sqft):
            return None

        net_rentable = acres * 43560 * yield_pct
        annual_noi   = net_rentable * rent_psf * occupancy * (1 - expense_ratio) * 12
        construction = cost_per_sqft * net_rentable
        total_cost   = construction + (asking_price or 0)
        return annual_noi / total_cost if total_cost else None
    except (TypeError, ZeroDivisionError):
        return None


def write_deal_to_db(
    listing_id: str,
    report_path: str,
    market: str,
    address: str,
    url: str,
    lat,
    lng,
    population_3mi,
    zip_code: str,
    zip_pool_count,
    first_seen: str,
    facilities: list,
    pop_gate_passed: str = None,
    city_name: str = None,
    recalc: bool = True,
):
    """
    Write a processed deal to SQLite. Reads proforma cells from the Excel
    report, calculates financial metrics, inserts/replaces in deals table,
    inserts comps rows, and recalculates all scores.

    Never raises — all errors are caught and logged to stderr.
    """
    import sys

    try:
        import openpyxl

        cells = {}
        if report_path and os.path.exists(report_path):
            try:
                wb = openpyxl.load_workbook(report_path, data_only=True)
                proforma = next(
                    (wb[n] for n in wb.sheetnames
                     if n.strip().lower() in {"proforma", "initial look proforma",
                                              "initial proforma"}),
                    None,
                )
                if proforma:
                    cells = calc_proforma_cells(proforma)

                # Count facilities from Facility List tab
                fac_tab = next(
                    (wb[n] for n in wb.sheetnames if "facility" in n.strip().lower()),
                    None,
                )
            except Exception as exc:
                print(f"  [db_utils] Warning: could not read report {report_path}: {exc}",
                      file=sys.stderr)
        else:
            fac_tab = None

        acres        = cells.get("acres")
        asking_price = cells.get("asking_price")
        avg_psf      = cells.get("avg_psf")
        cost_per_sqft = cells.get("cost_per_sqft")
        facility_type = cells.get("facility_type")

        price_per_acre = (asking_price / acres) if (asking_price and acres) else None
        if facility_type == "multi_story":
            # avg_psf here is a climate-controlled rate, not drive-up — no
            # drive-up rate exists for a multi-story facility.
            avg_psf_drive_up = None
        elif facility_type == "mixed":
            du_psf = cells.get("du_psf")
            avg_psf_drive_up = (du_psf + 0.05) if du_psf is not None else None
        else:
            avg_psf_drive_up = (avg_psf + 0.05) if avg_psf is not None else None
        yield_on_cost    = _calc_yoc(cells)
        nearby_facility_count = len(facilities) if facilities else (
            # Fall back to counting Facility List tab rows
            (sum(1 for r in fac_tab.iter_rows(min_row=2) if any(c.value for c in r))
             if fac_tab else None)
        )

        now = datetime.now(timezone.utc).isoformat(timespec="seconds")

        conn = get_db()
        try:
            conn.execute("""
                INSERT OR REPLACE INTO deals
                    (listing_id, market, address, zip_code, lat, lng,
                     asking_price, acres, price_per_acre, avg_psf, avg_psf_drive_up,
                     construction_cost_per_sqft, yield_on_cost, nearby_facility_count,
                     report_path, crexi_url, scraped_at, processed_at, zip_pool_count,
                     population_3mi, pop_gate_passed, city_name)
                VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                listing_id, market, address, zip_code, lat, lng,
                asking_price, acres, price_per_acre, avg_psf, avg_psf_drive_up,
                cost_per_sqft, yield_on_cost, nearby_facility_count,
                report_path, url, first_seen, now, zip_pool_count,
                population_3mi, pop_gate_passed, city_name,
            ))

            # Insert comps rows
            if facilities:
                for fac in facilities:
                    fac_name    = fac.get("name", "")
                    fac_address = fac.get("address", "")
                    dist        = fac.get("distance_miles")
                    for price in fac.get("pricing", []):
                        size      = price.get("size")
                        unit_type = price.get("type")
                        web_rate  = price.get("web_rate")
                        in_store  = price.get("in_store_rate")
                        # rate_per_sqft = web_rate / sqft
                        from comps_pipeline import UNIT_SF
                        sqft = UNIT_SF.get(size) if size else None
                        rate_psf = (web_rate / sqft) if (web_rate and sqft) else None
                        conn.execute("""
                            INSERT OR IGNORE INTO comps
                                (listing_id, facility_name, facility_address,
                                 distance_miles, unit_size, unit_type,
                                 web_rate, in_store_rate, rate_per_sqft, scraped_at)
                            VALUES (?,?,?,?,?,?,?,?,?,?)
                        """, (listing_id, fac_name, fac_address, dist,
                              size, unit_type, web_rate, in_store, rate_psf, now))

            conn.commit()
            if recalc:
                recalculate_scores(conn)
        finally:
            conn.close()

    except Exception as exc:
        print(f"  [db_utils] write_deal_to_db failed for {listing_id}: {exc}",
              file=sys.stderr)
