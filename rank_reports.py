"""
rank_reports.py

Build a ranked spreadsheet of all processed deals.

Two sections:
  1. RANKED   — deals that passed hard gates (YoC >= 10%, pop >= 30k),
                sorted by deal_score descending.
  2. UNGATED  — deals that failed the hard gates, sorted by YoC descending.

Each row links to the underlying report file and to the Crexi listing.

Usage:
    python rank_reports.py [--output PATH]
"""

import argparse
import os
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


HEADER_FILL_RANKED  = PatternFill("solid", fgColor="1A7A4A")  # green
HEADER_FILL_UNGATED = PatternFill("solid", fgColor="64748B")  # grey
SECTION_FILL        = PatternFill("solid", fgColor="E2EFDA")
THIN                = Side(style="thin", color="CCCCCC")
BORDER              = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    ("rank",            "Rank",         8),
    ("deal_score",      "Score",        8),
    ("city_name",       "City",         22),
    ("market",          "State",        14),
    ("address",         "Address",      36),
    ("acres",           "Acres",        8),
    ("asking_price",    "Asking $",     14),
    ("price_per_acre",  "$/Acre",       12),
    ("avg_psf",         "Avg PSF",      9),
    ("yield_on_cost",   "YoC",          8),
    ("population_3mi",  "Population",   12),
    ("pop_gate_passed", "Pop Gate",     14),
    ("crexi_url",       "Crexi Listing", 75),
    ("report_path",     "Report",        8),
]


def _fmt(v, kind):
    if v is None or v == "":
        return ""
    if kind == "usd":
        return f"${v:,.0f}"
    if kind == "psf":
        return f"${v:.2f}"
    if kind == "pct":
        return f"{v*100:.1f}%"
    if kind == "num":
        return f"{int(v):,}"
    if kind == "acres":
        return f"{v:.2f}"
    return v


def _kind(col_key):
    return {
        "asking_price":   "usd",
        "price_per_acre": "usd",
        "avg_psf":        "psf",
        "yield_on_cost":  "pct",
        "population_3mi": "num",
        "acres":          "acres",
    }.get(col_key)


def _write_header_row(ws, row: int, fill: PatternFill):
    for col_idx, (_, label, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=row, column=col_idx, value=label)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER


def _write_deal_row(ws, row: int, deal: dict, rank: int):
    for col_idx, (key, _, _) in enumerate(COLUMNS, 1):
        c = ws.cell(row=row, column=col_idx)
        if key == "rank":
            c.value = rank
            c.alignment = Alignment(horizontal="center")
        elif key == "crexi_url":
            url = deal.get("crexi_url")
            if url:
                c.value = url
                c.hyperlink = url
                c.style = "Hyperlink"
            c.alignment = Alignment(horizontal="left")
        elif key == "report_path":
            path = deal.get("report_path")
            if path and os.path.exists(path):
                c.value = "Report"
                c.hyperlink = f"file:///{os.path.abspath(path).replace(os.sep, '/')}"
                c.style = "Hyperlink"
            c.alignment = Alignment(horizontal="center")
        else:
            v = deal.get(key)
            kind = _kind(key)
            c.value = _fmt(v, kind) if kind else (v if v is not None else "")
        c.border = BORDER


def _write_section_header(ws, row: int, title: str, count: int):
    last_col = get_column_letter(len(COLUMNS))
    ws.merge_cells(f"A{row}:{last_col}{row}")
    c = ws.cell(row=row, column=1, value=f"{title}  ({count} deals)")
    c.font = Font(bold=True, size=13, color="0F172A")
    c.fill = SECTION_FILL
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[row].height = 22


def main():
    parser = argparse.ArgumentParser(description="Rank all processed reports into a spreadsheet")
    parser.add_argument("--output", default=None, help="Output xlsx path (default: output/Ranked_Reports_<date>.xlsx)")
    args = parser.parse_args()

    if not args.output:
        os.makedirs("output", exist_ok=True)
        args.output = os.path.join(
            "output", f"Ranked_Reports_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
        )

    from db_utils import get_db

    conn = get_db()
    rows = conn.execute("""
        SELECT listing_id, market, address, city_name, zip_code,
               acres, asking_price, price_per_acre, avg_psf,
               yield_on_cost, population_3mi, pop_gate_passed,
               deal_score, crexi_url, report_path
        FROM deals
        WHERE processed_at IS NOT NULL AND skip_reason IS NULL
          AND report_path IS NOT NULL
    """).fetchall()
    conn.close()

    deals = [dict(r) for r in rows]

    ranked  = sorted(
        [d for d in deals if d.get("deal_score") is not None],
        key=lambda d: d["deal_score"], reverse=True,
    )
    ungated = sorted(
        [d for d in deals if d.get("deal_score") is None],
        key=lambda d: d.get("yield_on_cost") or 0, reverse=True,
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Deals"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    # Title
    last_col = get_column_letter(len(COLUMNS))
    ws.merge_cells(f"A1:{last_col}1")
    c = ws.cell(row=1, column=1, value=f"Self-Storage Deal Rankings — {datetime.now().strftime('%B %d, %Y')}")
    c.font = Font(bold=True, size=16, color="0F172A")
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 28

    row = 3

    # Section 1: Ranked
    _write_section_header(ws, row, "RANKED — passed hard gates (YoC ≥ 10%, Pop ≥ 30k)", len(ranked))
    row += 1
    _write_header_row(ws, row, HEADER_FILL_RANKED)
    row += 1
    for i, d in enumerate(ranked, 1):
        _write_deal_row(ws, row, d, i)
        row += 1

    row += 2  # spacer

    # Section 2: Ungated
    _write_section_header(ws, row, "UNGATED — failed hard gates (sorted by YoC)", len(ungated))
    row += 1
    _write_header_row(ws, row, HEADER_FILL_UNGATED)
    row += 1
    for i, d in enumerate(ungated, 1):
        _write_deal_row(ws, row, d, i)
        row += 1

    # Column widths
    for col_idx, (_, _, width) in enumerate(COLUMNS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(args.output)
    print(f"Wrote ranked spreadsheet -> {args.output}")
    print(f"  Ranked:  {len(ranked)} deals")
    print(f"  Ungated: {len(ungated)} deals")
    print(f"  Total:   {len(deals)} reports")


if __name__ == "__main__":
    main()
