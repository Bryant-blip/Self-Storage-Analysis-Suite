"""
sort_reports.py

Reorganizes Excel reports in the reports/ folder into tiered subfolders
based on the higher of Drive-Up In-Store or CC In-Store average rent per sqft.

Tier folders (within reports/):
  Markets>1.50 PSF/  — max In-Store avg >= $1.50/sqft
  Markets>1.25 PSF/  — max In-Store avg >= $1.25/sqft
  Markets>1.00 PSF/  — max In-Store avg >= $1.00/sqft
  Markets>0.90 PSF/  — max In-Store avg >= $0.90/sqft

Each tier folder is sub-divided by state (e.g., Washington/, Utah/).
Files below $0.90 on both metrics are logged but not moved.

Usage:
    python sort_reports.py
    python sort_reports.py --dry-run    # preview moves without touching files
"""

import argparse
import os
import shutil
import sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install it with: pip install openpyxl")
    sys.exit(1)

# ── Config ────────────────────────────────────────────────────────────────────
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
REPORTS_DIR = os.path.join(PROJECT_DIR, "reports")

TIERS = [
    (1.50, "Markets 1.50+ PSF"),
    (1.25, "Markets 1.25+ PSF"),
    (1.00, "Markets 1.00+ PSF"),
    (0.90, "Markets 0.90+ PSF"),
]

TIER_FOLDER_NAMES = {name for _, name in TIERS}


# ── Excel reading ─────────────────────────────────────────────────────────────
import re as _re


def _parse_avg_refs(formula: str) -> tuple[list[str], int | None]:
    """
    Parse '=AVERAGE(E4,G4,J4)/25' → (['E4','G4','J4'], 25)
    Parse '=AVERAGE(C26,C27,C28)' → (['C26','C27','C28'], None)
    """
    m = _re.match(r"=AVERAGE\(([^)]+)\)(?:/(\d+))?", formula, _re.IGNORECASE)
    if not m:
        return [], None
    refs = [r.strip() for r in m.group(1).split(",")]
    divisor = int(m.group(2)) if m.group(2) else None
    return refs, divisor


def _compute_panel_avg(ws, total_avg_formula: str) -> float | None:
    """
    Trace the formula chain to compute the true average $/sqft.

    total_avg_formula = '=AVERAGE(C26,C27,C28,...)'  (per-size avg cells)
    Each per-size cell has   '=AVERAGE(E4,G4,...)/sqft'  (price cells / sqft)
    Price cells contain actual dollar values (numbers).
    """
    size_refs, _ = _parse_avg_refs(total_avg_formula)
    if not size_refs:
        return None

    per_size_psf = []
    for ref in size_refs:
        cell = ws[ref]
        if not isinstance(cell.value, str):
            continue
        price_refs, divisor = _parse_avg_refs(cell.value)
        if not price_refs or not divisor:
            continue
        prices = [ws[pr].value for pr in price_refs if isinstance(ws[pr].value, (int, float))]
        if prices:
            per_size_psf.append(sum(prices) / len(prices) / divisor)

    return round(sum(per_size_psf) / len(per_size_psf), 4) if per_size_psf else None


def read_total_averages(xlsx_path: str) -> tuple[float | None, float | None]:
    """
    Open the 'Market Comps' sheet and compute (drive_up_instore_avg, cc_instore_avg).

    Scans column B for 'total average' labels, reads the formula from column C,
    then traces the formula chain back to actual price values to compute the average.
    First 'total average' in col B = Drive-Up In-Store; second = CC In-Store.
    """
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except Exception as exc:
        print(f"  ERROR opening {os.path.basename(xlsx_path)}: {exc}")
        return None, None

    if "Market Comps" not in wb.sheetnames:
        print(f"  WARN: 'Market Comps' sheet not found in {os.path.basename(xlsx_path)}")
        return None, None

    ws = wb["Market Comps"]
    results = []

    for row in ws.iter_rows(min_col=2, max_col=3):
        label_cell, val_cell = row[0], row[1]
        if (isinstance(label_cell.value, str)
                and label_cell.value.strip().lower() == "total average"):
            v = val_cell.value
            if isinstance(v, (int, float)):
                # Older files: value written directly as a number
                results.append(float(v))
            elif isinstance(v, str) and v.startswith("="):
                # Newer files: live formula — trace back to price cells
                results.append(_compute_panel_avg(ws, v))
            else:
                results.append(None)
            if len(results) == 2:
                break

    du = results[0] if len(results) > 0 else None
    cc = results[1] if len(results) > 1 else None
    return du, cc


# ── State extraction ──────────────────────────────────────────────────────────
def extract_state(folder_name: str) -> str:
    """
    Extract state name from folder name pattern '{State}_{YYYY-MM-DD}'.
    'North_Carolina_2026-04-12' → 'North Carolina'
    'Utah_2026-04-12'           → 'Utah'
    """
    # Strip trailing date (last 11 chars: '_YYYY-MM-DD')
    if len(folder_name) > 11 and folder_name[-11] == "_":
        state_raw = folder_name[:-11]
    else:
        state_raw = folder_name
    return state_raw.replace("_", " ")


def is_tier_folder(path: str) -> bool:
    """Return True if any part of the path is a known tier folder name."""
    parts = path.replace("\\", "/").split("/")
    return any(part in TIER_FOLDER_NAMES for part in parts)


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Sort reports by rent PSF into tier folders")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview moves without touching any files")
    args = parser.parse_args()

    if not os.path.exists(REPORTS_DIR):
        print(f"ERROR: reports/ folder not found at {REPORTS_DIR}")
        sys.exit(1)

    dry = args.dry_run
    if dry:
        print("DRY RUN — no files will be moved\n")

    moved = 0
    skipped_low = 0
    skipped_unreadable = 0
    already_sorted = 0

    # Walk all .xlsx files under reports/
    for root, dirs, files in os.walk(REPORTS_DIR):
        for fname in sorted(files):
            if not fname.lower().endswith(".xlsx"):
                continue

            full_path = os.path.join(root, fname)
            rel_path  = os.path.relpath(full_path, REPORTS_DIR)

            # Skip files already in a tier folder
            if is_tier_folder(rel_path):
                already_sorted += 1
                continue

            # Extract state from parent folder name
            parent_folder = os.path.basename(root)
            state = extract_state(parent_folder)

            print(f"\n{rel_path}")
            du_avg, cc_avg = read_total_averages(full_path)

            # Format for display
            du_str = f"${du_avg:.2f}" if du_avg is not None else "n/a"
            cc_str = f"${cc_avg:.2f}" if cc_avg is not None else "n/a"
            print(f"  Drive-Up In-Store avg: {du_str}  |  CC In-Store avg: {cc_str}")

            avgs = [v for v in [du_avg, cc_avg] if v is not None]
            if not avgs:
                print("  SKIP — could not read averages (file may not have been opened in Excel)")
                skipped_unreadable += 1
                continue

            max_avg = max(avgs)
            tier_name = next((name for threshold, name in TIERS if max_avg >= threshold), None)

            if tier_name is None:
                print(f"  SKIP — below $0.90/sqft (max avg: ${max_avg:.2f})")
                skipped_low += 1
                continue

            # Destination: reports/{tier}/{state}/{filename}
            dest_dir  = os.path.join(REPORTS_DIR, tier_name, state)
            dest_path = os.path.join(dest_dir, fname)

            print(f"  -> {tier_name}/{state}/{fname}")

            if not dry:
                os.makedirs(dest_dir, exist_ok=True)
                if os.path.exists(dest_path):
                    # Avoid overwriting — append a suffix
                    base, ext = os.path.splitext(fname)
                    dest_path = os.path.join(dest_dir, f"{base}_dup{ext}")
                shutil.move(full_path, dest_path)

            moved += 1

    print(f"\n{'-'*60}")
    if dry:
        print("DRY RUN complete:")
        print(f"  Would move:     {moved}")
    else:
        print("Sort complete:")
        print(f"  Moved:          {moved}")
    print(f"  Already sorted: {already_sorted}")
    print(f"  Unreadable:     {skipped_unreadable}")
    print(f"  Below $0.90:    {skipped_low}")

    # Clean up empty source folders
    if not dry:
        _remove_empty_folders(REPORTS_DIR)


def _remove_empty_folders(root: str) -> None:
    """Remove empty subdirectories under root (not tier folders themselves)."""
    for dirpath, dirnames, filenames in os.walk(root, topdown=False):
        if dirpath == root:
            continue
        rel = os.path.relpath(dirpath, root)
        if is_tier_folder(rel):
            continue
        if not os.listdir(dirpath):
            os.rmdir(dirpath)
            print(f"  Removed empty folder: {rel}")


def sort_single_report(xlsx_path: str, market: str) -> str | None:
    """
    Sort a single newly-generated report into the appropriate tier folder.
    Called automatically by the pipeline after each Excel report is saved.

    Returns the new file path if moved, or the original path if not sorted
    (below threshold or unreadable).
    """
    du_avg, cc_avg = read_total_averages(xlsx_path)
    avgs = [v for v in [du_avg, cc_avg] if v is not None]
    if not avgs:
        return xlsx_path  # no comps data — leave in place

    max_avg = max(avgs)
    tier_name = next((name for threshold, name in TIERS if max_avg >= threshold), None)
    if tier_name is None:
        return xlsx_path  # below lowest threshold — leave in place

    state = market.strip() or "Unknown"
    dest_dir  = os.path.join(REPORTS_DIR, tier_name, state)
    dest_path = os.path.join(dest_dir, os.path.basename(xlsx_path))

    os.makedirs(dest_dir, exist_ok=True)
    if os.path.exists(dest_path):
        base, ext = os.path.splitext(os.path.basename(xlsx_path))
        dest_path = os.path.join(dest_dir, f"{base}_dup{ext}")

    shutil.move(xlsx_path, dest_path)
    return dest_path


if __name__ == "__main__":
    main()
