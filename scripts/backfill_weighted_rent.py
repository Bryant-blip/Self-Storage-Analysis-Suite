"""
backfill_weighted_rent.py

Re-reads the Drive-Up online prices from the Market Comps tab of every
existing Excel report, calculates the weighted average $/sqft using the
current UNIT_MIX_WEIGHTS, subtracts the $0.05 new-competitor discount,
and writes the result to Proforma cell E6.

Usage:
    python backfill_weighted_rent.py           # dry run — shows values
    python backfill_weighted_rent.py --apply   # writes to files
"""

import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # run from repo root or scripts/

import glob
import os
import sys

try:
    import openpyxl
except ImportError:
    print("Missing openpyxl — run: pip install openpyxl")
    sys.exit(1)

APPLY   = "--apply" in sys.argv
REPORTS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")

# Must match comps_pipeline.py
UNIT_SF = {"5x5": 25, "5x10": 50, "10x10": 100, "10x15": 150,
           "10x20": 200, "10x25": 250, "10x30": 300}
UNIT_MIX_WEIGHTS = {"5x5": 0.12, "5x10": 0.25, "10x10": 0.30, "10x15": 0.15,
                    "10x20": 0.12, "10x25": 0.00, "10x30": 0.06}
NEW_COMPETITOR_DISCOUNT = 0.05


def extract_driveup_online_rates(ws) -> dict[str, list[float]]:
    """
    Scan the Market Comps sheet for Drive-Up Online (discounted) price rows.

    Structure confirmed from live files:
      col 1 = "Online (Discounted) Rates"  ← section header (Drive-Up side)
      rows below: col 1 = sqft int, col 2 = size string, cols 3+ = facility prices
      Drive-Up is always on the LEFT (lowest column index for that label).

    Returns {size: [price1, price2, ...]} with raw dollar amounts (not $/sqft).
    """
    UNIT_SIZES = set(UNIT_SF.keys())

    # Find the row number of the Drive-Up "Online (Discounted) Rates" header.
    # There are two such headers (one for DU, one for CC); we want the one at
    # the lowest column index (Drive-Up is always on the left side).
    online_header_row = None
    online_header_col = None

    for row in ws.iter_rows():
        for cell in row:
            val = str(cell.value or "").strip().lower()
            if "online" in val and "discounted" in val or val == "online (discounted) rates":
                if online_header_col is None or cell.column < online_header_col:
                    online_header_row = cell.row
                    online_header_col = cell.column

    if online_header_row is None:
        return {}

    # Read data rows immediately below the header.
    # Stop when col 1 is no longer a sqft integer.
    size_prices: dict[str, list[float]] = {}
    for row in ws.iter_rows(min_row=online_header_row + 1,
                             max_row=online_header_row + 15):
        col1 = row[0].value
        col2 = row[1].value if len(row) > 1 else None

        # Data row signature: col1 is a known sqft value, col2 is a size string
        if col1 not in UNIT_SF.values() or col2 not in UNIT_SIZES:
            continue

        # Collect all positive numeric values from col 3 onward in the same row
        prices = [
            float(c.value)
            for c in row[2:]
            if isinstance(c.value, (int, float)) and c.value > 0
        ]
        if prices:
            size_prices[col2] = prices

    return size_prices


def calc_weighted_rent(size_prices: dict[str, list[float]]) -> float | None:
    """
    Calculate weighted average drive-up online $/sqft across all sizes,
    then subtract the new-competitor discount.
    Returns None if no usable data.
    """
    size_avgs: list[tuple[float, float]] = []  # (avg_$/sqft, weight)

    for size, prices in size_prices.items():
        w  = UNIT_MIX_WEIGHTS.get(size, 0.0)
        sf = UNIT_SF.get(size)
        if w <= 0 or not prices or not sf:
            continue
        avg_psf = (sum(prices) / len(prices)) / sf
        size_avgs.append((avg_psf, w))

    if not size_avgs:
        return None

    total_w  = sum(w for _, w in size_avgs)
    weighted = sum(avg * (w / total_w) for avg, w in size_avgs)
    return round(weighted - NEW_COMPETITOR_DISCOUNT, 2)


def update_report(path: str) -> tuple[str, float | None, float | None]:
    """
    Process a single report file.
    Returns (status, old_e6, new_e6).
    """
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as exc:
        return f"ERROR opening file: {exc}", None, None

    # Find proforma sheet
    proforma = next(
        (wb[n] for n in wb.sheetnames
         if n.strip().lower() in {"proforma", "initial look proforma", "initial proforma"}),
        None,
    )
    if proforma is None:
        return "ERROR: no proforma tab", None, None

    # Find Market Comps sheet
    comps = next(
        (wb[n] for n in wb.sheetnames if "market comps" in n.strip().lower()),
        None,
    )
    if comps is None:
        return "ERROR: no Market Comps tab", None, None

    old_e6  = proforma["E6"].value
    prices  = extract_driveup_online_rates(comps)
    new_e6  = calc_weighted_rent(prices)

    if new_e6 is None:
        return "SKIP: no drive-up online rates found", old_e6, None

    if APPLY:
        proforma["E6"] = new_e6
        wb.save(path)

    return "OK", old_e6, new_e6


def main():
    paths = sorted(glob.glob(os.path.join(REPORTS, "**", "*.xlsx"), recursive=True))
    if not paths:
        print("No .xlsx files found in reports/")
        return

    if not APPLY:
        print("DRY RUN — pass --apply to write changes\n")

    ok = skipped = errors = 0

    for path in paths:
        rel    = os.path.relpath(path, os.path.dirname(os.path.abspath(__file__)))
        status, old_e6, new_e6 = update_report(path)

        if status == "OK":
            old_str = f"${old_e6:.2f}" if isinstance(old_e6, (int, float)) else str(old_e6)
            new_str = f"${new_e6:.2f}"
            changed = "" if old_e6 == new_e6 else f"  ({old_str} -> {new_str})"
            print(f"{'WROTE' if APPLY else 'WOULD':5s}  {new_str}{changed}   {rel}")
            ok += 1
        elif status.startswith("SKIP"):
            print(f"SKIP   {status.split(':',1)[1].strip()}   {rel}")
            skipped += 1
        else:
            print(f"ERROR  {status}   {rel}")
            errors += 1

    print()
    print(f"{'Updated' if APPLY else 'Would update'}: {ok}  |  Skipped: {skipped}  |  Errors: {errors}")
    if not APPLY and ok:
        print("Run with --apply to write changes.")


if __name__ == "__main__":
    main()
