"""
backfill_land_cost.py

Reads every Excel report in the reports/ folder, re-scrapes the Crexi listing
URL stored in cell C2, and fills in C6 (Cost of Land) if it is blank.

Usage:
    python backfill_land_cost.py           # dry run — shows what would be written
    python backfill_land_cost.py --apply   # actually update the files

Each file costs 1 Firecrawl credit. Files that already have a value in C6
are skipped automatically.
"""

import glob
import os
import re
import sys
import time

from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"))

try:
    import openpyxl
except ImportError:
    print("Missing openpyxl — run: pip install openpyxl")
    sys.exit(1)

APPLY   = "--apply" in sys.argv
REPORTS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")
FIRECRAWL_KEY = os.environ.get("FIRECRAWL_API_KEY", "")

if not FIRECRAWL_KEY:
    print("ERROR: FIRECRAWL_API_KEY not set in .env")
    sys.exit(1)


# ── Price extraction (mirrors crexi/scraper.py logic) ─────────────────────────

def _extract_price_from_markdown(content: str):
    """Return asking price float or None. Checks first 800 chars for 'Contact Broker'."""
    price_header = content[:800]
    if re.search(r"contact\s+broker|unpriced|call\s+for\s+price|price\s+upon\s+request",
                 price_header, re.IGNORECASE):
        return None

    # Pattern 1: standalone $ line
    m = re.search(
        r"(?:^|\n)\s*\*{0,2}\$\s*([\d,]+(?:\.\d+)?)\s*([MmKk]?)\*{0,2}\s*(?:\r?\n|\||\Z)",
        content
    )
    # Pattern 2: labeled same line
    if not m:
        m = re.search(
            r"\*{0,2}(?:sale\s+price|asking\s+price|list\s+price|price)\*{0,2}"
            r"[\s\*:]{0,6}\$\s*([\d,]+(?:\.\d+)?)\s*([MmKk]?)",
            content, re.IGNORECASE
        )
    # Pattern 3: label then price on next line
    if not m:
        m = re.search(
            r"(?:sale\s+price|asking\s+price|list\s+price|price)[^\n]{0,20}\n"
            r"\s*\*{0,2}\$\s*([\d,]+(?:\.\d+)?)\s*([MmKk]?)",
            content, re.IGNORECASE
        )

    if m:
        try:
            val = float(m.group(1).replace(",", ""))
            suffix = m.group(2).upper()
            if suffix == "M":
                val *= 1_000_000
            elif suffix == "K":
                val *= 1_000
            if 50_000 <= val <= 100_000_000:
                return val
        except ValueError:
            pass

    # Pattern 4: last resort — first plausible land price in first 2000 chars
    for fm in re.finditer(r"\$\s*([\d,]+(?:\.\d+)?)\s*([MmKk]?)", content[:2000]):
        try:
            val = float(fm.group(1).replace(",", ""))
            suffix = fm.group(2).upper()
            if suffix == "M":
                val *= 1_000_000
            elif suffix == "K":
                val *= 1_000
            if 50_000 <= val <= 100_000_000:
                return val
        except ValueError:
            pass

    return None


def _scrape_price(url: str) -> float | None:
    """Firecrawl the Crexi listing URL and extract the asking price."""
    try:
        from firecrawl import FirecrawlApp
    except ImportError:
        print("  ERROR: firecrawl-py not installed — run: pip install firecrawl-py")
        return None

    try:
        app = FirecrawlApp(api_key=FIRECRAWL_KEY)
        result = app.scrape(url, formats=["markdown"])
        if result and hasattr(result, "markdown") and result.markdown:
            return _extract_price_from_markdown(result.markdown)
        if result and hasattr(result, "content") and result.content:
            return _extract_price_from_markdown(result.content)
    except Exception as exc:
        print(f"  Firecrawl error: {exc}")
    return None


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    paths = sorted(glob.glob(os.path.join(REPORTS, "**", "*.xlsx"), recursive=True))
    if not paths:
        print("No .xlsx files found in reports/")
        return

    if not APPLY:
        print("DRY RUN — pass --apply to actually update files\n")

    updated = skipped_has_price = skipped_no_url = errors = 0

    for path in paths:
        rel = os.path.relpath(path, os.path.dirname(os.path.abspath(__file__)))
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as exc:
            print(f"SKIP (can't open): {rel} — {exc}")
            errors += 1
            continue

        # Find proforma sheet
        ws = None
        for name in wb.sheetnames:
            if name.strip().lower() in {"proforma", "initial look proforma", "initial proforma"}:
                ws = wb[name]
                break
        if ws is None:
            print(f"SKIP (no proforma tab): {rel}")
            errors += 1
            continue

        # Check if C6 already has a value
        c6_val = ws["C6"].value
        if c6_val is not None and str(c6_val).strip() not in ("", "0"):
            skipped_has_price += 1
            print(f"OK   (already ${float(c6_val):,.0f}): {rel}")
            continue

        # Get Crexi URL from C2
        c2_cell = ws["C2"]
        url = None
        if c2_cell.hyperlink:
            url = str(c2_cell.hyperlink.target or c2_cell.hyperlink)
        if not url and c2_cell.value:
            url = str(c2_cell.value).strip()
        if not url or not url.startswith("http"):
            print(f"SKIP (no URL in C2): {rel}")
            skipped_no_url += 1
            continue

        print(f"Scraping: {rel}")
        print(f"  URL: {url}")
        price = _scrape_price(url)

        if price is None:
            print("  -> no price found")
            errors += 1
        else:
            print(f"  -> ${price:,.0f}")
            if APPLY:
                ws["C6"] = price
                wb.save(path)
                print("  -> saved")
            updated += 1

        time.sleep(1)  # rate limit between Firecrawl calls

    print()
    print("Done.")
    print(f"  Would update / updated:  {updated}")
    print(f"  Already had price:       {skipped_has_price}")
    print(f"  No URL in C2:            {skipped_no_url}")
    print(f"  Errors:                  {errors}")
    if not APPLY and updated:
        print("\nRun with --apply to write changes.")


if __name__ == "__main__":
    main()
