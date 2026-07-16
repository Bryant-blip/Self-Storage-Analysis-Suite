"""
backfill_market_averages.py

Replaces the "total average" rows in the Market Comps averages section of every
existing Excel report with proper normalized weighted-average formulas using
UNIT_MIX_WEIGHTS, matching the logic in comps_pipeline.py.

Usage:
    python backfill_market_averages.py           # dry run — shows what would change
    python backfill_market_averages.py --apply   # writes to files
"""

import os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # run from repo root or scripts/

import glob
import os
import sys

try:
    import openpyxl
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Missing openpyxl — run: pip install openpyxl")
    sys.exit(1)

APPLY   = "--apply" in sys.argv
REPORTS = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reports")

# Must match comps_pipeline.py
UNIT_SF = {"5x5": 25, "5x10": 50, "10x10": 100, "10x15": 150,
           "10x20": 200, "10x25": 250, "10x30": 300}
UNIT_MIX_WEIGHTS = {"5x5": 0.12, "5x10": 0.25, "10x10": 0.30, "10x15": 0.15,
                    "10x20": 0.12, "10x25": 0.00, "10x30": 0.06}
UNIT_SIZES = list(UNIT_SF.keys())


def build_weighted_formula(refs_weights: list[tuple[str, float]]) -> str:
    """
    Build a normalized weighted average formula from (cell_ref, weight) pairs.
    Weights are normalized across the sizes that are actually present.
    """
    total_w = sum(w for _, w in refs_weights)
    if total_w <= 0:
        refs = [r for r, _ in refs_weights]
        return f"=AVERAGE({','.join(refs)})"
    terms = [f"{ref}*{w / total_w:.6f}" for ref, w in refs_weights if w > 0]
    return f"={'+ '.join(terms)}"


def process_panel(ws, header_row: int, label_col: int, val_col: int, apply: bool):
    """
    Find the size rows and "total average" row for one panel.
    Replaces "total average" with a weighted formula.

    Returns: (status, old_formula, new_formula)
    """
    col_letter = get_column_letter(val_col)

    # Scan rows below the header for size rows and the total row
    size_rows: list[tuple[str, int]] = []   # (size, row_num)
    total_row: int | None = None

    for r in range(header_row + 1, header_row + 30):
        label_val = ws.cell(row=r, column=label_col).value
        if label_val is None:
            continue
        label_str = str(label_val).strip()

        if label_str in UNIT_SF:
            size_rows.append((label_str, r))
        elif label_str.lower() in ("total average", "weighted average", "total"):
            total_row = r
            break

    if not size_rows:
        return "SKIP_NO_SIZES", None, None

    # Build refs+weights list for sizes that have a non-zero weight
    refs_weights = [
        (f"{col_letter}{r}", UNIT_MIX_WEIGHTS.get(size, 0.0))
        for size, r in size_rows
        if UNIT_MIX_WEIGHTS.get(size, 0.0) > 0
    ]
    if not refs_weights:
        return "SKIP_NO_WEIGHTS", None, None

    new_formula = build_weighted_formula(refs_weights)

    if total_row is None:
        # No total row found — shouldn't happen in well-formed reports
        return "SKIP_NO_TOTAL_ROW", None, new_formula

    old_formula = ws.cell(row=total_row, column=val_col).value

    # Already a weighted formula (from a newer report) — skip
    if old_formula and isinstance(old_formula, str) and "*" in old_formula:
        ws.cell(row=total_row, column=label_col).value = "weighted average"
        if apply:
            ws.cell(row=total_row, column=val_col).value = new_formula
        return "ALREADY_WEIGHTED", old_formula, new_formula

    if apply:
        ws.cell(row=total_row, column=label_col).value = "weighted average"
        ws.cell(row=total_row, column=label_col).font = Font(bold=True)
        c = ws.cell(row=total_row, column=val_col)
        c.value = new_formula
        if not c.number_format or c.number_format == "General":
            c.number_format = '"$"#,##0.00'
        c.font = Font(bold=True)

    return "OK", old_formula, new_formula


def update_report(path: str) -> tuple[str, int, int]:
    """
    Process all four panels in one report.
    Returns (status_summary, updated_count, skipped_count).
    """
    try:
        wb = openpyxl.load_workbook(path)
    except Exception as exc:
        return f"ERROR opening: {exc}", 0, 0

    comps = next(
        (wb[n] for n in wb.sheetnames if "market comps" in n.strip().lower()),
        None,
    )
    if comps is None:
        return "ERROR: no Market Comps tab", 0, 0

    # Find header rows for the two section groups:
    #   "Drive-Up Units" / "Climate Controlled" in column 2
    # Each such header row anchors two panels (col 2/3 and col 5/6)
    section_header_rows: list[int] = []
    for row in comps.iter_rows():
        val = str(row[1].value or "").strip()   # column 2 (index 1)
        if val in ("Drive-Up Units", "Climate Controlled"):
            section_header_rows.append(row[0].row)

    if not section_header_rows:
        return "ERROR: averages section not found", 0, 0

    updated = 0
    skipped = 0

    # Each header row has two panels side by side: (label_col=2, val_col=3) and (label_col=5, val_col=6)
    for hdr_row in section_header_rows:
        for label_col, val_col in [(2, 3), (5, 6)]:
            status, old_f, new_f = process_panel(comps, hdr_row, label_col, val_col, APPLY)
            if status == "OK":
                updated += 1
            elif status == "ALREADY_WEIGHTED":
                skipped += 1   # already done — count as updated if label needs fixing
            else:
                skipped += 1

    if APPLY and updated > 0:
        try:
            wb.save(path)
        except PermissionError:
            return "ERROR: file locked (close it in Excel first)", 0, 0

    return "OK", updated, skipped


def main():
    paths = sorted(
        p for p in glob.glob(os.path.join(REPORTS, "**", "*.xlsx"), recursive=True)
        if os.path.basename(p) != "deal_rankings.xlsx"
    )
    if not paths:
        print("No .xlsx files found in reports/")
        return

    if not APPLY:
        print("DRY RUN — pass --apply to write changes\n")

    ok = errors = 0
    total_updated = 0

    for path in paths:
        rel = os.path.relpath(path, os.path.dirname(os.path.abspath(__file__)))
        status, updated, skipped = update_report(path)

        if status == "OK":
            total_updated += updated
            verb = "WROTE" if APPLY else "WOULD"
            print(f"{verb:5s}  {updated} panel(s) updated   {rel}")
            ok += 1
        else:
            print(f"ERROR  {status}   {rel}")
            errors += 1

    print()
    print(f"{'Updated' if APPLY else 'Would update'}: {ok} files ({total_updated} panels)  |  Errors: {errors}")
    if not APPLY and ok:
        print("Run with --apply to write changes.")


if __name__ == "__main__":
    main()
