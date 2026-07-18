"""
migrate_existing_deals.py

One-time script that backfills the SQLite deals database (data/deals.db)
from the existing seen_deals_*.json files and corresponding Excel reports.

Usage:
    python migrate_existing_deals.py

What it does:
  - Reads all data/seen_deals_*.json and data/seen_deals.json
  - Deduplicates entries by listing_id across files
  - Inserts processed deals (reads Excel report for metrics + comps)
  - Inserts skipped deals (bare record with skip_reason)
  - Inserts pending/unprocessed deals (scraped but not yet processed)
  - Calls recalculate_scores() at the end
  - Prints per-market summary
"""

import os
import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))  # run from repo root or scripts/

import glob
import json
import os
import re as _re
import sys
from datetime import datetime, timezone

# ── Bootstrap path so db_utils can be imported ────────────────────────────────
PROJECT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, PROJECT_DIR)

try:
    from db_utils import init_db, get_db, calc_proforma_cells, _calc_yoc, recalculate_scores
except ImportError as e:
    print(f"ERROR: Could not import db_utils: {e}")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    print("Missing openpyxl — run: pip install openpyxl")
    sys.exit(1)

try:
    from comps_pipeline import UNIT_SF
except ImportError:
    # Fallback if comps_pipeline not available
    UNIT_SF = {
        "5x5": 25, "5x10": 50, "10x10": 100,
        "10x15": 150, "10x20": 200, "10x25": 250, "10x30": 300,
    }

DATA_DIR = os.path.join(PROJECT_DIR, "data")
REPORTS_DIR = os.path.join(PROJECT_DIR, "reports")


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _find_proforma_sheet(wb):
    return next(
        (wb[n] for n in wb.sheetnames
         if n.strip().lower() in {"proforma", "initial look proforma", "initial proforma"}),
        None,
    )


def _count_facility_list_rows(wb) -> int | None:
    """Count data rows in the Facility List tab (minus header)."""
    fac = next(
        (wb[n] for n in wb.sheetnames if "facility" in n.strip().lower()),
        None,
    )
    if not fac:
        return None
    count = 0
    for row in fac.iter_rows(min_row=2):
        if any(c.value is not None for c in row):
            count += 1
    return count or None


def _get_cc_weighted_avg(wb) -> float | None:
    """
    Try to read the CC Online weighted average $/sqft from the Market Comps sheet.
    Looks for the second 'weighted average' row in column 5 (E), value in column 6 (F).
    Returns None if not found or formula not evaluated.
    """
    comps = next(
        (wb[n] for n in wb.sheetnames if "market comps" in n.strip().lower()),
        None,
    )
    if not comps:
        return None

    found = 0
    for row in comps.iter_rows():
        label_cell = row[4] if len(row) > 4 else None  # col 5 (index 4)
        if label_cell and str(label_cell.value or "").strip().lower() == "weighted average":
            found += 1
            if found == 2:  # Second occurrence = CC Online
                val_cell = row[5] if len(row) > 5 else None
                if val_cell and isinstance(val_cell.value, (int, float)):
                    return float(val_cell.value)
    return None


def _resolve_report_path(report_path: str) -> str | None:
    """
    Return the actual file path for a report, handling stale paths from before
    the flat-folder refactor. Falls back to searching reports/ by filename.
    """
    if not report_path:
        return None
    if os.path.exists(report_path):
        return report_path
    # Stale path — try basename in flat reports/ directory
    basename = os.path.basename(report_path)
    alt = os.path.join(REPORTS_DIR, basename)
    return alt if os.path.exists(alt) else None


def _read_excel_metrics(wb) -> dict:
    """
    Extract financial metrics from an already-open workbook.
    Returns a dict with keys matching the deals table columns.
    """
    result = {}
    proforma = _find_proforma_sheet(wb)
    if proforma:
        cells = calc_proforma_cells(proforma)
        acres         = cells.get("acres")
        asking_price  = cells.get("asking_price")
        avg_psf       = cells.get("avg_psf")
        cost_per_sqft = cells.get("cost_per_sqft")

        result["acres"]                    = acres
        result["asking_price"]             = asking_price
        result["avg_psf"]                  = avg_psf
        result["construction_cost_per_sqft"] = cost_per_sqft
        result["avg_psf_drive_up"]         = (avg_psf + 0.05) if avg_psf is not None else None
        result["price_per_acre"]           = (asking_price / acres) if (asking_price and acres) else None
        result["yield_on_cost"]            = _calc_yoc(cells)

    result["avg_psf_climate"]    = _get_cc_weighted_avg(wb)
    result["nearby_facility_count"] = _count_facility_list_rows(wb)

    return result


def _read_comps_from_excel(wb, listing_id: str, now: str) -> list[tuple]:
    """
    Parse the Market Comps tab and Facility List tab to extract
    per-facility, per-unit-size pricing data.

    Returns list of 10-tuples ready for INSERT into the comps table:
    (listing_id, facility_name, facility_address, distance_miles,
     unit_size, unit_type, web_rate, in_store_rate, rate_per_sqft, scraped_at)
    """
    comps_ws = next(
        (wb[n] for n in wb.sheetnames if "market comps" in n.strip().lower()),
        None,
    )
    if not comps_ws:
        return []

    fac_ws = next(
        (wb[n] for n in wb.sheetnames if "facility" in n.strip().lower()),
        None,
    )

    # Build facility name (lowercase) → (address, distance_miles) from Facility List
    fac_info: dict[str, tuple] = {}
    if fac_ws:
        for row in fac_ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            name = str(row[0]).strip()
            addr = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            dist = row[2] if len(row) > 2 else None
            if name:
                fac_info[name.lower()] = (addr, dist)

    # Read all rows as value tuples
    all_rows = list(comps_ws.iter_rows(values_only=True))
    if not all_rows:
        return []

    # Find the CC section start column by scanning row 1 for "CLIMATE"
    cc_start_col: int | None = None
    row1 = all_rows[0]
    for col_idx, cell in enumerate(row1):
        if cell and "climate" in str(cell).lower():
            cc_start_col = col_idx
            break

    # Accumulator: (facility_name, unit_size, unit_type) → rate dict
    rate_map: dict[tuple, dict] = {}

    def _parse_side(side_start_col: int, side_end_col: int, unit_type: str):
        """Parse one side (drive_up or climate_control) of the Market Comps grid."""
        # Find header rows: row where cell[side_start_col] == "Sq Ft"
        # and cell[side_start_col+1] == "Size"
        header_rows = []
        for row_idx, row in enumerate(all_rows):
            if len(row) <= side_start_col + 1:
                continue
            a = str(row[side_start_col] or "").strip()
            b = str(row[side_start_col + 1] or "").strip()
            if a == "Sq Ft" and b == "Size":
                header_rows.append(row_idx)

        if not header_rows:
            return

        for header_num, header_row_idx in enumerate(header_rows):
            # First header = in-store, second = online (web rate)
            rate_key = "in_store_rate" if header_num % 2 == 0 else "web_rate"
            header = all_rows[header_row_idx]

            # Collect facility columns (cols start+2 onwards until empty/out of range)
            fac_cols = []
            for col_idx in range(side_start_col + 2, side_end_col):
                if col_idx >= len(header):
                    break
                val = header[col_idx]
                if val is None:
                    break
                s = str(val).strip()
                if not s:
                    break
                # Strip "(X.X mi)" suffix to get clean name
                dist_match = _re.search(r'\((\d+\.?\d*)\s*mi\)', s)
                inline_dist = float(dist_match.group(1)) if dist_match else None
                name = _re.sub(r'\s*\(\d+\.?\d*\s*mi\)', '', s).strip()
                fac_cols.append((col_idx, name, inline_dist))

            if not fac_cols:
                continue

            # Data rows: between this header and the next (or end of sheet)
            next_header_idx = (header_rows[header_num + 1]
                               if header_num + 1 < len(header_rows) else len(all_rows))

            for row_idx in range(header_row_idx + 1, next_header_idx):
                row = all_rows[row_idx]
                if len(row) <= side_start_col + 1:
                    continue
                size_val = row[side_start_col + 1]
                if not size_val:
                    continue
                size_str = str(size_val).strip()
                # Only rows matching "NxM" pattern (e.g. 10x10)
                if not _re.match(r'^\d+x\d+$', size_str):
                    continue

                for col_idx, fac_name, inline_dist in fac_cols:
                    if col_idx >= len(row):
                        continue
                    price = row[col_idx]
                    if price is None or not isinstance(price, (int, float)):
                        continue
                    price = float(price)

                    key = (fac_name, size_str, unit_type)
                    if key not in rate_map:
                        # Look up address & distance from Facility List
                        addr, fac_dist = fac_info.get(fac_name.lower(), ("", inline_dist))
                        rate_map[key] = {
                            "address":      addr,
                            "distance":     fac_dist if fac_dist is not None else inline_dist,
                            "web_rate":     None,
                            "in_store_rate": None,
                        }
                    rate_map[key][rate_key] = price

    # Parse Drive-Up side (cols 0 → cc_start_col or end of row)
    du_end = cc_start_col if cc_start_col is not None else len(row1)
    _parse_side(0, du_end, "drive_up")

    # Parse Climate Controlled side (if present)
    if cc_start_col is not None:
        _parse_side(cc_start_col, len(row1), "climate_control")

    # Build result tuples
    result = []
    for (fac_name, unit_size, unit_type), data in rate_map.items():
        web_rate  = data["web_rate"]
        in_store  = data["in_store_rate"]
        sqft      = UNIT_SF.get(unit_size)
        rate_psf  = (web_rate / sqft) if (web_rate and sqft) else None
        result.append((
            listing_id, fac_name, data["address"], data["distance"],
            unit_size, unit_type, web_rate, in_store, rate_psf, now,
        ))
    return result


# ── JSON loaders ──────────────────────────────────────────────────────────────

def load_all_seen_deals() -> dict:
    """
    Load and deduplicate all seen_deals_*.json and seen_deals.json.
    Returns {listing_id: entry_dict}.
    """
    combined = {}
    json_files = sorted(glob.glob(os.path.join(DATA_DIR, "seen_deals_*.json")))
    legacy     = os.path.join(DATA_DIR, "seen_deals.json")
    if os.path.exists(legacy):
        json_files.append(legacy)

    for path in json_files:
        try:
            with open(path, encoding="utf-8") as f:
                data = json.load(f)
            for lid, entry in data.items():
                if lid not in combined:
                    entry["_source_file"] = os.path.basename(path)
                    combined[lid] = entry
        except Exception as exc:
            print(f"  WARN: Could not read {path}: {exc}")

    return combined


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("=" * 60)
    print("Storage Intel — Deal Migration")
    print("=" * 60)
    print()

    init_db()
    conn = get_db()

    # Clear existing rows for a clean re-run
    conn.execute("DELETE FROM comps")
    conn.execute("DELETE FROM deals")
    conn.commit()
    print("Cleared existing DB rows.")
    print()

    all_entries = load_all_seen_deals()
    print(f"Found {len(all_entries)} total entries across all seen_deals files")
    print()

    now = datetime.now(timezone.utc).isoformat(timespec="seconds")

    # Per-market counters
    market_counts: dict[str, dict] = {}

    def _market_cnt(market):
        if market not in market_counts:
            market_counts[market] = {"inserted": 0, "skipped": 0, "pending": 0, "errors": 0}
        return market_counts[market]

    total_inserted = total_skipped = total_pending = total_errors = 0
    total_comps = 0

    for lid, entry in all_entries.items():
        market  = entry.get("market") or "Unknown"
        cnt     = _market_cnt(market)
        address = entry.get("address") or ""
        url     = entry.get("url") or ""
        first_seen = entry.get("first_seen") or now
        skip_reason = entry.get("skip_reason")
        report_path = entry.get("report_path")
        processed = entry.get("processed", False)
        zip_code  = entry.get("zip_code")
        population_3mi = entry.get("population_3mi")
        zip_pool_count = entry.get("zip_pool_count")

        display_name = (address or lid)[:55]

        try:
            if skip_reason:
                # Permanently skipped deal
                conn.execute("""
                    INSERT OR IGNORE INTO deals
                        (listing_id, market, address, zip_code, crexi_url,
                         scraped_at, skip_reason)
                    VALUES (?,?,?,?,?,?,?)
                """, (lid, market, address, zip_code, url, first_seen, skip_reason))
                cnt["skipped"] += 1
                total_skipped += 1

            elif processed:
                # Fully processed deal — open Excel for metrics + comps
                resolved_path = _resolve_report_path(report_path)
                metrics = {}
                comp_rows = []

                if resolved_path:
                    try:
                        wb = openpyxl.load_workbook(resolved_path, data_only=True)
                        metrics   = _read_excel_metrics(wb)
                        comp_rows = _read_comps_from_excel(wb, lid, now)
                    except Exception as exc:
                        print(f"    WARN: Could not read {os.path.basename(resolved_path or '')}: {exc}")
                else:
                    print(f"    WARN: Report not found for {display_name}")

                processed_at = entry.get("last_seen") or now

                conn.execute("""
                    INSERT OR IGNORE INTO deals
                        (listing_id, market, address, zip_code, crexi_url,
                         scraped_at, processed_at,
                         asking_price, acres, price_per_acre,
                         avg_psf, avg_psf_drive_up, avg_psf_climate,
                         construction_cost_per_sqft, yield_on_cost,
                         nearby_facility_count, report_path,
                         population_3mi, zip_pool_count)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                """, (
                    lid, market, address, zip_code, url,
                    first_seen, processed_at,
                    metrics.get("asking_price"),
                    metrics.get("acres"),
                    metrics.get("price_per_acre"),
                    metrics.get("avg_psf"),
                    metrics.get("avg_psf_drive_up"),
                    metrics.get("avg_psf_climate"),
                    metrics.get("construction_cost_per_sqft"),
                    metrics.get("yield_on_cost"),
                    metrics.get("nearby_facility_count"),
                    resolved_path,
                    population_3mi,
                    zip_pool_count,
                ))

                # Insert comps rows
                for comp_tuple in comp_rows:
                    conn.execute("""
                        INSERT OR IGNORE INTO comps
                            (listing_id, facility_name, facility_address,
                             distance_miles, unit_size, unit_type,
                             web_rate, in_store_rate, rate_per_sqft, scraped_at)
                        VALUES (?,?,?,?,?,?,?,?,?,?)
                    """, comp_tuple)
                    total_comps += 1

                cnt["inserted"] += 1
                total_inserted += 1
                psf_str = f"  psf=${metrics.get('avg_psf') or 0:.2f}" if metrics.get("avg_psf") else ""
                comps_str = f"  {len(comp_rows)} comps" if comp_rows else ""
                print(f"  OK  {market:15s}  {display_name}{psf_str}{comps_str}")

            else:
                # Pending / unprocessed
                conn.execute("""
                    INSERT OR IGNORE INTO deals
                        (listing_id, market, address, zip_code, crexi_url, scraped_at)
                    VALUES (?,?,?,?,?,?)
                """, (lid, market, address, zip_code, url, first_seen))
                cnt["pending"] += 1
                total_pending += 1

        except Exception as exc:
            print(f"  ERROR  {market:15s}  {display_name}: {exc}")
            cnt["errors"] += 1
            total_errors += 1

    conn.commit()

    print()
    print("Recalculating deal scores...")
    recalculate_scores(conn)
    conn.close()

    # Summary
    print()
    print("=" * 60)
    print(f"{'Market':<20} {'Inserted':>9} {'Skipped':>8} {'Pending':>8} {'Errors':>7}")
    print("-" * 60)
    for market, cnt in sorted(market_counts.items()):
        print(f"{market:<20} {cnt['inserted']:>9} {cnt['skipped']:>8} "
              f"{cnt['pending']:>8} {cnt['errors']:>7}")
    print("-" * 60)
    print(f"{'TOTAL':<20} {total_inserted:>9} {total_skipped:>8} "
          f"{total_pending:>8} {total_errors:>7}")
    print()
    print(f"Comps rows inserted: {total_comps}")
    print(f"Database: {os.path.relpath(os.path.join(DATA_DIR, 'deals.db'), PROJECT_DIR)}")
    print("Migration complete.")


if __name__ == "__main__":
    main()
