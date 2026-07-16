#!/usr/bin/env python3
"""
comps_pipeline.py — Self Storage Market Comps Pipeline

Workflow:
  1. Geocode subject location       (Google Geocoding API)
  2. Find nearby competitors        (Google Places Nearby Search)
  3. Per facility:
     a. Get address / phone / URL   (Google Place Details)
     b. Extract website content     (Tavily Extract — direct from facility site)
     c. Parse structured pricing    (Claude 3.5 Haiku)
  4. Write Excel report             (openpyxl — 2-tab format)

Cost estimate: ~$0.02–$0.05 per facility
"""

import json
import logging
import math
import os
try:
    import winreg
except ImportError:  # non-Windows platforms
    winreg = None
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed

# Firecrawl Hobby plan caps concurrent requests at 5. The outer pool runs more
# workers than that (for pipelining Google/Claude stages), so gate Firecrawl
# calls through a shared semaphore.
_FIRECRAWL_SEMAPHORE = threading.Semaphore(5)

AGGREGATOR_DOMAINS = (
    "sparefoot.com", "storagesearch.com", "selfstorageunit.com",
    "yelp.com", "yellowpages.com", "mapquest.com", "tripadvisor.com",
    "facebook.com", "google.com",
)

def _is_aggregator(url: str) -> bool:
    url_lower = (url or "").lower()
    return any(d in url_lower for d in AGGREGATOR_DOMAINS)

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


def _get_env(name: str) -> str:
    """
    Read an environment variable, falling back to the Windows User env store.
    Needed because Claude Code sets some vars (e.g. ANTHROPIC_API_KEY) to an
    empty string in the process environment, shadowing the user's real value.
    """
    val = os.environ.get(name, "")
    if val:
        return val
    if winreg is None:
        return ""
    try:
        with winreg.OpenKey(winreg.HKEY_CURRENT_USER, r"Environment") as key:
            val, _ = winreg.QueryValueEx(key, name)
            return val or ""
    except (FileNotFoundError, OSError):
        return ""


# ── Constants ──────────────────────────────────────────────────────────────────

UNIT_SIZES    = ["5x5", "5x10", "10x10", "10x15", "10x20", "10x25", "10x30"]
UNIT_SF       = {"5x5": 25, "5x10": 50, "10x10": 100, "10x15": 150, "10x20": 200, "10x25": 250, "10x30": 300}
# Unit mix weights for the weighted average $/sqft row in the averages section.
# 10x25 is set to 0 (uncommon size — excluded from weighted avg).
# Weights are normalized at runtime so missing sizes don't throw off the result.
UNIT_MIX_WEIGHTS = {"5x5": 0.12, "5x10": 0.25, "10x10": 0.30, "10x15": 0.15,
                    "10x20": 0.12, "10x25": 0.00, "10x30": 0.06}

# ── Facility-type constants ───────────────────────────────────────────────────
# Lehi, UT reference: 1.6 acres → 115K gross / 85K rentable
# Business assumptions below — override via .env / environment variables without
# touching code (defaults preserve the original hardcoded values).
_MULTI_STORY_YIELD  = float(os.environ.get("MULTI_STORY_YIELD", str(round(85_000 / (1.6 * 43_560), 4))))  # ~1.2196 (122%)
_SINGLE_STORY_YIELD = float(os.environ.get("SINGLE_STORY_YIELD", "0.40"))                                   # 40%
_MIXED_TARGET_SQFT  = float(os.environ.get("MIXED_TARGET_SQFT", "90000"))

_MULTI_STORY_COST  = float(os.environ.get("MULTI_STORY_COST", "95.0"))
_SINGLE_STORY_COST = float(os.environ.get("SINGLE_STORY_COST", "50.0"))


def classify_facility(acres: float | None) -> str:
    if acres is None:
        return "single_story"
    if acres < 2.0:
        return "multi_story"
    if acres <= 4.0:
        return "mixed"
    return "single_story"


def calc_facility_assumptions(facility_type: str, acres: float = None) -> dict:
    if facility_type == "multi_story":
        return {"yield_pct": _MULTI_STORY_YIELD, "cost_per_sqft": _MULTI_STORY_COST}

    if facility_type == "mixed" and acres:
        land_sqft = acres * 43_560
        ms_frac = (_MIXED_TARGET_SQFT / land_sqft - _SINGLE_STORY_YIELD) \
                  / (_MULTI_STORY_YIELD - _SINGLE_STORY_YIELD)
        ms_frac = max(0.0, min(1.0, ms_frac))
        ss_frac = 1.0 - ms_frac

        ms_sqft = ms_frac * land_sqft * _MULTI_STORY_YIELD
        ss_sqft = ss_frac * land_sqft * _SINGLE_STORY_YIELD
        total_sqft = ms_sqft + ss_sqft
        total_cost = ms_sqft * _MULTI_STORY_COST + ss_sqft * _SINGLE_STORY_COST

        eff_yield = total_sqft / land_sqft
        eff_cost = total_cost / total_sqft if total_sqft else _SINGLE_STORY_COST

        return {
            "yield_pct": round(eff_yield, 4),
            "cost_per_sqft": round(eff_cost, 2),
            "ms_frac": round(ms_frac, 4),
            "ss_frac": round(ss_frac, 4),
            "ms_sqft": round(ms_sqft),
            "ss_sqft": round(ss_sqft),
        }

    return {"yield_pct": _SINGLE_STORY_YIELD, "cost_per_sqft": _SINGLE_STORY_COST}


DRIVE_MPH     = 25.0
TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "claude excel model template.xlsx")
MIXED_TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                                   "mixed_proforma_template.xlsx")
ORANGE_HEX = "FCE4D6"
GREEN_HEX  = "E2EFDA"

EXTRACTION_SYSTEM = """You are a data extraction assistant for self-storage pricing.
You will receive search result snippets about a specific storage facility.
Extract any unit pricing you find. Return ONLY a valid JSON array — no explanation, no markdown.

Each element must have these fields:
  {"size": "10x10", "type": "drive_up", "in_store_rate": 129.00, "web_rate": 99.00}

Size rules:
- Normalize all size formats to NxN: "10 x 10", "10X10", "10'x10'" all become "10x10"
- Only include these sizes: 5x5, 5x10, 10x10, 10x15, 10x20, 10x25, 10x30
- Skip any other size
- Only extract prices for units explicitly shown as AVAILABLE at this specific location.
  Do NOT extract prices from a general size guide, "all locations" pricing table, or
  "how to choose a unit" reference page. If the page is a national price list rather
  than inventory for one specific location, return [].

Type rules:
- The "type" field must be EXACTLY one of two values: "drive_up" or "climate_control"
- "climate_control": unit is explicitly described as climate controlled, air conditioned,
  temperature controlled, heated and cooled, or AC. Interior/indoor alone does NOT mean
  climate controlled — only use this if there is explicit temperature control language.
- "drive_up": everything else — outdoor, drive-up, walk-up, standard, ground floor,
  non-climate, unconditioned, or indoor/interior without temperature control language.
- Do NOT use any other value for type. Default to "drive_up" if unclear.

Price rules:
- web_rate: the lower online/discounted price. Labels: "Web Rate", "Online Rate", "Online-Only Price",
  "Reserve Online", "Internet Price", "eRate", or any price explicitly marked as discounted/online
- in_store_rate: the higher regular/walk-in price. Labels: "Standard Price", "Street Rate",
  "Regular Price", "Regular Rate", "In-Store Rate", "In-Store Price", or any price shown
  crossed out / struck through next to a lower price
- The in_store_rate is ALWAYS higher than web_rate for the same unit. If you see two prices
  and one is crossed out, the crossed-out one is in_store_rate.
- If only one price is mentioned, put it in web_rate and set in_store_rate to null
- Extract dollar amounts from text like "$89/mo", "starting at $89", "from $89", "$89.00/month"
- Never invent prices — only extract prices explicitly stated
- Do NOT extract prices from size guide pages, example pricing, or "starting at" estimates
  that don't correspond to actual available inventory at this location

Return [] if no prices are found. One object per (size, type) combination."""


# ── Google APIs ────────────────────────────────────────────────────────────────

def _geocode(address: str, api_key: str) -> tuple:
    """Convert an address to (lat, lng) using Google Geocoding API."""
    r = requests.get(
        "https://maps.googleapis.com/maps/api/geocode/json",
        params={"address": address, "key": api_key},
        timeout=10,
    )
    r.raise_for_status()
    data = r.json()
    if data["status"] != "OK" or not data.get("results"):
        raise ValueError(
            f"Could not geocode '{address}' — Google status: {data.get('status')}"
        )
    loc = data["results"][0]["geometry"]["location"]
    return loc["lat"], loc["lng"]


def _nearby_search(lat: float, lng: float, radius_m: int, api_key: str) -> list:
    """Return up to 40 self-storage facilities from Google Places Nearby Search."""
    url = "https://maps.googleapis.com/maps/api/place/nearbysearch/json"
    params = {
        "location": f"{lat},{lng}",
        "radius": radius_m,
        "keyword": "self storage",
        "key": api_key,
    }
    places = []
    while True:
        r = requests.get(url, params=params, timeout=10)
        r.raise_for_status()
        data = r.json()
        status = data.get("status")
        if status not in ("OK", "ZERO_RESULTS"):
            error_msg = data.get("error_message", "no details provided")
            raise ValueError(
                f"Google Places API error: {status} — {error_msg}\n"
                "Check that the Places API is enabled on your Google Cloud project."
            )
        for p in data.get("results", []):
            geo = p["geometry"]["location"]
            places.append({
                "place_id": p["place_id"],
                "name":     p["name"],
                "vicinity": p.get("vicinity", ""),
                "lat":      geo["lat"],
                "lng":      geo["lng"],
            })
        token = data.get("next_page_token")
        if not token or len(places) >= 40:
            break
        time.sleep(2)  # Google requires a short pause before using next_page_token
        params = {"pagetoken": token, "key": api_key}
    return places


def _text_search(lat: float, lng: float, radius_m: int, location_hint: str, api_key: str) -> list:
    """Fallback: Google Places Text Search, more flexible than Nearby Search."""
    r = requests.get(
        "https://maps.googleapis.com/maps/api/place/textsearch/json",
        params={
            "query":    f"self storage near {location_hint}",
            "location": f"{lat},{lng}",
            "radius":   radius_m,
            "key":      api_key,
        },
        timeout=10,
    )
    r.raise_for_status()
    data = r.json()
    status = data.get("status")
    if status not in ("OK", "ZERO_RESULTS"):
        error_msg = data.get("error_message", "no details provided")
        raise ValueError(
            f"Google Places Text Search error: {status} — {error_msg}\n"
            "Check that the Places API is enabled on your Google Cloud project."
        )
    places = []
    for p in data.get("results", []):
        geo = p["geometry"]["location"]
        places.append({
            "place_id": p["place_id"],
            "name":     p["name"],
            "vicinity": p.get("formatted_address", p.get("vicinity", "")),
            "lat":      geo["lat"],
            "lng":      geo["lng"],
        })
    return places


def _place_details(place_id: str, api_key: str) -> dict:
    """Fetch address, phone, and website for a single place."""
    r = requests.get(
        "https://maps.googleapis.com/maps/api/place/details/json",
        params={
            "place_id": place_id,
            "fields":   "formatted_address,formatted_phone_number,website",
            "key":      api_key,
        },
        timeout=10,
    )
    r.raise_for_status()
    result = r.json().get("result", {})
    return {
        "address": result.get("formatted_address", ""),
        "phone":   result.get("formatted_phone_number", ""),
        "website": result.get("website", ""),
    }


# ── Distance ───────────────────────────────────────────────────────────────────

def _distance(lat1: float, lng1: float, lat2: float, lng2: float) -> tuple:
    """Return (distance_miles, drive_time_minutes) using Haversine + 25 mph."""
    R = 3958.8
    dlat = math.radians(lat2 - lat1)
    dlng = math.radians(lng2 - lng1)
    a = (
        math.sin(dlat / 2) ** 2
        + math.cos(math.radians(lat1))
        * math.cos(math.radians(lat2))
        * math.sin(dlng / 2) ** 2
    )
    miles   = R * 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    minutes = int(round((miles / DRIVE_MPH) * 60))
    return round(miles, 2), minutes


# ── Firecrawl Pricing Scrape ───────────────────────────────────────────────────

def _firecrawl_scrape_pricing(website_url: str, api_key: str) -> str:
    """
    Scrape a facility's own website for pricing using Firecrawl.
    Firecrawl handles JavaScript rendering and Cloudflare challenges automatically.
    Uses the website URL from Google Places — direct source, no aggregator middleman.
    Returns Markdown content, or empty string if no website or scrape fails.
    """
    if not website_url or not api_key:
        return ""
    from firecrawl import FirecrawlApp
    app = FirecrawlApp(api_key=api_key)
    last_exc = None
    for attempt in range(2):  # initial try + 1 retry
        try:
            with _FIRECRAWL_SEMAPHORE:
                result = app.scrape(website_url, formats=["markdown"])
            if result and hasattr(result, "markdown") and result.markdown:
                return result.markdown
            if result and hasattr(result, "content") and result.content:
                return result.content
            logger.warning("Firecrawl returned no content for %s (result type: %s)", website_url, type(result).__name__)
            return ""
        except Exception as exc:
            last_exc = exc
            if attempt == 0:
                logger.info("Firecrawl scrape error for %s (%s) — retrying in 2.5s", website_url, exc)
                time.sleep(2.5)
    logger.warning("Firecrawl scrape failed for %s after retry: %s", website_url, last_exc)
    return ""


def _firecrawl_search_pricing(facility_name: str, address: str, api_key: str) -> tuple[str, str]:
    """
    Fallback: when Google Places has no website URL, search for the facility's
    pricing page using Firecrawl's search endpoint, then scrape the top result.

    Returns (content, url_used) — content is empty string if search/scrape fails.
    Skips aggregator domains (sparefoot, storagesearch, etc.) to keep pricing direct.
    """
    if not facility_name or not api_key:
        return "", ""

    query = f"{facility_name} self storage unit prices rates"
    if address:
        # Add city/state from address to narrow search
        parts = address.split(",")
        if len(parts) >= 2:
            query += f" {parts[-2].strip()} {parts[-1].strip()}"

    try:
        from firecrawl import FirecrawlApp
        app = FirecrawlApp(api_key=api_key)
        with _FIRECRAWL_SEMAPHORE:
            search_results = app.search(query, limit=5)
        # SDK v4.21.1: SearchData object — results are in .web, not .data
        results_list = []
        if hasattr(search_results, "web") and search_results.web:
            results_list = search_results.web
        elif isinstance(search_results, list):
            results_list = search_results

        for item in results_list:
            url = ""
            if hasattr(item, "url"):
                url = item.url or ""
            elif isinstance(item, dict):
                url = item.get("url", "")

            if not url:
                continue
            if _is_aggregator(url):
                logger.debug("Search fallback: skipping aggregator %s", url)
                continue

            logger.info("Search fallback: scraping %s for '%s'", url, facility_name)
            content = _firecrawl_scrape_pricing(url, api_key)
            if content:
                return content, url

    except Exception as exc:
        logger.warning("Firecrawl search fallback failed for '%s': %s", facility_name, exc)

    return "", ""


# ── Claude 3.5 Haiku Pricing Extraction ───────────────────────────────────────

def _extract_pricing(raw_text: str, facility_name: str, api_key: str) -> list:
    """
    Use Claude 3.5 Haiku to parse raw website text into structured pricing JSON.
    Returns a list of dicts: [{size, type, in_store_rate, web_rate}, ...]
    """
    if not raw_text.strip():
        return []
    try:
        import anthropic
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=2048,
            system=EXTRACTION_SYSTEM,
            messages=[{
                "role":    "user",
                "content": f"Facility: {facility_name}\n\nWebsite content:\n{raw_text[:50000]}",
            }],
        )
        text = resp.content[0].text.strip()
        logger.debug("Claude raw response for %s: %s", facility_name, text[:200])
        # Strip markdown code fences if present
        if "```" in text:
            for part in text.split("```"):
                part = part.strip().lstrip("json").strip()
                try:
                    parsed = json.loads(part)
                    if isinstance(parsed, list):
                        return _normalize_pricing(parsed)
                except (json.JSONDecodeError, ValueError):
                    continue
        parsed = json.loads(text)
        return _normalize_pricing(parsed) if isinstance(parsed, list) else []
    except json.JSONDecodeError as exc:
        logger.warning("Claude returned non-JSON for %s: %s", facility_name, exc)
        return []
    except Exception as exc:
        logger.warning("Pricing extraction failed for %s: %s", facility_name, exc)
        return []


def _normalize_pricing(entries: list) -> list:
    """
    Normalize Claude's output:
    - Convert type strings → "drive_up" or "climate_control"
    - Convert "$129.00" strings → 129.00 floats
    - Filter to valid sizes only
    """
    def parse_rate(val):
        if val is None:
            return None
        try:
            return float(str(val).replace("$", "").replace(",", "").strip())
        except (ValueError, TypeError):
            return None

    def normalize_type(t: str) -> str:
        t_lower = str(t).lower()
        if any(w in t_lower for w in ("climate", "air condition", "temperature", "heated", "cooled", "ac ", "a/c")):
            return "climate_control"
        return "drive_up"

    # Parse all valid entries
    parsed = []
    for e in entries:
        size = str(e.get("size", "")).strip().lower().replace(" ", "").replace("'", "").replace('"', "").replace("ft", "")
        if size not in UNIT_SIZES:
            continue
        parsed.append({
            "size":          size,
            "type":          normalize_type(e.get("type", "")),
            "in_store_rate": parse_rate(e.get("in_store_rate")),
            "web_rate":      parse_rate(e.get("web_rate")),
        })

    # Deduplicate: for each (size, type) keep the entry with the lowest web_rate
    best = {}
    for e in parsed:
        key = (e["size"], e["type"])
        existing = best.get(key)
        if existing is None:
            best[key] = e
        else:
            # Prefer the entry with the lower web_rate
            e_web      = e["web_rate"] if e["web_rate"] is not None else float("inf")
            exist_web  = existing["web_rate"] if existing["web_rate"] is not None else float("inf")
            if e_web < exist_web:
                best[key] = e

    return list(best.values())


# ── Per-Facility Processing ────────────────────────────────────────────────────

def _process_facility(facility: dict, subj_lat: float, subj_lng: float, api_keys: dict) -> dict:
    """Run the full pipeline for a single facility (details → content → parse)."""

    # 1. Get full address, phone, website from Google Place Details
    try:
        details = _place_details(facility["place_id"], api_keys["google"])
        facility.update(details)
    except Exception:
        facility.setdefault("address", facility.get("vicinity", ""))
        facility.setdefault("phone", "")
        facility.setdefault("website", "")

    # 2. Calculate distance from subject property
    facility["distance_miles"], facility["drive_time_min"] = _distance(
        subj_lat, subj_lng, facility["lat"], facility["lng"]
    )

    # 3. Scrape facility's own website for pricing (URL from Google Places).
    # Skip aggregator URLs (Yelp, SpareFoot, etc.) — they won't yield clean
    # pricing; go straight to the search fallback instead.
    website = facility.get("website", "")
    if website and _is_aggregator(website):
        logger.debug("Skipping aggregator website for %s: %s", facility.get("name"), website)
        website = ""
    raw_text = _firecrawl_scrape_pricing(website, api_keys["firecrawl"]) if website else ""

    # 3b. Fallback: if no website URL or scrape returned empty, search by name
    if not raw_text:
        raw_text, found_url = _firecrawl_search_pricing(
            facility["name"], facility.get("address", ""), api_keys["firecrawl"]
        )
        if raw_text and found_url and not website:
            facility["website"] = found_url  # use the found URL in Facility List tab

    facility["_content_length"] = len(raw_text)

    # 4. Extract structured pricing with Claude
    facility["pricing"] = (
        _extract_pricing(raw_text, facility["name"], api_keys["anthropic"])
        if raw_text else []
    )

    return facility


# ── Excel Writer ───────────────────────────────────────────────────────────────

def _write_side(ws, start_col: int, facilities: list, fill_hex: str, label: str, unit_type: str) -> int:
    """
    Write one side (Drive-Up or Climate Controlled) of the Market Comps tab.
    Returns the next available column index (after a 1-column gap).
    """
    fill  = PatternFill("solid", fgColor=fill_hex)
    bold  = Font(bold=True)
    n_fac = len(facilities)
    end_col = start_col + 1 + n_fac  # Sq Ft + Size + N facility cols

    row = 1

    # Section header spanning all columns
    cell = ws.cell(row=row, column=start_col, value=label)
    cell.font = bold
    cell.fill = fill
    if n_fac > 0:
        ws.merge_cells(
            start_row=row, start_column=start_col,
            end_row=row,   end_column=end_col,
        )
    row += 1

    cell_map = {}  # {(size, rate_key): [cell_ref, ...]} — for formula building in averages

    def write_rate_block(section_label: str, rate_key: str):
        nonlocal row
        # Sub-header
        ws.cell(row=row, column=start_col, value=section_label).font = bold
        row += 1
        # Column headers
        ws.cell(row=row, column=start_col,   value="Sq Ft").font = bold
        ws.cell(row=row, column=start_col+1, value="Size").font  = bold
        for i, f in enumerate(facilities):
            dist = f.get("distance_miles")
            label = f["name"] if dist is None else f"{f['name']} ({dist:.1f} mi)"
            c = ws.cell(row=row, column=start_col + 2 + i, value=label)
            c.font      = bold
            c.fill      = fill
            c.alignment = Alignment(wrap_text=True, horizontal="center")
        row += 1
        # Data rows — one per unit size (only if at least one facility has data)
        for size in UNIT_SIZES:
            has_data = any(
                p.get(rate_key) is not None
                for f in facilities
                for p in f.get("pricing", [])
                if p.get("size") == size and p.get("type") == unit_type
            )
            if not has_data:
                continue
            ws.cell(row=row, column=start_col,   value=UNIT_SF[size])
            ws.cell(row=row, column=start_col+1, value=size)
            for i, f in enumerate(facilities):
                candidates = [
                    p.get(rate_key)
                    for p in f.get("pricing", [])
                    if p.get("size") == size
                    and p.get("type") == unit_type
                    and p.get(rate_key) is not None
                ]
                price = min(candidates) if candidates else None
                if price is not None:
                    col_idx = start_col + 2 + i
                    c = ws.cell(row=row, column=col_idx, value=float(price))
                    c.number_format = '"$"#,##0.00'
                    # Track cell address for averages formulas
                    ref = f"{get_column_letter(col_idx)}{row}"
                    cell_map.setdefault((size, rate_key), []).append(ref)
            row += 1
        row += 1  # blank separator between rate blocks

    write_rate_block("In-Store Rates",           "in_store_rate")
    write_rate_block("Online (Discounted) Rates", "web_rate")

    return end_col + 2, row, cell_map  # next available col, last row written, cell map


def _write_averages_section(ws, cell_map_du: dict, cell_map_cc: dict, start_row: int) -> None:
    """
    Write the market averages summary below the comps grid.
    Four panels: Drive-Up In-Store | Drive-Up Online | CC In-Store | CC Online
    Each cell is a live =AVERAGE(...)/sqft formula pointing back to the price grid above.
    """
    orange_fill = PatternFill("solid", fgColor=ORANGE_HEX)
    green_fill  = PatternFill("solid", fgColor=GREEN_HEX)
    bold        = Font(bold=True)
    pct_fmt     = '"$"#,##0.00'

    def make_avg_formula(cell_map: dict, size: str, rate_key: str) -> str | None:
        """Build =AVERAGE(ref1,ref2,...)/sqft formula for the given size+rate combination."""
        refs = cell_map.get((size, rate_key), [])
        if not refs:
            return None
        sf = UNIT_SF[size]
        return f"=AVERAGE({','.join(refs)})/{sf}"

    def write_panel(title: str, subtitle: str, cell_map: dict, rate_key: str, hdr_fill,
                    label_col: int, val_col: int, base_row: int) -> int:
        """Write one panel (title + per-size formula rows + total). Returns next row after panel."""
        r = base_row
        # Header row
        hdr = ws.cell(row=r, column=label_col, value=title)
        hdr.font = bold
        hdr.fill = hdr_fill
        sub = ws.cell(row=r, column=val_col, value=subtitle)
        sub.font = bold
        sub.fill = hdr_fill
        r += 1
        # Sub-label
        ws.cell(row=r, column=val_col, value="average cost per SqFt").font = Font(italic=True)
        r += 1
        # Per-size rows — live formulas
        size_avg_refs = []
        size_avg_weights = []
        for size in UNIT_SIZES:
            formula = make_avg_formula(cell_map, size, rate_key)
            if formula is None:
                continue
            ws.cell(row=r, column=label_col, value=size)
            c = ws.cell(row=r, column=val_col, value=formula)
            c.number_format = pct_fmt
            size_avg_refs.append(f"{get_column_letter(val_col)}{r}")
            size_avg_weights.append(UNIT_MIX_WEIGHTS.get(size, 0.0))
            r += 1
        # Weighted average — normalize weights across whichever sizes have data
        if size_avg_refs:
            total_w = sum(size_avg_weights)
            if total_w > 0:
                terms = [f"{ref}*{w / total_w:.6f}"
                         for ref, w in zip(size_avg_refs, size_avg_weights)
                         if w > 0]
                weighted_formula = f"={'+ '.join(terms)}" if terms else f"=AVERAGE({','.join(size_avg_refs)})"
            else:
                weighted_formula = f"=AVERAGE({','.join(size_avg_refs)})"
            ws.cell(row=r, column=label_col, value="weighted average").font = bold
            c = ws.cell(row=r, column=val_col, value=weighted_formula)
            c.number_format = pct_fmt
            c.font = bold
            r += 1
        return r + 1  # blank gap after panel

    # Drive-Up row: left panel (cols 2-3) and right panel (cols 5-6) side by side
    next_du = write_panel("Drive-Up Units", "In-Store rates",
                          cell_map_du, "in_store_rate", orange_fill, 2, 3, start_row)
    write_panel(          "Drive-Up Units", "Online rates (discounted)",
                          cell_map_du, "web_rate",      orange_fill, 5, 6, start_row)

    # Climate Controlled row: left panel and right panel side by side
    write_panel("Climate Controlled", "In-Store rates",
                cell_map_cc, "in_store_rate", green_fill, 2, 3, next_du)
    write_panel("Climate Controlled", "Online rates (discounted)",
                cell_map_cc, "web_rate",      green_fill, 5, 6, next_du)


def _calc_avg_rent_per_sqft(facilities: list):
    """Average web_rate/sqft across all valid pricing entries from all facilities."""
    rates = []
    for f in facilities:
        for p in f.get("pricing", []):
            web = p.get("web_rate")
            sf  = UNIT_SF.get(p.get("size", ""))
            if web and sf:
                rates.append(web / sf)
    return round(sum(rates) / len(rates), 2) if rates else None


def _calc_weighted_rent_per_sqft(facilities: list, unit_type: str) -> float | None:
    """
    Calculate the weighted average in-store $/sqft (falling back to web_rate when
    in_store_rate isn't published) for the given unit_type, using UNIT_MIX_WEIGHTS.
    For each size, averages the rate/sqft across all matching facilities that have
    that size, then blends by weight. Weights are normalized to account for sizes
    with no data. Returns None if no matching pricing data is available.
    """
    size_rates: dict[str, list[float]] = {s: [] for s in UNIT_SIZES}
    for f in facilities:
        for p in f.get("pricing", []):
            if p.get("unit_type", p.get("type", "")) != unit_type:
                continue
            size = p.get("size", "")
            rate = p.get("in_store_rate") or p.get("web_rate")
            sf   = UNIT_SF.get(size)
            if rate and sf and size in size_rates:
                size_rates[size].append(rate / sf)

    # Compute per-size averages and pair with weights
    size_avgs: list[tuple[float, float]] = []  # (avg_rate_per_sqft, weight)
    for size, rates in size_rates.items():
        if not rates:
            continue
        w = UNIT_MIX_WEIGHTS.get(size, 0.0)
        if w > 0:
            size_avgs.append((sum(rates) / len(rates), w))

    if not size_avgs:
        return None

    total_w = sum(w for _, w in size_avgs)
    weighted = sum(avg * (w / total_w) for avg, w in size_avgs)
    return round(weighted, 2)


def _load_proforma_from_template(
    location: str,
    acres: float = None,
    asking_price: float = None,
    crexi_url: str = "",
    rent_per_sqft: float = None,
    yield_pct: float = None,
    cost_per_sqft: float = None,
    facility_type: str = None,
):
    """
    Load the Excel template and return a workbook with the proforma tab populated.
    Clears hardcoded input cells so user fills them in fresh each run.
    All formatting, colors, borders, and formulas come directly from the template file.

    Uses a case-insensitive search for the proforma tab so a minor rename in
    the template (e.g. capitalisation change) never breaks the pipeline.
    """
    template = MIXED_TEMPLATE_PATH if facility_type == "mixed" else TEMPLATE_PATH
    wb = openpyxl.load_workbook(template)

    # Case-insensitive lookup for the proforma sheet
    PROFORMA_NAMES = {"initial look proforma", "proforma", "initial proforma"}
    proforma_sheet = next(
        (wb[name] for name in wb.sheetnames
         if name.strip().lower() in PROFORMA_NAMES),
        None,
    )
    if proforma_sheet is None:
        available = ", ".join(f'"{s}"' for s in wb.sheetnames)
        raise KeyError(
            f"Could not find a proforma tab in '{TEMPLATE_PATH}'.\n"
            f"  Available sheets: {available}\n"
            f"  Expected one of: {sorted(PROFORMA_NAMES)}\n"
            "  Please rename the proforma tab in the template file to match one of the expected names."
        )
    ws = proforma_sheet
    ws.title = "Proforma"
    ws["B3"] = location or ""
    ws["C5"] = acres
    ws["C6"] = asking_price
    if facility_type != "mixed":
        ws["E6"] = rent_per_sqft
        if yield_pct is not None:
            ws["E5"] = yield_pct
        if cost_per_sqft is not None:
            ws["E10"] = cost_per_sqft
    if facility_type:
        if facility_type != "mixed":
            ws["D3"] = "Facility Type"
            ws["E3"] = facility_type
    if crexi_url:
        ws["C2"] = crexi_url
        ws["C2"].hyperlink = crexi_url
        ws["C2"].style = "Hyperlink"
    else:
        ws["C2"] = None
        ws["C2"].hyperlink = None

    ws.sheet_view.showGridLines = False
    return wb


def _write_mixed_breakdown(ws, assumptions: dict, cc_rent: float | None, du_rent: float | None) -> None:
    """
    Fill dynamic values into the mixed proforma template.
    All layout, labels, formulas, and formatting come from mixed_proforma_template.xlsx.
    Only the per-deal values need to be written. If a rate is missing (None), the
    corresponding cell is left untouched/blank — never guessed or substituted.
    """
    ws["F12"] = f"{assumptions['ms_frac']:.0%} multi-story / {assumptions['ss_frac']:.0%} single-story"
    ws["B15"] = assumptions["ms_sqft"]
    if cc_rent is not None:
        ws["D15"] = cc_rent
    ws["D18"] = _MULTI_STORY_COST
    ws["B24"] = assumptions["ss_sqft"]
    if du_rent is not None:
        ws["D24"] = du_rent
    ws["D27"] = _SINGLE_STORY_COST


def write_comps_excel(facilities: list, output_path: str, location: str = "",
                      acres: float = None, asking_price: float = None,
                      crexi_url: str = "") -> None:
    """Write the 3-tab comps Excel file to output_path."""
    # ── Tab 1: Proforma — loaded from template ─────────────────────────────────
    facility_type = classify_facility(acres)
    assumptions = calc_facility_assumptions(facility_type, acres=acres)

    # Compute both weighted rates once — reused across the main proforma cell
    # and (for mixed) the CC/DU breakdown rows, instead of recomputing 2-3x.
    cc_rate = _calc_weighted_rent_per_sqft(facilities, "climate_control")
    du_rate = _calc_weighted_rent_per_sqft(facilities, "drive_up")

    if facility_type == "multi_story":
        raw_rate = cc_rate if cc_rate is not None else du_rate
    elif facility_type == "mixed":
        ms_frac = assumptions.get("ms_frac", 0.5)
        if cc_rate is not None and du_rate is not None:
            raw_rate = round(ms_frac * cc_rate + (1 - ms_frac) * du_rate, 2)
        else:
            raw_rate = cc_rate if cc_rate is not None else du_rate
    else:
        raw_rate = du_rate

    rent_per_sqft = round(raw_rate - 0.05, 2) if raw_rate is not None else None
    wb = _load_proforma_from_template(location, acres=acres,
                                      asking_price=asking_price, crexi_url=crexi_url,
                                      rent_per_sqft=rent_per_sqft,
                                      yield_pct=assumptions["yield_pct"],
                                      cost_per_sqft=assumptions["cost_per_sqft"],
                                      facility_type=facility_type)

    # For mixed: write breakdown rows showing the land split. Missing rates are
    # passed through as None — never fabricated or substituted with the other
    # type's blended rent (repo rule: leave the cell blank, never guess).
    if facility_type == "mixed" and "ms_frac" in assumptions:
        cc_rent = round(cc_rate - 0.05, 2) if cc_rate is not None else None
        du_rent = round(du_rate - 0.05, 2) if du_rate is not None else None
        proforma_ws = wb["Proforma"]
        _write_mixed_breakdown(proforma_ws, assumptions, cc_rent, du_rent)

    # Remove any existing Market Comps / Facility List sheets from the template
    for name in ["Market Comps", "Facility List"]:
        if name in wb.sheetnames:
            del wb[name]

    # ── Tab 2: Market Comps ────────────────────────────────────────────────────
    ws1 = wb.create_sheet("Market Comps")
    ws1.sheet_view.showGridLines = False

    du_facs = [f for f in facilities if any(p.get("type") == "drive_up"       for p in f.get("pricing", []))]
    cc_facs = [f for f in facilities if any(p.get("type") == "climate_control" for p in f.get("pricing", []))]

    next_col, last_row_du, cell_map_du = _write_side(ws1, 1,         du_facs, ORANGE_HEX, "DRIVE-UP / STANDARD UNITS", "drive_up")
    _,         last_row_cc, cell_map_cc = _write_side(ws1, next_col, cc_facs, GREEN_HEX,  "CLIMATE CONTROLLED UNITS",  "climate_control")

    avgs_start_row = max(last_row_du, last_row_cc) + 2
    _write_averages_section(ws1, cell_map_du, cell_map_cc, avgs_start_row)

    for col in ws1.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws1.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 32)

    # ── Tab 2: Facility List ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Facility List")
    ws2.sheet_view.showGridLines = False

    header_fill = PatternFill("solid", fgColor=ORANGE_HEX)
    headers = ["Facility Name", "Address", "Distance (mi)", "Drive Time (min)", "Phone", "Website"]
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = Font(bold=True)
        c.fill = header_fill

    for row, f in enumerate(facilities, 2):
        ws2.cell(row=row, column=1, value=f.get("name", ""))
        ws2.cell(row=row, column=2, value=f.get("address") or f.get("vicinity", ""))
        ws2.cell(row=row, column=3, value=f.get("distance_miles"))
        ws2.cell(row=row, column=4, value=f.get("drive_time_min"))
        ws2.cell(row=row, column=5, value=f.get("phone", ""))
        website = f.get("website", "")
        if website:
            c = ws2.cell(row=row, column=6, value=website)
            c.hyperlink = website
            c.style = "Hyperlink"
        else:
            ws2.cell(row=row, column=6, value="No website")

    for col in ws2.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(c.value)) for c in col if c.value), default=0)
        ws2.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 55)

    wb.save(output_path)


# ── Main Pipeline Entry Point ──────────────────────────────────────────────────

def run_comps_pipeline(
    location: str,
    radius_miles: float,
    output_path: str,
    api_keys: dict,
    progress_cb=None,
    stop_flag=None,
    acres: float = None,
    asking_price: float = None,
    crexi_url: str = "",
) -> None:
    """
    Run the full comps pipeline from location string to Excel file.

    Args:
        location:     Subject property address, city/state, or zip code
        radius_miles: Search radius in miles
        output_path:  Destination .xlsx file path
        api_keys:     {"google": "...", "firecrawl": "...", "anthropic": "..."}
        progress_cb:  Optional callable(pct: float | None, msg: str)
        stop_flag:    Optional callable() -> bool; return True to cancel
    """

    def emit(pct, msg):
        if progress_cb:
            progress_cb(pct, msg)

    def stopped():
        return bool(stop_flag and stop_flag())

    # Step 1 — Geocode subject location
    emit(5, f"Geocoding '{location}'...")
    subj_lat, subj_lng = _geocode(location, api_keys["google"])
    emit(10, f"Subject coordinates: {subj_lat:.5f}, {subj_lng:.5f}")

    if stopped():
        return

    # Step 2 — Find nearby facilities
    emit(12, "Searching Google Places for nearby self-storage facilities...")
    radius_m = int(radius_miles * 1609.34)
    places = _nearby_search(subj_lat, subj_lng, radius_m, api_keys["google"])

    if not places:
        emit(None, "Nearby Search returned no results — trying Text Search...")
        places = _text_search(subj_lat, subj_lng, radius_m, location, api_keys["google"])

    if not places:
        raise ValueError(
            f"No self-storage facilities found within {radius_miles} miles of '{location}'."
        )

    # Filter out non-self-storage businesses that Google Places sometimes returns
    _EXCLUDE = ("pods", "moving & storage", "moving and storage", "u-haul truck")
    places = [p for p in places if not any(ex in p["name"].lower() for ex in _EXCLUDE)]

    # Hard distance filter — Google Places radius is a hint, not a guarantee
    places = [
        p for p in places
        if _distance(subj_lat, subj_lng, p["lat"], p["lng"])[0] <= radius_miles
    ]

    emit(18, f"Found {len(places)} facilities — fetching pricing data...")

    if stopped():
        return

    # Step 3 — Process facilities in parallel (details + extract + parse)
    results = []
    n = len(places)
    completed = 0

    with ThreadPoolExecutor(max_workers=10) as executor:
        future_map = {
            executor.submit(_process_facility, place, subj_lat, subj_lng, api_keys): place["name"]
            for place in places
        }
        for future in as_completed(future_map):
            if stopped():
                break
            name = future_map[future]
            completed += 1
            try:
                fac = future.result()
                results.append(fac)
                n_prices = len(fac.get("pricing", []))
                chars    = fac.get("_content_length", 0)
                if n_prices:
                    status = f"{n_prices} unit prices found"
                elif chars > 0:
                    status = f"content fetched ({chars} chars) but no pricing extracted"
                else:
                    status = "no website found (Google Places + search fallback) — pricing skipped"
                pct = 18 + int((completed / n) * 72)
                emit(pct, f"  [{completed}/{n}] {name} — {status}")
            except Exception as exc:
                emit(None, f"  [{completed}/{n}] {name} — error: {exc}")

    if stopped():
        return

    # Sort by distance (closest first)
    results.sort(key=lambda f: f.get("distance_miles") or 9999)

    # Step 4 — Write Excel report
    emit(92, "Writing Excel report...")
    write_comps_excel(results, output_path, location=location,
                      acres=acres, asking_price=asking_price, crexi_url=crexi_url)
    emit(100, f"Saved: {os.path.basename(output_path)}")
    return output_path, results
