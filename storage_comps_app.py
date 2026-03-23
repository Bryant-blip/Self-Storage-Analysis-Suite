#!/usr/bin/env python3
"""
Self Storage Tools — Desktop App

Tabs:
  1. Market Comps  — AI-powered rent comp analysis
  2. Cost Estimator — Construction cost estimate by city
"""

import os
import sys
import threading
import asyncio
import subprocess
from datetime import date
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox

import anyio
from claude_agent_sdk import (
    query,
    ClaudeAgentOptions,
    ResultMessage,
    AssistantMessage,
    TextBlock,
)

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_DIR = os.path.join(BASE_DIR, "output")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Colors ─────────────────────────────────────────────────────────────────────
BG         = "#0d1117"
CARD       = "#161b27"
BORDER     = "#21293a"
ACCENT     = "#4f8ef7"
ACCENT_HV  = "#3b7de8"
SUCCESS    = "#34d399"
SUCCESS_HV = "#28c789"
DANGER     = "#f43f5e"
DANGER_HV  = "#e0284d"
TEXT       = "#e2e8f0"
MUTED      = "#64748b"
DIM        = "#1e2736"
LOG_BG     = "#080c14"
LOG_FG     = "#8899bb"
INPUT_BG   = "#0d1117"
ORANGE     = "#f59e0b"
ORANGE_HV  = "#d97706"

# ── System prompt ─────────────────────────────────────────────────────────────
SYSTEM_PROMPT = """
You are a self-storage market research analyst. Your job is to find EVERY
self-storage facility within the search radius and collect ACTUAL PRICING
for as many unit types as possible. Output a formatted Excel file.

## Research Strategy

### Phase 1 — Area-level search (do this FIRST — saves many turns)
1. WebSearch: "self storage units near [location] site:storageunits.com"
2. WebFetch the top StorageUnits.com area page — these list MULTIPLE facilities
   with prices on a SINGLE page. Extract every facility name, address, and price
   you can find. This one page may give you most of your data.
3. If StorageUnits.com didn't work, try: "self storage [city] [state] site:selfstorage.com"
   and fetch that area page instead.
4. Also do ONE general search: "self storage near [location]" to catch any
   facilities not listed on aggregators.

### Phase 2 — Write the Excel file IMMEDIATELY
After Phase 1, write the Excel file with whatever data you have so far.
Use N/A for any missing prices. This ensures a file always exists even if
you run out of turns later.

### Phase 3 — Fill in missing prices
For facilities that still have N/A prices after Phase 1:
1. WebSearch: "[facility name] [city] self storage unit prices"
2. **Check search snippets first** — prices often appear directly in result text.
3. **Fetch priority order** (try until you get pricing):
   a. StorageUnits.com or SelfStorage.com result (static HTML, most reliable)
   b. StorageCafe.com result
   c. The facility's own website (WARNING: many use JavaScript and return empty
      content — if fetch returns no pricing, move on immediately)
   d. SpareFoot result (often blocked — try last)
4. If the first fetch has no pricing, try ONE more page then move on.
5. After looking up each facility, REWRITE the Excel file with updated data.
   This ensures the latest data is always saved.

### Common issues:
- **JavaScript sites:** Major chains (Public Storage, Extra Space, CubeSmart,
  Life Storage) render prices via JS. Their direct websites return empty pricing.
  ALWAYS use aggregator sites for these chains.
- **Empty/blocked pages:** If WebFetch returns a CAPTCHA, "Access Denied," or
  very little text, do NOT retry that domain. Move to the next source.
- **Rate formats:** Prices appear as "$89", "$89.00", "$89/mo", "From $89".
  Extract the dollar amount.

## Target Unit Types
Collect pricing for ALL of these sizes when available, for BOTH
Climate Controlled AND Drive-Up units:
5x5, 5x10, 10x10, 10x15, 10x20, 10x25, 10x30, Parking/Vehicle Storage

If a facility offers a 5x5 climate controlled unit, include it. If they offer
a 5x5 drive-up unit, include that too. Do NOT skip smaller sizes like 5x5.
Include every unit size the facility offers.

## Collect per facility:
Name | Address | Phone | Website | Online rates by unit size | In-store rates |
Climate controlled (Yes/No/Drive-Up) | Promotions | Distance (mi) & drive time (min)
from subject using Haversine formula, 25 mph estimate.

Include only facilities within the search radius. Sort closest first.

## Excel — 3 Tabs

Tab 1 "Comps Detail" — one row per facility x unit type
Columns: Facility Name | Address | Distance (mi) | Unit Type | Sq Ft |
  Climate Controlled | Online Rate ($/mo) | In-Store Rate ($/mo) |
  Notes/Promotions | Date Pulled | Source URL | Drive Time (min)
Format: bold header (#BDD7EE), currency, auto-width columns, freeze row 1

Tab 2 "Market Summary" — two sections separated by a blank row
  Section 1 header: "CLIMATE CONTROLLED UNITS" (bold, green #E2EFDA)
  Section 2 header: "DRIVE UP / STANDARD UNITS" (bold, orange #FCE4D6)
  Each section columns: Unit Type | Sq Ft | Avg Online | Min Online | Max Online |
    Avg In-Store | Min In-Store | Max In-Store | # Comps
  Currency format on all rate columns.

Tab 3 "Facility List" — one row per facility
Columns: Facility Name | Address | Distance (mi) | Drive Time (min) | Phone | Website
Format: bold header (#FCE4D6), sorted by distance

## Rules
- Never fabricate pricing. Mark missing data as N/A.
- Distance required on every row — exclude facility if unknown.
- Write Excel via openpyxl using the Bash tool.
- A facility with ALL N/A prices is useless — try harder to find at least one price.
"""

# ── Fact-check agent system prompt ───────────────────────────────────────────
FACT_CHECK_PROMPT = """
You are an aggressive, skeptical price auditor. Your job is to tear apart a
self-storage market comps spreadsheet and find every error, inconsistency, and
suspicious data point. Assume nothing is correct until you prove it. Be blunt,
be specific, and do not sugarcoat findings. This report goes to investors who
need to know exactly how trustworthy this data is before making million-dollar
decisions.

## Step 1 — Read the Excel file
Use the Bash tool to run a Python script that reads the workbook with openpyxl
and prints ALL data from EVERY tab (Comps Detail, Market Summary, Facility List).
Print every row so you have the complete dataset.

## Step 2 — Verify prices
Work through facilities ONE AT A TIME. For each unique facility name:

### Search strategy (follow this order):
1. Do ONE WebSearch: "[facility name] [city] self storage unit prices"
2. **CHECK SEARCH SNIPPETS FIRST.** Search results often show prices directly
   in the snippet text (e.g., "5x10 starting at $89/mo"). Extract any prices
   you can find from the snippets before fetching any pages.
3. **Fetch priority order** — try these in order until you get pricing data:
   a. StorageUnits.com or SelfStorage.com result (these return static HTML
      that is reliably parseable — they are your BEST sources)
   b. The facility's own website (but note: many use JavaScript-heavy pages
      that may return empty content — if the fetch returns no pricing data,
      move on immediately)
   c. iStorage, StorageCafe.com, or other aggregator results
   d. SpareFoot (often blocked — try last)
4. If the first fetch returns no pricing, try up to 2 MORE pages from the
   search results before giving up.
5. If all fetches fail, try ONE alternative search:
   "site:storageunits.com [facility name] [city]" or
   "site:selfstorage.com [facility name] [city]"
   and fetch the top result.
6. From the results, extract rates for ALL unit types at that facility
   and compare against every row in the spreadsheet for that facility.

### Common issues to handle:
- **Empty/blocked pages:** If a WebFetch returns very little text, a CAPTCHA
  page, or "Access Denied," do NOT retry that domain. Move to the next source.
- **JavaScript sites:** Major chains (Public Storage, Extra Space, CubeSmart,
  Life Storage) render prices via JavaScript. Their direct websites will
  likely return empty pricing. Always prefer aggregator sites for these chains.
- **Rate format variations:** Prices may appear as "$89", "$89.00", "$89/mo",
  "From $89". Extract the dollar amount and compare.

IMPORTANT: Do NOT search separately for each unit type — search ONCE per
facility and match all unit types from the results. This saves turns.

Prices are compared as whole dollar amounts. "$129.00" and "$129" are the
same price. Strip cents for comparison — round to nearest dollar.

## Step 3 — Verify Market Summary math
Recalculate EVERY number in the Market Summary tab from the raw Comps Detail data:
- Avg, Min, Max for each unit type × climate type (Drive-Up vs Climate Controlled)
- Verify that "# Comps" counts match actual row counts in Comps Detail
- Flag ANY discrepancy, even $0.01 off. Show your calculated value vs the report value.
- Check that the correct rows are being included/excluded (e.g., N/A rates should
  not be counted in averages).

## Step 4 — Write the verification report
Use the Bash tool to write an Excel report with openpyxl. Save it to the
specified output path.

### Tab 1 "Price Verification"
Row 1: "PRICE VERIFICATION REPORT" (bold, size 14)
Row 2: "File reviewed: [filename]"
Row 3: "Audit date: [today]"
Row 4: blank
Row 5: Header row (bold, #BDD7EE fill):
  Facility Name | Unit Type | Climate | Report Online Rate | Report In-Store Rate |
  Verified Rate | Difference | Source | Status
Row 6+: One row per facility × unit type combination you were able to check.
- Status values: "Verified" (exact dollar match), "Mismatch" (any difference),
  "Not Found" (couldn't find pricing for that unit type)
- Difference column: "$0" for Verified, "+$X" or "-$X" for Mismatch, "N/A" for Not Found
- Green (#E2EFDA) fill for Verified rows, Red (#FCE4D6) fill for Mismatch rows
- Auto-width all columns

### Tab 2 "Math Verification"
Columns: Unit Type | Climate | Metric | Report Value | Calculated Value | Difference | Status
- One row per Avg/Min/Max/Count per unit type × climate type
- Status: "Correct" or "ERROR — off by $X"
- Green fill for correct, Red fill for errors
- Bold header row, auto-width

### Tab 3 "Analysis Notes"
This is your written critique. Be harsh. Be specific. Hold nothing back.
Row 1: "ANALYSIS NOTES" (bold, size 14)
Row 2: "Auditor: AI Price Verification Agent"
Row 3: "Date: [today]"
Row 4: blank

Then write the following sections, each with a bold header row:

**"PRICE ACCURACY ASSESSMENT"**
- For each facility, write 1-2 sentences on whether its prices checked out.
  If a price was wrong, state exactly what the report says vs what you found
  and where you found it. Call out patterns — is the report consistently
  high, low, or using stale data?

**"SUSPICIOUS PRICING"**
- Flag any rate that looks fabricated, rounded suspiciously, or is a clear
  outlier vs the market. Be specific: "Facility X lists 10x10 at $250/mo
  while every other facility in the radius is $120-$160. This is either
  premium pricing that needs explanation or a data error."
- Flag any facility where online and in-store rates are identical — this
  is unusual and may indicate the agent only found one rate and copied it.
- Flag any rate below $20/mo or above $600/mo as likely erroneous.

**"MATH ERRORS"**
- List every Market Summary calculation that doesn't match your recalculation.
  Show the math: "Report says 10x10 Drive-Up Avg = $142.50 but actual average
  of [$130, $145, $155, $140] = $142.50 — Correct" or "Report says Avg = $150
  but I calculate $142.50 from the data — ERROR, overstated by $7.50."

**"DATA GAPS & CONCERNS"**
- How many unit types have zero verified prices? What % of rates are "N/A"?
- Are there facilities with no pricing at all? Why were they included?
- Any unit types completely missing from the market that should be there?

**"BOTTOM LINE"**
- 2-3 sentences maximum. Is this data reliable enough to base an investment
  decision on? Give a straight answer. If the prices are mostly wrong, say so
  plainly. If they're solid, acknowledge it — but still note any caveats.

Format: Use openpyxl to write each section. Bold headers in size 12. Regular
text in size 10. Wrap text in the notes column. Set column A width to 120
characters. Leave a blank row between sections.

### Tab 4 "Summary"
Row 1: "VERIFICATION SUMMARY" (bold, size 14)
Row 2: blank
Row 3+: Stats table:
  - Total facilities in report
  - Total facilities checked
  - Total prices verified
  - Prices confirmed accurate (count and %)
  - Prices with mismatches (count and %)
  - Prices unable to verify (count and %)
  - Market Summary calculations checked
  - Market Summary errors found
  - Overall accuracy score: X/10

Scoring:
- 9-10: All prices verified, math is correct, no concerns
- 7-8: Most prices check out, minor math rounding issues
- 5-6: Multiple price mismatches or math errors that affect reliability
- 3-4: Significant inaccuracies — do not rely on this data without re-running
- 1-2: Widespread fabrication or errors — data is unreliable

## Rules
- Never fabricate verification data — if you can't find a price, mark "Not Found."
- Format all currency values with $ and commas.
- Auto-width all columns in the Excel output.
- A price is "Verified" ONLY if it is an exact dollar match to the report rate
  (rounded to whole dollars — $129.00 = $129). Any difference, even $1, is a
  "Mismatch." Always show the difference amount in the Difference column.
- Be ruthlessly honest in Analysis Notes. Do not hedge or use soft language.
  Say "wrong," "fabricated," "error" — not "slightly different" or "minor gap."
"""

# ── Cost data ──────────────────────────────────────────────────────────────────
# Base $/SF at national average, "Average" quality (2025-2026 estimates)
DRIVEUP_COSTS = [
    ("Site Work & Grading",        5.50),
    ("Concrete Slab / Foundation", 8.00),
    ("Steel Structure",           18.00),
    ("Metal Roofing",              5.50),
    ("Electrical & Lighting",      4.00),
    ("Paving & Parking",           4.50),
]
DRIVEUP_LUMP = [
    ("Roll-Up Doors",    "~1 per 125 SF", lambda sf: int(sf / 125) * 1100),
    ("Security System",  "Lump sum",      lambda sf: 30000),
    ("Office Buildout",  "~400 SF",       lambda sf: 400 * 130),
]

CC_COSTS = [
    ("Site Work & Grading",        5.50),
    ("Concrete Slab / Foundation", 9.50),
    ("Steel Structure",           22.00),
    ("Metal Roofing",              5.50),
    ("HVAC System",               12.00),
    ("Insulation",                 4.00),
    ("Interior Corridors",         6.00),
    ("Fire Suppression",           3.50),
    ("Electrical & Lighting",      5.00),
    ("Paving & Parking",           4.50),
]
CC_LUMP = [
    ("Roll-Up / Entry Doors", "~1 per 100 SF", lambda sf: int(sf / 100) * 1200),
    ("Elevator",              "If > 20k SF",    lambda sf: 120000 if sf > 20000 else 0),
    ("Security System",       "Lump sum",       lambda sf: 40000),
    ("Office Buildout",       "~400 SF",        lambda sf: 400 * 140),
]

QUALITY_MULT = {"Economy": 0.85, "Average": 1.00, "Premium": 1.15}

# Soft costs as % of hard costs (itemized)
SOFT_COSTS = [
    ("Architectural & Engineering",   0.050),
    ("Permits & Impact Fees",         0.025),
    ("Geotechnical / Environmental",  0.008),
    ("Survey & Land Planning",        0.004),
    ("Legal & Closing",               0.008),
    ("Builder's Risk Insurance",      0.007),
    ("Construction Loan Interest",    0.040),
    ("Property Taxes During Const.",  0.008),
    ("Contingency",                   0.075),
]
# Total soft cost % = sum of above = ~22.5%

# ── Cost agent system prompt ──────────────────────────────────────────────────
COST_AGENT_PROMPT = """
You are a self-storage construction cost analyst. Your job is to research CURRENT,
ACCURATE construction costs for a specific city and building type, then write an
Excel cost estimate.

MINIMIZE web fetches — every extra fetch costs money. Max 3 fetches total.

## Research Strategy
1. ONE WebSearch: "self storage construction cost per square foot [city] [year]"
2. ONE WebSearch: "USACE area cost factor [city] [year]" OR
   "RSMeans construction cost [city] [year]"
3. Fetch the 1-2 most relevant results to get real numbers.

## What to find:
- Current $/SF for self-storage construction in or near that city
- Location cost factor vs national average
- Breakdown by component if available (site work, steel, concrete, HVAC, etc.)
- Any recent project cost data for self-storage in the area

## Output
Write an Excel file using openpyxl with these details:

Tab 1 "Cost Estimate":
Row 1: "Self Storage Construction Cost Estimate" (bold, size 14)
Row 2: Building Type
Row 3: Total SF
Row 4: City + location factor found
Row 5: Quality level
Row 6: Date + sources used
Row 7: blank
Row 8: Header row (bold, #BDD7EE fill): Component | $/SF | Total Cost
Rows 9+: One row per cost component with real researched $/SF values
Then: Hard Cost Subtotal (bold)
Then: Itemized Soft Costs (~22.5% total: A&E 5%, Permits 2.5%, Geotech 0.8%, Survey 0.4%, Legal 0.8%, Insurance 0.7%, Loan Interest 4%, Property Tax 0.8%, Contingency 7.5%)
Then: TOTAL ESTIMATED COST (bold, #E2EFDA fill) with $/SF and total

Tab 2 "Sources":
List every URL and source used with what data came from each.

## Rules
- Use REAL data from your research, not generic estimates.
- If you cannot find city-specific data, use the closest metro area and note that.
- Always include the source of each number.
- Currency format on all dollar columns.
- Auto-width all columns.
"""

# USACE-based location cost factors (subset of major metros)
LOCATION_FACTORS = {
    "new york": 1.42, "manhattan": 1.48, "brooklyn": 1.42, "bronx": 1.42,
    "los angeles": 1.18, "chicago": 1.12, "houston": 0.88, "phoenix": 0.92,
    "philadelphia": 1.15, "san antonio": 0.85, "san diego": 1.15,
    "dallas": 0.90, "austin": 0.92, "fort worth": 0.89,
    "jacksonville": 0.87, "san francisco": 1.38, "san jose": 1.30,
    "columbus": 0.93, "charlotte": 0.88, "indianapolis": 0.92,
    "seattle": 1.15, "denver": 0.96, "nashville": 0.90,
    "atlanta": 0.90, "portland": 1.08, "las vegas": 0.98,
    "memphis": 0.85, "louisville": 0.90, "baltimore": 0.98,
    "milwaukee": 1.02, "albuquerque": 0.90, "tucson": 0.90,
    "fresno": 1.05, "sacramento": 1.12, "miami": 0.95,
    "tampa": 0.90, "orlando": 0.90, "st louis": 0.98,
    "pittsburgh": 1.00, "raleigh": 0.88, "minneapolis": 1.05,
    "cleveland": 0.98, "detroit": 1.00, "boston": 1.25,
    "honolulu": 1.35, "anchorage": 1.28, "kansas city": 0.95,
    "oklahoma city": 0.85, "omaha": 0.90, "virginia beach": 0.90,
    "colorado springs": 0.93, "tulsa": 0.84, "arlington": 0.89,
    "new orleans": 0.88, "bakersfield": 1.05, "boise": 0.92,
    "richmond": 0.90, "des moines": 0.92, "salt lake city": 0.93,
    "birmingham": 0.85, "spokane": 1.00, "rochester": 1.02,
}


def _lookup_location_factor(city_text: str) -> tuple[float, str]:
    """Return (factor, matched_city) or (1.0, '') if not found."""
    lower = city_text.lower().strip().rstrip(",").strip()
    # Try direct match
    for city, factor in LOCATION_FACTORS.items():
        if city in lower:
            return factor, city.title()
    return 1.00, ""


# ── App ─────────────────────────────────────────────────────────────────────────
class StorageCompsApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Storage Tools")
        self.geometry("960x750")
        self.minsize(760, 580)
        self.configure(bg=BG)

        self._running = False
        self._thread = None
        self._output_file = None
        self._step = 0
        self._active_tab = "comps"

        # Fact-check state
        self._fc_running = False
        self._fc_thread: threading.Thread | None = None
        self._fc_output_file: str = ""
        self._fc_log_widget: scrolledtext.ScrolledText | None = None

        # Widget references (set in _build_ui)
        self.radius_var: tk.StringVar
        self.location_var: tk.StringVar
        self.step_var: tk.StringVar
        self.pct_var: tk.StringVar
        self.step_lbl: tk.Label
        self.progress: ttk.Progressbar
        self.cost_type_var: tk.StringVar
        self.cost_sf_var: tk.StringVar
        self.cost_city_var: tk.StringVar
        self.cost_quality_var: tk.StringVar
        self._quick_btn: tk.Button
        self._accurate_btn: tk.Button
        self._cost_running = False
        self._cost_thread: threading.Thread | None = None
        self._cost_output_file: str = ""
        self._cost_log_widget: scrolledtext.ScrolledText | None = None
        self._cost_export_btn: tk.Button
        self._cost_export_frame: tk.Frame
        self._results_inner: tk.Frame
        self._results_canvas: tk.Canvas
        self._cost_location_lbl: tk.Label
        self._cost_placeholder: tk.Label
        self._driveup_btn: tk.Button
        self._cc_btn: tk.Button
        self._last_estimate: dict | None = None
        self._tab_btns: dict = {}
        self._comps_frame: tk.Frame
        self._cost_frame: tk.Frame
        self._content: tk.Frame

        self._build_ui()

    # ── UI construction ────────────────────────────────────────────────────────
    def _build_ui(self):

        # ── Top accent line ──
        tk.Frame(self, bg=ACCENT, height=3).pack(fill="x")

        # ── Header ──
        header = tk.Frame(self, bg=BG, padx=24, pady=14)
        header.pack(fill="x")

        tk.Label(
            header, text="Storage Tools",
            font=("Segoe UI", 20, "bold"), bg=BG, fg=TEXT,
        ).pack(side="left")

        tk.Label(
            header, text="  ·  Self Storage Analysis Suite",
            font=("Segoe UI", 11), bg=BG, fg=MUTED,
        ).pack(side="left", pady=(6, 0))

        # ── Tab bar ──
        tab_bar = tk.Frame(self, bg=BG, padx=20)
        tab_bar.pack(fill="x")

        self._tab_btns = {}
        for tab_id, label in [("comps", "Market Comps"), ("cost", "Cost Estimator")]:
            btn = tk.Button(
                tab_bar, text=label,
                font=("Segoe UI", 10, "bold"),
                relief="flat", bd=0, padx=20, pady=8, cursor="hand2",
                command=lambda t=tab_id: self._switch_tab(t),
            )
            btn.pack(side="left", padx=(0, 4))
            self._tab_btns[tab_id] = btn

        tk.Frame(self, bg=BORDER, height=1).pack(fill="x", padx=20, pady=(6, 0))

        # ── Content area ──
        self._content = tk.Frame(self, bg=BG)
        self._content.pack(fill="both", expand=True)

        # ── Build both tabs ──
        self._comps_frame = tk.Frame(self._content, bg=BG)
        self._cost_frame = tk.Frame(self._content, bg=BG)
        self._build_comps_tab(self._comps_frame)
        self._build_cost_tab(self._cost_frame)

        # ── Status bar (shared) ──
        status_bar = tk.Frame(self, bg=CARD, height=40,
                              highlightthickness=1, highlightbackground=BORDER)
        status_bar.pack(fill="x", side="bottom")
        status_bar.pack_propagate(False)

        self.status_dot = tk.Label(
            status_bar, text="\u25cf",
            font=("Segoe UI", 9), bg=CARD, fg=MUTED,
        )
        self.status_dot.pack(side="left", padx=(16, 5), pady=10)

        self.status_var = tk.StringVar(value="Ready")
        self.status_lbl = tk.Label(
            status_bar, textvariable=self.status_var,
            font=("Segoe UI", 9), bg=CARD, fg=MUTED, anchor="w",
        )
        self.status_lbl.pack(side="left", pady=10)

        self.open_btn = tk.Button(
            status_bar, text="Open Excel  \u25b6",
            font=("Segoe UI", 9, "bold"),
            bg=SUCCESS, fg="#051a0e", activebackground=SUCCESS_HV,
            relief="flat", padx=16, pady=5, cursor="hand2",
            command=self._open_file, state="disabled",
        )
        self.open_btn.pack(side="right", padx=14, pady=7)

        # Show default tab
        self._switch_tab("comps")

    # ── Tab switching ─────────────────────────────────────────────────────────
    def _switch_tab(self, tab_id: str):
        self._active_tab = tab_id
        # Hide all
        self._comps_frame.pack_forget()
        self._cost_frame.pack_forget()
        # Show selected
        if tab_id == "comps":
            self._comps_frame.pack(fill="both", expand=True)
        else:
            self._cost_frame.pack(fill="both", expand=True)
        # Update tab button styles
        for tid, btn in self._tab_btns.items():
            if tid == tab_id:
                btn.config(bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HV)
            else:
                btn.config(bg=DIM, fg=MUTED, activebackground=BORDER)

    # ── Comps tab ─────────────────────────────────────────────────────────────
    def _build_comps_tab(self, parent):
        # Input card
        card_wrap = tk.Frame(parent, bg=BG)
        card_wrap.pack(fill="x", padx=20, pady=(12, 0))

        card = tk.Frame(card_wrap, bg=CARD, highlightthickness=1,
                        highlightbackground=BORDER)
        card.pack(fill="x")

        inner = tk.Frame(card, bg=CARD, padx=22, pady=18)
        inner.pack(fill="x")

        tk.Label(inner, text="Subject Property Location",
                 font=("Segoe UI", 9), bg=CARD, fg=MUTED,
                 ).grid(row=0, column=0, sticky="w", pady=(0, 5))
        tk.Label(inner, text="Radius (mi)",
                 font=("Segoe UI", 9), bg=CARD, fg=MUTED,
                 ).grid(row=0, column=1, sticky="w", padx=(14, 0), pady=(0, 5))

        HINT = "e.g. 78701  or  Austin, TX  or  123 Main St, Denver CO"
        self.location_var = tk.StringVar()
        loc_entry = tk.Entry(
            inner, textvariable=self.location_var,
            font=("Segoe UI", 12), width=42,
            bg=INPUT_BG, fg=MUTED, insertbackground=ACCENT,
            relief="flat", bd=0,
            highlightthickness=1, highlightbackground=BORDER, highlightcolor=ACCENT,
        )
        loc_entry.grid(row=1, column=0, ipady=9, padx=(0, 14))
        loc_entry.insert(0, HINT)
        loc_entry.bind("<Return>", lambda e: self._start())

        def _clear_hint(e):
            if loc_entry.get() == HINT:
                loc_entry.delete(0, "end")
                loc_entry.config(fg=TEXT)

        def _restore_hint(e):
            if not loc_entry.get().strip():
                loc_entry.insert(0, HINT)
                loc_entry.config(fg=MUTED)

        loc_entry.bind("<FocusIn>", _clear_hint)
        loc_entry.bind("<FocusOut>", _restore_hint)

        self.radius_var = tk.StringVar(value="5")
        tk.Spinbox(
            inner, textvariable=self.radius_var,
            from_=1, to=7, increment=1, width=5,
            font=("Segoe UI", 12),
            bg=INPUT_BG, fg=TEXT, buttonbackground=CARD,
            relief="flat", bd=0,
            highlightthickness=1, highlightbackground=BORDER, highlightcolor=ACCENT,
        ).grid(row=1, column=1, ipady=9, padx=(0, 14))

        btn_frame = tk.Frame(inner, bg=CARD)
        btn_frame.grid(row=1, column=2)

        self.run_btn = tk.Button(
            btn_frame, text="\u25b6  Run Analysis",
            font=("Segoe UI", 10, "bold"),
            bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HV,
            relief="flat", padx=20, pady=10, cursor="hand2",
            command=self._start,
        )
        self.run_btn.pack(side="left", padx=(0, 8))

        self.stop_btn = tk.Button(
            btn_frame, text="\u25a0  Stop",
            font=("Segoe UI", 10, "bold"),
            bg=DIM, fg=MUTED, activebackground=DANGER_HV,
            relief="flat", padx=16, pady=10, cursor="hand2",
            command=self._stop, state="disabled",
        )
        self.stop_btn.pack(side="left", padx=(0, 8))

        self.fc_btn = tk.Button(
            btn_frame, text="\u2714  Fact-Check",
            font=("Segoe UI", 10, "bold"),
            bg=ORANGE, fg="#1a1000", activebackground=ORANGE_HV,
            relief="flat", padx=16, pady=10, cursor="hand2",
            command=self._start_fact_check, state="disabled",
        )
        self.fc_btn.pack(side="left")

        # Log
        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", padx=20, pady=(14, 0))
        log_outer = tk.Frame(parent, bg=BG)
        log_outer.pack(fill="both", expand=True, padx=20, pady=(10, 0))

        tk.Label(log_outer, text="Activity Log",
                 font=("Segoe UI", 9, "bold"),
                 bg=BG, fg=MUTED).pack(anchor="w", pady=(0, 6))

        self.log = scrolledtext.ScrolledText(
            log_outer, font=("Consolas", 9),
            bg=LOG_BG, fg=LOG_FG, insertbackground=LOG_FG,
            relief="flat", bd=0, wrap="word", state="disabled",
            highlightthickness=1, highlightbackground=BORDER,
        )
        self.log.pack(fill="both", expand=True, pady=(0, 10))

        self.log.tag_config("info",    foreground=LOG_FG)
        self.log.tag_config("header",  foreground=ACCENT, font=("Consolas", 9, "bold"))
        self.log.tag_config("success", foreground=SUCCESS)
        self.log.tag_config("muted",   foreground=MUTED)
        self.log.tag_config("error",   foreground=DANGER)

        # Progress bar (below Activity Log)
        progress_frame = tk.Frame(parent, bg=BG)
        progress_frame.pack(fill="x", padx=20, pady=(0, 6))

        progress_top = tk.Frame(progress_frame, bg=BG)
        progress_top.pack(fill="x", pady=(0, 6))

        self.step_var = tk.StringVar(value="")
        self.step_lbl = tk.Label(
            progress_top, textvariable=self.step_var,
            font=("Segoe UI", 9), bg=BG, fg=MUTED, anchor="w",
        )
        self.step_lbl.pack(side="left")

        self.pct_var = tk.StringVar(value="")
        tk.Label(
            progress_top, textvariable=self.pct_var,
            font=("Segoe UI", 9), bg=BG, fg=MUTED, anchor="e",
        ).pack(side="right")

        style = ttk.Style()
        style.theme_use("default")
        style.configure("Custom.Horizontal.TProgressbar",
                        troughcolor=BORDER, background=ACCENT,
                        darkcolor=ACCENT, lightcolor=ACCENT,
                        bordercolor=BORDER, thickness=6)
        self.progress = ttk.Progressbar(
            progress_frame, style="Custom.Horizontal.TProgressbar",
            orient="horizontal", length=100, mode="determinate",
            maximum=100, value=0,
        )
        self.progress.pack(fill="x")

    # ── Cost Estimator tab ────────────────────────────────────────────────────
    def _build_cost_tab(self, parent):
        # Input card
        card_wrap = tk.Frame(parent, bg=BG)
        card_wrap.pack(fill="x", padx=20, pady=(12, 0))

        card = tk.Frame(card_wrap, bg=CARD, highlightthickness=1,
                        highlightbackground=BORDER)
        card.pack(fill="x")

        inner = tk.Frame(card, bg=CARD, padx=22, pady=18)
        inner.pack(fill="x")

        # ── Building type toggle ──
        tk.Label(inner, text="Building Type",
                 font=("Segoe UI", 9), bg=CARD, fg=MUTED,
                 ).grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 6))

        toggle_frame = tk.Frame(inner, bg=CARD)
        toggle_frame.grid(row=1, column=0, columnspan=2, sticky="w", pady=(0, 14))

        self.cost_type_var = tk.StringVar(value="driveup")

        self._driveup_btn = tk.Button(
            toggle_frame, text="Drive-Up",
            font=("Segoe UI", 10, "bold"),
            bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HV,
            relief="flat", padx=24, pady=8, cursor="hand2",
            command=lambda: self._set_building_type("driveup"),
        )
        self._driveup_btn.pack(side="left", padx=(0, 2))

        self._cc_btn = tk.Button(
            toggle_frame, text="Climate Controlled",
            font=("Segoe UI", 10, "bold"),
            bg=DIM, fg=MUTED, activebackground=BORDER,
            relief="flat", padx=24, pady=8, cursor="hand2",
            command=lambda: self._set_building_type("cc"),
        )
        self._cc_btn.pack(side="left")

        # ── Input fields ──
        fields_frame = tk.Frame(inner, bg=CARD)
        fields_frame.grid(row=2, column=0, columnspan=3, sticky="w")

        # Total SF
        tk.Label(fields_frame, text="Total Building SF",
                 font=("Segoe UI", 9), bg=CARD, fg=MUTED,
                 ).grid(row=0, column=0, sticky="w", pady=(0, 5))
        self.cost_sf_var = tk.StringVar(value="50000")
        tk.Entry(
            fields_frame, textvariable=self.cost_sf_var,
            font=("Segoe UI", 12), width=14,
            bg=INPUT_BG, fg=TEXT, insertbackground=ACCENT,
            relief="flat", bd=0,
            highlightthickness=1, highlightbackground=BORDER, highlightcolor=ACCENT,
        ).grid(row=1, column=0, ipady=9, padx=(0, 14))

        # City
        tk.Label(fields_frame, text="City / Metro",
                 font=("Segoe UI", 9), bg=CARD, fg=MUTED,
                 ).grid(row=0, column=1, sticky="w", padx=(0, 0), pady=(0, 5))
        self.cost_city_var = tk.StringVar(value="")
        city_entry = tk.Entry(
            fields_frame, textvariable=self.cost_city_var,
            font=("Segoe UI", 12), width=22,
            bg=INPUT_BG, fg=TEXT, insertbackground=ACCENT,
            relief="flat", bd=0,
            highlightthickness=1, highlightbackground=BORDER, highlightcolor=ACCENT,
        )
        city_entry.grid(row=1, column=1, ipady=9, padx=(0, 14))

        CITY_HINT = "e.g. Austin, Denver, Miami"
        city_entry.insert(0, CITY_HINT)
        city_entry.config(fg=MUTED)

        def _clear_city(e):
            if city_entry.get() == CITY_HINT:
                city_entry.delete(0, "end")
                city_entry.config(fg=TEXT)

        def _restore_city(e):
            if not city_entry.get().strip():
                city_entry.insert(0, CITY_HINT)
                city_entry.config(fg=MUTED)

        city_entry.bind("<FocusIn>", _clear_city)
        city_entry.bind("<FocusOut>", _restore_city)

        # Quality
        tk.Label(fields_frame, text="Quality",
                 font=("Segoe UI", 9), bg=CARD, fg=MUTED,
                 ).grid(row=0, column=2, sticky="w", pady=(0, 5))
        self.cost_quality_var = tk.StringVar(value="Average")
        quality_menu = ttk.Combobox(
            fields_frame, textvariable=self.cost_quality_var,
            values=["Economy", "Average", "Premium"],
            font=("Segoe UI", 11), width=10, state="readonly",
        )
        quality_menu.grid(row=1, column=2, ipady=6, padx=(0, 14))

        # Estimate buttons
        est_btn_frame = tk.Frame(fields_frame, bg=CARD)
        est_btn_frame.grid(row=1, column=3, padx=(4, 0))

        self._quick_btn = tk.Button(
            est_btn_frame, text="\u25b6  Quick Estimate",
            font=("Segoe UI", 9, "bold"),
            bg=ORANGE, fg="#1a1000", activebackground=ORANGE_HV,
            relief="flat", padx=14, pady=10, cursor="hand2",
            command=self._estimate_cost,
        )
        self._quick_btn.pack(side="left", padx=(0, 6))

        self._accurate_btn = tk.Button(
            est_btn_frame, text="\u25b6  Accurate Estimate",
            font=("Segoe UI", 9, "bold"),
            bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HV,
            relief="flat", padx=14, pady=10, cursor="hand2",
            command=self._accurate_estimate,
        )
        self._accurate_btn.pack(side="left")

        # ── Divider ──
        tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", padx=20, pady=(14, 0))

        # ── Results header ──
        results_header = tk.Frame(parent, bg=BG)
        results_header.pack(fill="x", padx=20, pady=(12, 0))

        tk.Label(results_header, text="Cost Breakdown",
                 font=("Segoe UI", 9, "bold"),
                 bg=BG, fg=MUTED).pack(side="left")

        self._cost_location_lbl = tk.Label(
            results_header, text="",
            font=("Segoe UI", 9), bg=BG, fg=MUTED,
        )
        self._cost_location_lbl.pack(side="right")

        # ── Results table ──
        table_outer = tk.Frame(parent, bg=BG)
        table_outer.pack(fill="both", expand=True, padx=20, pady=(8, 0))

        # Canvas + scrollbar for results
        self._results_canvas = tk.Canvas(
            table_outer, bg=LOG_BG, bd=0, highlightthickness=1,
            highlightbackground=BORDER,
        )
        scrollbar = ttk.Scrollbar(table_outer, orient="vertical",
                                  command=self._results_canvas.yview)
        self._results_inner = tk.Frame(self._results_canvas, bg=LOG_BG)

        self._results_inner.bind(
            "<Configure>",
            lambda e: self._results_canvas.configure(
                scrollregion=self._results_canvas.bbox("all")
            ),
        )
        self._results_canvas.create_window((0, 0), window=self._results_inner,
                                           anchor="nw")
        self._results_canvas.configure(yscrollcommand=scrollbar.set)

        self._results_canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        # Placeholder
        self._cost_placeholder = tk.Label(
            self._results_inner,
            text="Enter building details above and click Estimate Cost",
            font=("Segoe UI", 11), bg=LOG_BG, fg=MUTED, pady=40,
        )
        self._cost_placeholder.pack()

        # Export button (hidden until results exist)
        self._cost_export_frame = tk.Frame(parent, bg=BG)
        self._cost_export_frame.pack(fill="x", padx=20, pady=(8, 10))

        self._cost_export_btn = tk.Button(
            self._cost_export_frame, text="Export to Excel",
            font=("Segoe UI", 9, "bold"),
            bg=SUCCESS, fg="#051a0e", activebackground=SUCCESS_HV,
            relief="flat", padx=16, pady=6, cursor="hand2",
            command=self._export_cost_excel, state="disabled",
        )
        self._cost_export_btn.pack(side="right")

        # Store last estimate for export
        self._last_estimate = None

    # ── Building type toggle ──────────────────────────────────────────────────
    def _set_building_type(self, btype: str):
        self.cost_type_var.set(btype)
        if btype == "driveup":
            self._driveup_btn.config(bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HV)
            self._cc_btn.config(bg=DIM, fg=MUTED, activebackground=BORDER)
        else:
            self._cc_btn.config(bg=ACCENT, fg="#ffffff", activebackground=ACCENT_HV)
            self._driveup_btn.config(bg=DIM, fg=MUTED, activebackground=BORDER)

    # ── Cost estimation ───────────────────────────────────────────────────────
    def _estimate_cost(self):
        # Validate SF
        try:
            sf = float(self.cost_sf_var.get().replace(",", ""))
            if sf <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Invalid SF", "Please enter a valid building square footage.")
            return

        city_text = self.cost_city_var.get().strip()
        CITY_HINT = "e.g. Austin, Denver, Miami"
        if not city_text or city_text == CITY_HINT:
            city_text = ""

        quality = self.cost_quality_var.get()
        btype = self.cost_type_var.get()
        q_mult = QUALITY_MULT.get(quality, 1.0)

        # Location factor
        loc_factor, matched_city = _lookup_location_factor(city_text)
        if matched_city:
            self._cost_location_lbl.config(
                text=f"Location factor: {loc_factor:.2f}x  ({matched_city})",
                fg=ACCENT,
            )
        elif city_text:
            self._cost_location_lbl.config(
                text=f"City not found — using national average (1.00x)",
                fg=ORANGE,
            )
        else:
            self._cost_location_lbl.config(
                text=f"No city — using national average (1.00x)",
                fg=MUTED,
            )

        # Pick cost tables
        per_sf_items = CC_COSTS if btype == "cc" else DRIVEUP_COSTS
        lump_items = CC_LUMP if btype == "cc" else DRIVEUP_LUMP

        # Calculate
        rows = []
        total_hard = 0.0

        for name, base_psf in per_sf_items:
            adj_psf = base_psf * q_mult * loc_factor
            cost = adj_psf * sf
            rows.append((name, f"${adj_psf:,.2f}", f"${cost:,.0f}"))
            total_hard += cost

        for name, note, calc_fn in lump_items:
            cost = calc_fn(sf) * q_mult * loc_factor
            if cost > 0:
                rows.append((name, note, f"${cost:,.0f}"))
                total_hard += cost

        soft_items = [(name, pct, total_hard * pct) for name, pct in SOFT_COSTS]
        total_soft = sum(amt for _, _, amt in soft_items)
        grand_total = total_hard + total_soft
        total_psf = grand_total / sf if sf > 0 else 0

        # Clear previous results
        for w in self._results_inner.winfo_children():
            w.destroy()

        # ── Draw table ──
        HDR_BG = "#1a2233"
        ROW_BG1 = LOG_BG
        ROW_BG2 = "#0c1018"
        TOTAL_BG = "#132030"

        # Header row
        headers = ["Component", "$/SF or Note", "Total Cost"]
        col_widths = [30, 16, 16]
        for col, (h, w) in enumerate(zip(headers, col_widths)):
            tk.Label(
                self._results_inner, text=h,
                font=("Segoe UI", 9, "bold"), bg=HDR_BG, fg=ACCENT,
                anchor="w", padx=12, pady=8, width=w,
            ).grid(row=0, column=col, sticky="ew")

        # Data rows
        for i, (name, psf, cost) in enumerate(rows):
            bg = ROW_BG1 if i % 2 == 0 else ROW_BG2
            tk.Label(self._results_inner, text=name,
                     font=("Segoe UI", 9), bg=bg, fg=TEXT,
                     anchor="w", padx=12, pady=6, width=col_widths[0],
                     ).grid(row=i + 1, column=0, sticky="ew")
            tk.Label(self._results_inner, text=psf,
                     font=("Consolas", 9), bg=bg, fg=LOG_FG,
                     anchor="e", padx=12, pady=6, width=col_widths[1],
                     ).grid(row=i + 1, column=1, sticky="ew")
            tk.Label(self._results_inner, text=cost,
                     font=("Consolas", 9), bg=bg, fg=TEXT,
                     anchor="e", padx=12, pady=6, width=col_widths[2],
                     ).grid(row=i + 1, column=2, sticky="ew")

        # Subtotal row
        r = len(rows) + 1
        tk.Label(self._results_inner, text="Hard Cost Subtotal",
                 font=("Segoe UI", 9, "bold"), bg=TOTAL_BG, fg=TEXT,
                 anchor="w", padx=12, pady=8, width=col_widths[0],
                 ).grid(row=r, column=0, sticky="ew")
        tk.Label(self._results_inner, text="",
                 bg=TOTAL_BG, width=col_widths[1],
                 ).grid(row=r, column=1, sticky="ew")
        tk.Label(self._results_inner, text=f"${total_hard:,.0f}",
                 font=("Consolas", 9, "bold"), bg=TOTAL_BG, fg=TEXT,
                 anchor="e", padx=12, pady=8, width=col_widths[2],
                 ).grid(row=r, column=2, sticky="ew")

        # Soft cost rows (itemized)
        for si, (sname, spct, samt) in enumerate(soft_items):
            r += 1
            bg = ROW_BG1 if si % 2 == 0 else ROW_BG2
            tk.Label(self._results_inner, text=f"  {sname}",
                     font=("Segoe UI", 9), bg=bg, fg=MUTED,
                     anchor="w", padx=12, pady=4, width=col_widths[0],
                     ).grid(row=r, column=0, sticky="ew")
            tk.Label(self._results_inner, text=f"{spct:.1%}",
                     font=("Consolas", 9), bg=bg, fg=MUTED,
                     anchor="e", padx=12, pady=4, width=col_widths[1],
                     ).grid(row=r, column=1, sticky="ew")
            tk.Label(self._results_inner, text=f"${samt:,.0f}",
                     font=("Consolas", 9), bg=bg, fg=MUTED,
                     anchor="e", padx=12, pady=4, width=col_widths[2],
                     ).grid(row=r, column=2, sticky="ew")

        # Soft cost subtotal
        r += 1
        tk.Label(self._results_inner, text="Soft Cost Subtotal",
                 font=("Segoe UI", 9, "bold"), bg=TOTAL_BG, fg=MUTED,
                 anchor="w", padx=12, pady=6, width=col_widths[0],
                 ).grid(row=r, column=0, sticky="ew")
        tk.Label(self._results_inner, text="",
                 bg=TOTAL_BG, width=col_widths[1],
                 ).grid(row=r, column=1, sticky="ew")
        tk.Label(self._results_inner, text=f"${total_soft:,.0f}",
                 font=("Consolas", 9, "bold"), bg=TOTAL_BG, fg=MUTED,
                 anchor="e", padx=12, pady=6, width=col_widths[2],
                 ).grid(row=r, column=2, sticky="ew")

        # Grand total row
        r += 1
        tk.Label(self._results_inner, text="TOTAL ESTIMATED COST",
                 font=("Segoe UI", 10, "bold"), bg=ACCENT, fg="#ffffff",
                 anchor="w", padx=12, pady=10, width=col_widths[0],
                 ).grid(row=r, column=0, sticky="ew")
        tk.Label(self._results_inner, text=f"${total_psf:,.2f} / SF",
                 font=("Consolas", 10, "bold"), bg=ACCENT, fg="#ffffff",
                 anchor="e", padx=12, pady=10, width=col_widths[1],
                 ).grid(row=r, column=1, sticky="ew")
        tk.Label(self._results_inner, text=f"${grand_total:,.0f}",
                 font=("Consolas", 10, "bold"), bg=ACCENT, fg="#ffffff",
                 anchor="e", padx=12, pady=10, width=col_widths[2],
                 ).grid(row=r, column=2, sticky="ew")

        # Enable export
        self._last_estimate = {
            "btype": "Climate Controlled" if btype == "cc" else "Drive-Up",
            "sf": sf, "city": city_text or "National Avg",
            "quality": quality, "loc_factor": loc_factor,
            "rows": rows, "total_hard": total_hard,
            "soft_items": soft_items, "total_soft": total_soft,
            "grand_total": grand_total, "total_psf": total_psf,
        }
        self._cost_export_btn.config(state="normal")
        self._set_status(
            f"Estimate: {self._last_estimate['btype']} — "
            f"{sf:,.0f} SF — ${grand_total:,.0f} total (${total_psf:,.2f}/SF)",
            SUCCESS,
        )

    # ── Export cost estimate to Excel ──────────────────────────────────────────
    def _export_cost_excel(self):
        if not self._last_estimate:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, numbers

            est = self._last_estimate
            wb = Workbook()
            ws = wb.active
            ws.title = "Cost Estimate"

            # Header info
            ws.append(["Self Storage Construction Cost Estimate"])
            ws["A1"].font = Font(size=14, bold=True)
            ws.append([f"Building Type: {est['btype']}"])
            ws.append([f"Total SF: {est['sf']:,.0f}"])
            ws.append([f"City: {est['city']}  (Location Factor: {est['loc_factor']:.2f}x)"])
            ws.append([f"Quality: {est['quality']}"])
            ws.append([f"Date: {date.today().strftime('%B %d, %Y')}"])
            ws.append([])

            # Table header
            hdr_fill = PatternFill("solid", fgColor="BDD7EE")
            hdr_font = Font(bold=True)
            for col, h in enumerate(["Component", "$/SF or Note", "Total Cost"], 1):
                cell = ws.cell(row=8, column=col, value=h)
                cell.font = hdr_font
                cell.fill = hdr_fill

            # Data rows
            row = 9
            for name, psf, cost in est["rows"]:
                ws.cell(row=row, column=1, value=name)
                ws.cell(row=row, column=2, value=psf)
                ws.cell(row=row, column=3, value=cost)
                row += 1

            # Totals
            ws.cell(row=row, column=1, value="Hard Cost Subtotal").font = Font(bold=True)
            ws.cell(row=row, column=3, value=f"${est['total_hard']:,.0f}").font = Font(bold=True)
            row += 1

            # Itemized soft costs
            soft_fill = PatternFill("solid", fgColor="F2F2F2")
            for sname, spct, samt in est["soft_items"]:
                ws.cell(row=row, column=1, value=f"  {sname}")
                ws.cell(row=row, column=2, value=f"{spct:.1%}")
                ws.cell(row=row, column=3, value=f"${samt:,.0f}")
                for c in range(1, 4):
                    ws.cell(row=row, column=c).fill = soft_fill
                row += 1

            ws.cell(row=row, column=1, value="Soft Cost Subtotal").font = Font(bold=True)
            ws.cell(row=row, column=3, value=f"${est['total_soft']:,.0f}").font = Font(bold=True)
            row += 1

            total_fill = PatternFill("solid", fgColor="E2EFDA")
            ws.cell(row=row, column=1, value="TOTAL ESTIMATED COST").font = Font(bold=True, size=11)
            ws.cell(row=row, column=1).fill = total_fill
            ws.cell(row=row, column=2, value=f"${est['total_psf']:,.2f} / SF").font = Font(bold=True)
            ws.cell(row=row, column=2).fill = total_fill
            ws.cell(row=row, column=3, value=f"${est['grand_total']:,.0f}").font = Font(bold=True, size=11)
            ws.cell(row=row, column=3).fill = total_fill

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = max(max_len + 4, 14)

            # Save
            today = date.today().strftime("%b-%d-%y")
            safe_city = (est["city"] or "estimate").replace(" ", "_").replace(",", "")
            fname = f"cost_estimate_{safe_city}_{today}.xlsx"
            path = os.path.join(OUTPUT_DIR, fname)
            wb.save(path)

            self._set_status(f"Exported: {fname}", SUCCESS)
            os.startfile(path)

        except Exception as exc:
            messagebox.showerror("Export Error", f"Failed to export:\n{exc}")

    # ── Accurate (AI-powered) cost estimate ──────────────────────────────────
    def _accurate_estimate(self):
        # Validate inputs
        try:
            sf = float(self.cost_sf_var.get().replace(",", ""))
            if sf <= 0:
                raise ValueError
        except ValueError:
            messagebox.showwarning("Invalid SF", "Please enter a valid building square footage.")
            return

        city_text = self.cost_city_var.get().strip()
        CITY_HINT = "e.g. Austin, Denver, Miami"
        if not city_text or city_text == CITY_HINT:
            messagebox.showwarning("City Required",
                                   "Please enter a city for the accurate estimate.")
            return

        if self._cost_running:
            messagebox.showinfo("Already Running", "An accurate estimate is already in progress.")
            return

        quality = self.cost_quality_var.get()
        btype = self.cost_type_var.get()
        btype_label = "Climate Controlled" if btype == "cc" else "Drive-Up"

        # Clear results and show progress
        for w in self._results_inner.winfo_children():
            w.destroy()

        self._cost_log_widget = scrolledtext.ScrolledText(
            self._results_inner, font=("Consolas", 9),
            bg=LOG_BG, fg=LOG_FG, insertbackground=LOG_FG,
            relief="flat", bd=0, wrap="word", state="disabled",
            highlightthickness=0, width=80, height=20,
        )
        self._cost_log_widget.pack(fill="both", expand=True, padx=4, pady=4)
        self._cost_log_widget.tag_config("info",    foreground=LOG_FG)
        self._cost_log_widget.tag_config("header",  foreground=ACCENT, font=("Consolas", 9, "bold"))
        self._cost_log_widget.tag_config("success", foreground=SUCCESS)
        self._cost_log_widget.tag_config("error",   foreground=DANGER)

        self._cost_running = True
        self._accurate_btn.config(state="disabled", bg=DIM, fg=MUTED)
        self._quick_btn.config(state="disabled")
        self._cost_export_btn.config(state="disabled")
        self._set_status(f"Researching construction costs for {city_text}...", ACCENT)
        self._pulse_dot()

        # Build output path
        today = date.today().strftime("%b-%d-%y")
        safe_city = city_text.replace(" ", "_").replace(",", "")
        self._cost_output_file = os.path.join(
            OUTPUT_DIR, f"cost_estimate_{safe_city}_{today}.xlsx"
        )

        self._cost_log("=" * 52, "header")
        self._cost_log("  ACCURATE COST ESTIMATE", "header")
        self._cost_log("=" * 52, "header")
        self._cost_log(f"  Type     : {btype_label}", "info")
        self._cost_log(f"  SF       : {sf:,.0f}", "info")
        self._cost_log(f"  City     : {city_text}", "info")
        self._cost_log(f"  Quality  : {quality}", "info")
        self._cost_log("=" * 52, "header")
        self._cost_log("")

        self._cost_thread = threading.Thread(
            target=self._run_cost_agent,
            args=(sf, city_text, quality, btype_label),
            daemon=True,
        )
        self._cost_thread.start()

    def _cost_log(self, text: str, tag: str = "info"):
        def _append():
            self._cost_log_widget.config(state="normal")
            self._cost_log_widget.insert("end", text + "\n", tag)
            self._cost_log_widget.see("end")
            self._cost_log_widget.config(state="disabled")
        self.after(0, _append)

    def _run_cost_agent(self, sf: float, city: str, quality: str, btype: str):
        try:
            asyncio.run(self._cost_agent_query(sf, city, quality, btype))
        except Exception as exc:
            self._cost_log(f"\nError: {exc}", "error")
            self._set_status(f"Error: {exc}", DANGER)
        finally:
            self.after(0, self._on_cost_done)

    async def _cost_agent_query(self, sf: float, city: str, quality: str, btype: str):
        prompt = f"""
Research and create a construction cost estimate for:
  Building Type : {btype} self-storage
  Total SF      : {sf:,.0f}
  City          : {city}
  Quality       : {quality}
  Date          : {date.today().strftime("%B %d, %Y")}
  Save to       : {self._cost_output_file}

Search for CURRENT construction costs specific to {city} or the nearest major metro.
Find real $/SF data, not generic national averages.
Include itemized soft costs (~22.5% of hard costs total): A&E 5%, Permits & Impact Fees 2.5%, Geotech/Environmental 0.8%, Survey & Land Planning 0.4%, Legal & Closing 0.8%, Builder's Risk Insurance 0.7%, Construction Loan Interest 4%, Property Taxes During Construction 0.8%, Contingency 7.5%.
Write the Excel file per the system prompt format.
"""
        async for message in query(
            prompt=prompt,
            options=ClaudeAgentOptions(
                system_prompt=COST_AGENT_PROMPT,
                allowed_tools=["WebSearch", "WebFetch", "Bash", "Write"],
                permission_mode="acceptEdits",
                cwd=BASE_DIR,
                max_turns=10,
                model="claude-haiku-4-5",
            ),
        ):
            if not self._cost_running:
                break

            if isinstance(message, AssistantMessage):
                for block in message.content:
                    if isinstance(block, TextBlock) and block.text.strip():
                        self._cost_log(block.text)
            elif isinstance(message, ResultMessage):
                self._cost_log("")
                self._cost_log("=" * 52, "success")
                self._cost_log("  COMPLETE", "success")
                self._cost_log("=" * 52, "success")
                if message.result:
                    self._cost_log(message.result, "success")

    def _on_cost_done(self):
        self._cost_running = False
        self._running = False  # stop dot pulse
        self._accurate_btn.config(state="normal", bg=ACCENT, fg="#ffffff")
        self._quick_btn.config(state="normal")

        if hasattr(self, '_cost_output_file') and os.path.exists(self._cost_output_file):
            self._set_status(
                f"Done! Saved: {os.path.basename(self._cost_output_file)}", SUCCESS
            )
            self._cost_export_btn.config(state="normal")
            # Repurpose export button to open the file
            self._cost_export_btn.config(
                text="Open Excel",
                command=lambda: os.startfile(self._cost_output_file),
            )
            self._cost_log(f"\nFile saved to:\n  {self._cost_output_file}", "success")
        else:
            self._set_status("Finished (no file produced)", MUTED)

    # ── Logging helpers ────────────────────────────────────────────────────────
    def _log(self, text: str, tag: str = "info"):
        def _append():
            self.log.config(state="normal")
            self.log.insert("end", text + "\n", tag)
            self.log.see("end")
            self.log.config(state="disabled")
        self.after(0, _append)
        if text.strip():
            self._advance_progress(text)

    def _set_status(self, text: str, color: str = MUTED):
        def _update():
            self.status_var.set(text)
            self.status_lbl.config(fg=color)
            self.status_dot.config(fg=color)
        self.after(0, _update)

    def _pulse_dot(self):
        if not self._running:
            self.status_dot.config(fg=MUTED)
            return
        current = self.status_dot.cget("fg")
        self.status_dot.config(fg=ACCENT if current == MUTED else MUTED)
        self.after(600, self._pulse_dot)

    # ── Progress tracking ─────────────────────────────────────────────────────
    _STEPS = [
        (["searching", "search", "sparefoot", "websearch"], "Searching for facilities...", 15),
        (["found", "facilities", "facility"], "Facilities found", 30),
        (["fetching", "webfetch", "pricing", "rates", "price"], "Collecting pricing data...", 50),
        (["calculating", "distance", "drive time", "haversine"], "Calculating distances...", 65),
        (["writing", "openpyxl", "excel", "xlsx", "spreadsheet"], "Building Excel report...", 80),
        (["saved", "complete", "done", "summary"], "Finalizing...", 95),
    ]

    def _advance_progress(self, log_text: str):
        lower = log_text.lower()
        for keywords, label, pct in self._STEPS:
            if any(kw in lower for kw in keywords):
                if pct > self._step:
                    self._step = pct
                    def _update(l=label, p=pct):
                        self.progress["value"] = p
                        self.step_var.set(l)
                        self.pct_var.set(f"{p}%")
                    self.after(0, _update)
                    break

    # ── Run / Stop ─────────────────────────────────────────────────────────────
    def _start(self):
        location = self.location_var.get().strip()
        HINT = "e.g. 78701  or  Austin, TX  or  123 Main St, Denver CO"
        if not location or location == HINT:
            messagebox.showwarning("Missing Location",
                                   "Please enter a subject property location.")
            return

        try:
            radius = float(self.radius_var.get())
        except ValueError:
            messagebox.showwarning("Invalid Radius",
                                   "Please enter a valid number for the radius.")
            return

        if radius > 7:
            messagebox.showwarning("Radius Too Large",
                                   "Maximum search radius is 7 miles.")
            self.radius_var.set("7")
            return

        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        self.log.config(state="disabled")

        self._output_file = None
        self._step = 0
        self.progress["value"] = 0
        self.step_var.set("Starting...")
        self.pct_var.set("0%")
        self.open_btn.config(state="disabled")
        self.run_btn.config(state="disabled")
        self.stop_btn.config(state="normal", bg=DANGER, fg="#ffffff")
        self._running = True
        self._pulse_dot()

        today = date.today().strftime("%b-%d-%y")
        safe_loc = location.replace(" ", "_").replace(",", "").replace("/", "-")
        self._output_file = os.path.join(OUTPUT_DIR, f"storage_comps_{safe_loc}_{today}.xlsx")

        self._log("=" * 62, "header")
        self._log("  SELF STORAGE MARKET RENT COMPS", "header")
        self._log("=" * 62, "header")
        self._log(f"  Location : {location}", "info")
        self._log(f"  Radius   : {radius} miles", "info")
        self._log(f"  Output   : {self._output_file}", "muted")
        self._log("=" * 62, "header")
        self._log("")

        self._set_status(f"Searching near {location}...", ACCENT)

        self._thread = threading.Thread(
            target=self._run_in_thread,
            args=(location, radius),
            daemon=True,
        )
        self._thread.start()

    def _stop(self):
        self._running = False
        self._log("\n[Stopping \u2014 please wait for the current step to finish]", "error")
        self._set_status("Stopping...", DANGER)

    def _run_in_thread(self, location: str, radius: float):
        try:
            asyncio.run(self._run_agent(location, radius))
        except Exception as exc:
            self._log(f"\nError: {exc}", "error")
            self._set_status(f"Error: {exc}", DANGER)
        finally:
            self.after(0, self._on_done)

    async def _run_agent(self, location: str, radius: float):
        prompt = f"""
Find self-storage market rent comps for:
  Location : {location}
  Radius   : {radius} miles
  Date     : {date.today().strftime("%B %d, %Y")}
  Save to  : {self._output_file}

Instructions:
1. Find ALL self-storage facilities within {radius} miles of {location}.
2. For each facility, search for pricing (use aggregator sites like StorageUnits.com
   and SelfStorage.com \u2014 they return static HTML with actual prices).
3. Collect ALL unit sizes (5x5, 5x10, 10x10, 10x15, 10x20, 10x25, 10x30).
4. Calculate distance/drive time from "{location}" for each facility.
5. Write the Excel file using openpyxl (3-tab format per system prompt).
6. Print a brief summary: facilities found, price ranges by unit size.

No fabricated data \u2014 mark missing as N/A.
"""
        async for message in query(
            prompt=prompt,
            options=ClaudeAgentOptions(
                system_prompt=SYSTEM_PROMPT,
                allowed_tools=["WebSearch", "WebFetch", "Bash", "Write"],
                permission_mode="acceptEdits",
                cwd=BASE_DIR,
                max_turns=50,
                model="claude-haiku-4-5-20251001",
            ),
        ):
            if not self._running:
                break

            if isinstance(message, AssistantMessage):
                for block in message.content:
                    if isinstance(block, TextBlock) and block.text.strip():
                        self._log(block.text)
            elif isinstance(message, ResultMessage):
                self._log("")
                self._log("=" * 62, "success")
                self._log("  COMPLETE", "success")
                self._log("=" * 62, "success")
                if message.result:
                    self._log(message.result, "success")

    def _on_done(self):
        self._running = False
        self.run_btn.config(state="normal")
        self.stop_btn.config(state="disabled", bg=DIM, fg=MUTED)
        self.progress["value"] = 100
        self.pct_var.set("100%")

        if self._output_file and os.path.exists(self._output_file):
            self.step_var.set("Complete!")
            self._set_status(
                f"Done!  File saved: {os.path.basename(self._output_file)}", SUCCESS
            )
            self.open_btn.config(state="normal")
            self.fc_btn.config(state="normal")
            self._log(f"\nFile saved to:\n  {self._output_file}", "success")
        else:
            self.step_var.set("Finished")
            self._set_status("Finished (no file produced)", MUTED)

    def _open_file(self):
        if self._output_file and os.path.exists(self._output_file):
            os.startfile(self._output_file)
        else:
            messagebox.showinfo("File Not Found",
                                "The output file was not found. The run may not have completed.")

    # ── Fact-Check agent ───────────────────────────────────────────────────────
    def _start_fact_check(self):
        if not self._output_file or not os.path.exists(self._output_file):
            messagebox.showwarning("No Comps File",
                                   "Run a market comps analysis first, then fact-check it.")
            return

        if self._fc_running:
            messagebox.showinfo("Already Running", "A fact-check is already in progress.")
            return

        self._fc_running = True
        self.fc_btn.config(state="disabled", bg=DIM, fg=MUTED)
        self.run_btn.config(state="disabled")

        today = date.today().strftime("%b-%d-%y")
        basename = os.path.splitext(os.path.basename(self._output_file))[0]
        self._fc_output_file = os.path.join(OUTPUT_DIR, f"audit_{basename}_{today}.xlsx")

        # Clear comps log and repurpose for fact-check output
        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        self.log.config(state="disabled")

        self._step = 0
        self.progress["value"] = 0
        self.step_var.set("Auditing...")
        self.pct_var.set("0%")
        self._running = True
        self._pulse_dot()

        self._log("=" * 62, "header")
        self._log("  MARKET COMPS FACT-CHECK & AUDIT", "header")
        self._log("=" * 62, "header")
        self._log(f"  File     : {self._output_file}", "info")
        self._log(f"  Output   : {self._fc_output_file}", "muted")
        self._log("=" * 62, "header")
        self._log("")

        self._set_status("Auditing comps report...", ORANGE)

        self._fc_thread = threading.Thread(
            target=self._run_fc_thread,
            daemon=True,
        )
        self._fc_thread.start()

    def _run_fc_thread(self):
        try:
            asyncio.run(self._run_fc_agent())
        except Exception as exc:
            self._log(f"\nError: {exc}", "error")
            self._set_status(f"Error: {exc}", DANGER)
        finally:
            self.after(0, self._on_fc_done)

    async def _run_fc_agent(self):
        prompt = f"""
Audit the self-storage market comps spreadsheet:
  File     : {self._output_file}
  Date     : {date.today().strftime("%B %d, %Y")}
  Save to  : {self._fc_output_file}

Follow the steps in the system prompt:
1. Read all data from the Excel file.
2. Verify prices for every facility (search once per facility, match all unit types).
3. Recalculate and verify all Market Summary math.
4. Write the verification report Excel file.
"""
        async for message in query(
            prompt=prompt,
            options=ClaudeAgentOptions(
                system_prompt=FACT_CHECK_PROMPT,
                allowed_tools=["WebSearch", "WebFetch", "Bash", "Read"],
                permission_mode="acceptEdits",
                cwd=BASE_DIR,
                max_turns=50,
                model="claude-sonnet-4-6",
            ),
        ):
            if not self._fc_running:
                break

            if isinstance(message, AssistantMessage):
                for block in message.content:
                    if isinstance(block, TextBlock) and block.text.strip():
                        self._log(block.text)
            elif isinstance(message, ResultMessage):
                self._log("")
                self._log("=" * 62, "success")
                self._log("  AUDIT COMPLETE", "success")
                self._log("=" * 62, "success")
                if message.result:
                    self._log(message.result, "success")

    def _on_fc_done(self):
        self._fc_running = False
        self._running = False
        self.run_btn.config(state="normal")
        self.fc_btn.config(state="normal", bg=ORANGE, fg="#1a1000")
        self.progress["value"] = 100
        self.pct_var.set("100%")

        if self._fc_output_file and os.path.exists(self._fc_output_file):
            self.step_var.set("Audit Complete!")
            self._set_status(
                f"Audit saved: {os.path.basename(self._fc_output_file)}", SUCCESS
            )
            self.open_btn.config(
                state="normal",
                command=lambda: os.startfile(self._fc_output_file),
            )
            self._log(f"\nAudit saved to:\n  {self._fc_output_file}", "success")
        else:
            self.step_var.set("Audit Finished")
            self._set_status("Audit finished (no file produced)", MUTED)


# ── Entry point ─────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    app = StorageCompsApp()
    app.mainloop()
