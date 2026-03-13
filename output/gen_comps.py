import math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from pathlib import Path

output_dir = Path(__file__).resolve().parent
output_dir.mkdir(parents=True, exist_ok=True)

# Subject property - Rambler Ave area
subject_lat, subject_lon = 34.5850, -118.1150

def haversine(lat1, lon1, lat2, lon2):
    R = 3959
    dlat = math.radians(lat2 - lat1)
    dlon = math.radians(lon2 - lon1)
    a = math.sin(dlat/2)**2 + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2)) * math.sin(dlon/2)**2
    c = 2 * math.asin(math.sqrt(a))
    return R * c

def get_drive_time(dist):
    return round((dist / 25) * 60)

facilities = [
    {'name': 'AV Self Storage', 'addr': '850 E Ave P8, Palmdale, CA 93550', 'phone': '(661) 268-6209', 'web': 'avselfstorage.com', 'lat': 34.5795, 'lon': -118.1195, 'units': [
        {'sz': '5x5', 'sf': 25, 'on': 55, 'in': 55, 'cc': True},
        {'sz': '10x10', 'sf': 100, 'on': 'N/A', 'in': 'N/A', 'cc': True},
        {'sz': '10x20', 'sf': 200, 'on': 139, 'in': 139, 'cc': True}
    ], 'note': 'Climate control available'},
    {'name': 'Nova Storage', 'addr': '3305 E Palmdale Blvd, Palmdale, CA 93550', 'phone': '(661) 266-1200', 'web': 'novastorage.com', 'lat': 34.5720, 'lon': -118.0920, 'units': [
        {'sz': '5x5', 'sf': 25, 'on': 104, 'in': 104, 'cc': True},
        {'sz': '5x7', 'sf': 35, 'on': 157.5, 'in': 157.5, 'cc': True},
        {'sz': '10x10', 'sf': 100, 'on': 'N/A', 'in': 'N/A', 'cc': True},
        {'sz': '10x20', 'sf': 200, 'on': 'N/A', 'in': 'N/A', 'cc': False}
    ], 'note': 'Pricing varies by unit'},
    {'name': 'A-American (10th St E)', 'addr': '37228 10th St E, Palmdale, CA 93550', 'phone': '(661) 443-6552', 'web': 'aamericanselfstorage.com', 'lat': 34.5210, 'lon': -118.1835, 'units': [
        {'sz': '5x5', 'sf': 25, 'on': 'N/A', 'in': 'N/A', 'cc': False},
        {'sz': '10x10', 'sf': 100, 'on': 'N/A', 'in': 'N/A', 'cc': False},
        {'sz': '10x20', 'sf': 200, 'on': 'N/A', 'in': 'N/A', 'cc': False}
    ], 'note': 'Pricing from $15'},
    {'name': 'SecurCare (10th St E)', 'addr': '37909 10th St E, Palmdale, CA 93550', 'phone': '(661) 947-7589', 'web': 'securcareselfstorage.com', 'lat': 34.5135, 'lon': -118.1620, 'units': [
        {'sz': '5x5', 'sf': 25, 'on': 49, 'in': 49, 'cc': True},
        {'sz': '10x10', 'sf': 100, 'on': 115, 'in': 115, 'cc': True},
        {'sz': '10x20', 'sf': 200, 'on': 132, 'in': 132, 'cc': True}
    ], 'note': 'Admin fee: $29'},
    {'name': 'Extra Space (Sierra Hwy)', 'addr': '37352 Sierra Hwy, Palmdale, CA 93550', 'phone': '(661) 575-0100', 'web': 'extraspace.com', 'lat': 34.5200, 'lon': -118.2480, 'units': [
        {'sz': '5x5', 'sf': 25, 'on': 32, 'in': 32, 'cc': False},
        {'sz': '10x10', 'sf': 100, 'on': 126, 'in': 126, 'cc': True},
        {'sz': '10x20', 'sf': 200, 'on': 194, 'in': 194, 'cc': True}
    ], 'note': 'Move-in deals available'},
    {'name': 'Extra Space (30th St E)', 'addr': '38910 30th St E, Palmdale, CA 93550', 'phone': '(661) 434-4651', 'web': 'extraspace.com', 'lat': 34.5040, 'lon': -118.2320, 'units': [
        {'sz': '5x5', 'sf': 25, 'on': 38, 'in': 68, 'cc': False},
        {'sz': '5x10', 'sf': 50, 'on': 62, 'in': 110, 'cc': False},
        {'sz': '10x10', 'sf': 100, 'on': 99, 'in': 178, 'cc': True},
        {'sz': '10x20', 'sf': 200, 'on': 'N/A', 'in': 'N/A', 'cc': True}
    ], 'note': 'Web-only specials'},
    {'name': 'SecurCare (Sierra Hwy)', 'addr': '38963 N Sierra Hwy, Palmdale, CA 93550', 'phone': 'N/A', 'web': 'securcareselfstorage.com', 'lat': 34.5240, 'lon': -118.2680, 'units': [
        {'sz': '5x5', 'sf': 25, 'on': 19, 'in': 19, 'cc': True},
        {'sz': '10x10', 'sf': 100, 'on': 70, 'in': 70, 'cc': False},
        {'sz': '10x20', 'sf': 200, 'on': 170, 'in': 170, 'cc': False}
    ], 'note': 'Climate varies'},
]

results = []
for f in facilities:
    d = haversine(subject_lat, subject_lon, f['lat'], f['lon'])
    if d <= 3.0:
        f['dist'] = round(d, 2)
        f['time'] = get_drive_time(d)
        results.append(f)

results.sort(key=lambda x: x['dist'])
print('Found {} facilities within 3 miles'.format(len(results)))
for r in results:
    print('  {}: {} mi, {} min'.format(r['name'], r['dist'], r['time']))

wb = Workbook()
ws1 = wb.active
ws1.title = 'Comps Detail'
ws2 = wb.create_sheet('Market Summary')
ws3 = wb.create_sheet('Facility List')

# TAB 1
headers = ['Facility Name', 'Address', 'Distance (mi)', 'Unit Type', 'Sq Ft', 'Climate Controlled', 'Online Rate ($/mo)', 'In-Store Rate ($/mo)', 'Notes/Promotions', 'Date Pulled', 'Source URL', 'Drive Time (min)']
ws1.append(headers)

fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
font = Font(bold=True)
for cell in ws1[1]:
    cell.fill = fill
    cell.font = font

for f in results:
    for u in f['units']:
        on_rate = u['on'] if isinstance(u['on'], str) else '${:.2f}'.format(u['on'])
        in_rate = u['in'] if isinstance(u['in'], str) else '${:.2f}'.format(u['in'])
        ws1.append([f['name'], f['addr'], f['dist'], u['sz'], u['sf'], 'Yes' if u['cc'] else 'No', on_rate, in_rate, f['note'], '3/11/2026', f['web'], f['time']])

ws1.column_dimensions['A'].width = 32
ws1.column_dimensions['B'].width = 40
ws1.column_dimensions['C'].width = 14
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 8
ws1.column_dimensions['F'].width = 18
ws1.column_dimensions['G'].width = 18
ws1.column_dimensions['H'].width = 18
ws1.column_dimensions['I'].width = 25
ws1.column_dimensions['J'].width = 12
ws1.column_dimensions['K'].width = 20
ws1.column_dimensions['L'].width = 14
ws1.freeze_panes = 'A2'

# TAB 2
row = 1
cc_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
ws2.cell(row, 1, 'CLIMATE CONTROLLED UNITS').font = Font(bold=True)
ws2.cell(row, 1).fill = cc_fill

row = 2
cc_hdrs = ['Unit Type', 'Sq Ft', 'Avg Online', 'Min Online', 'Max Online', 'Avg In-Store', 'Min In-Store', 'Max In-Store', '# Comps']
for col, hdr in enumerate(cc_hdrs, 1):
    cell = ws2.cell(row, col, hdr)
    cell.fill = cc_fill
    cell.font = Font(bold=True)

cc_data = {}
for f in results:
    for u in f['units']:
        if u['cc']:
            key = (u['sz'], u['sf'])
            if key not in cc_data:
                cc_data[key] = {'on': [], 'in': []}
            if isinstance(u['on'], (int, float)):
                cc_data[key]['on'].append(u['on'])
            if isinstance(u['in'], (int, float)):
                cc_data[key]['in'].append(u['in'])

row = 3
for (sz, sf), d in sorted(cc_data.items()):
    on_vals = d['on']
    in_vals = d['in']
    ws2.cell(row, 1, sz)
    ws2.cell(row, 2, sf)
    if on_vals:
        ws2.cell(row, 3, '${:.2f}'.format(sum(on_vals)/len(on_vals)))
        ws2.cell(row, 4, '${:.2f}'.format(min(on_vals)))
        ws2.cell(row, 5, '${:.2f}'.format(max(on_vals)))
    if in_vals:
        ws2.cell(row, 6, '${:.2f}'.format(sum(in_vals)/len(in_vals)))
        ws2.cell(row, 7, '${:.2f}'.format(min(in_vals)))
        ws2.cell(row, 8, '${:.2f}'.format(max(in_vals)))
    ws2.cell(row, 9, len(on_vals) if on_vals else len(in_vals))
    row += 1

row += 1

du_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
ws2.cell(row, 1, 'DRIVE UP / STANDARD UNITS').font = Font(bold=True)
ws2.cell(row, 1).fill = du_fill

row += 1
for col, hdr in enumerate(cc_hdrs, 1):
    cell = ws2.cell(row, col, hdr)
    cell.fill = du_fill
    cell.font = Font(bold=True)

du_data = {}
for f in results:
    for u in f['units']:
        if not u['cc']:
            key = (u['sz'], u['sf'])
            if key not in du_data:
                du_data[key] = {'on': [], 'in': []}
            if isinstance(u['on'], (int, float)):
                du_data[key]['on'].append(u['on'])
            if isinstance(u['in'], (int, float)):
                du_data[key]['in'].append(u['in'])

row += 1
for (sz, sf), d in sorted(du_data.items()):
    on_vals = d['on']
    in_vals = d['in']
    ws2.cell(row, 1, sz)
    ws2.cell(row, 2, sf)
    if on_vals:
        ws2.cell(row, 3, '${:.2f}'.format(sum(on_vals)/len(on_vals)))
        ws2.cell(row, 4, '${:.2f}'.format(min(on_vals)))
        ws2.cell(row, 5, '${:.2f}'.format(max(on_vals)))
    if in_vals:
        ws2.cell(row, 6, '${:.2f}'.format(sum(in_vals)/len(in_vals)))
        ws2.cell(row, 7, '${:.2f}'.format(min(in_vals)))
        ws2.cell(row, 8, '${:.2f}'.format(max(in_vals)))
    ws2.cell(row, 9, len(on_vals) if on_vals else len(in_vals))
    row += 1

for c in 'ABCDEFGHI':
    ws2.column_dimensions[c].width = 14

# TAB 3
ws3.append(['Facility Name', 'Address', 'Distance (mi)', 'Drive Time (min)', 'Phone', 'Website'])
fac_fill = PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid')
for cell in ws3[1]:
    cell.fill = fac_fill
    cell.font = Font(bold=True)

for f in results:
    ws3.append([f['name'], f['addr'], f['dist'], f['time'], f['phone'], f['web']])

ws3.column_dimensions['A'].width = 32
ws3.column_dimensions['B'].width = 40
ws3.column_dimensions['C'].width = 14
ws3.column_dimensions['D'].width = 16
ws3.column_dimensions['E'].width = 16
ws3.column_dimensions['F'].width = 20
ws3.freeze_panes = 'A2'

out = output_dir / 'storage_comps_Rambler_Ave_Palmdale_CA_93550_Mar-11-26.xlsx'
wb.save(str(out))
print('Excel saved: {}'.format(out))
